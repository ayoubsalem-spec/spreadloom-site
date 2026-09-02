"""
BuildIQ -- Darycet's unified technology platform.

One Flask app, one login, one database. Sections are prefixed by URL path
rather than split into separate Flask Blueprint objects, to keep this first
version simple to read top-to-bottom. Sections:
  /            -- home (app picker)
  /tracker/... -- Bid Tracker (formerly Command Center: projects + quotes)
  /sitepulse/... -- SitePulse (equipment + outside rentals)
  /inventory/... -- Site Inventory (concrete requests + material inventory)

This is the first migration pass. Ported fully: auth, home, SitePulse
(equipment + rentals), Site Inventory (concrete requests + materials), and
Bid Tracker's core (projects, quotes, dashboard). NOT yet ported from the
original Command Center: AI-generated RFQ/follow-up emails, quote file
attachments, and Excel export -- flagged here rather than silently dropped,
to be added in a follow-up pass.
"""
import os
import sqlite3
import uuid
import hashlib
import json
import secrets
import threading
import time
import base64
import io
import requests
import markdown as md_lib
import bleach
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas as pdf_canvas
from flask import Flask, render_template, request, redirect, url_for, flash, g, send_file, send_from_directory, session, Response, stream_with_context
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_wtf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)

# ---------------------------------------------------------------------------
# SECRET_KEY hardening -- production must never silently start with a
# known/default secret. APP_ENV defaults to "production" (fail-closed) so
# a deployment that forgets to set it is treated as production, not as an
# accidental opt-in to the insecure dev fallback. Only an explicit
# APP_ENV=development (or dev/test/testing, for CI/local test runs)
# unlocks the dev-only fallback key below.
# ---------------------------------------------------------------------------
APP_ENV = os.environ.get("APP_ENV", "production").strip().lower()
_DEV_ENVS = ("development", "dev", "test", "testing")

# Dedicated, explicit flag for the temporary Atlas Voice Diagnostics panel
# (templates/assistant.html) -- deliberately NOT derived from APP_ENV/
# _DEV_ENVS. Railway's TEST deployment intentionally runs with
# APP_ENV=production (to exercise production-like security behavior and
# require a real SECRET_KEY), so APP_ENV cannot be used to gate this.
# Defaults to false/off; must be explicitly set to enable. LIVE will not
# set this. Does not affect SECRET_KEY handling, APP_ENV, or any other
# security configuration.
ATLAS_VOICE_DIAGNOSTICS = os.environ.get("ATLAS_VOICE_DIAGNOSTICS", "").strip().lower() in ("1", "true", "yes")
_DEV_ONLY_SECRET_KEY = "dev-only-insecure-secret-do-not-use-in-production"

_secret_key = os.environ.get("SECRET_KEY", "").strip()
if not _secret_key:
    if APP_ENV in _DEV_ENVS:
        _secret_key = _DEV_ONLY_SECRET_KEY
    else:
        raise RuntimeError(
            "SECRET_KEY is not set. Refusing to start with APP_ENV="
            f"'{APP_ENV}' and no secret key configured. Set the SECRET_KEY "
            "environment variable, or set APP_ENV=development for local "
            "development/testing only (never in production)."
        )
elif _secret_key == _DEV_ONLY_SECRET_KEY and APP_ENV not in _DEV_ENVS:
    # Someone explicitly set SECRET_KEY to the known dev value outside a
    # dev/test environment -- treat that the same as "no secret set".
    raise RuntimeError(
        "SECRET_KEY is set to the known development-only value while "
        f"APP_ENV='{APP_ENV}'. Refusing to start. Set a real SECRET_KEY."
    )
app.secret_key = _secret_key
del _secret_key  # never leave the resolved value sitting in a module-level name
csrf = CSRFProtect(app)

# Atlas conversation state lives here, server-side, keyed by a small token
# stored in the person's session cookie. Streamed responses can't safely
# rewrite the session cookie mid-stream (headers are already sent by the
# time the body starts flowing), so the actual draft -- mode, collected
# fields, and real message history -- is kept here instead, and the cookie
# only ever holds the lookup token. In-memory, so it resets on redeploy;
# that's fine for a conversational scratchpad.
ATLAS_SESSIONS = {}

# Guards the "claim" step of a pending concrete-request write confirmation
# (see assistant_confirm_write) -- ATLAS_SESSIONS is a plain in-memory
# dict with no other concurrency protection, so without this, two
# simultaneous or retried requests carrying the same valid token could
# both read pending_write as still-present before either one clears it,
# and both go on to call execute_tool -- a real double-write. The lock
# only needs to protect the short check-token-and-clear step; the
# potentially slow execute_tool() call itself deliberately runs outside
# it (see assistant_confirm_write's docstring for the full reasoning).
# A single process-wide lock is appropriate here, not a per-token lock:
# ATLAS_SESSIONS is already documented as in-memory/single-process only
# (won't survive a restart, not safe for multi-process deployment) --
# this lock matches that same architectural scope, not a new constraint.
ATLAS_WRITE_CONFIRM_LOCK = threading.Lock()

# How long a pending_write token remains valid after being issued. Keeps
# an abandoned write authorization (client never called confirm_write --
# barge-in, tab closed, network dropped) from sitting valid in
# ATLAS_SESSIONS indefinitely. Long enough for the client to finish
# receiving the tail of one SSE response and immediately call
# confirm_write; short enough that a genuinely abandoned token doesn't
# linger as a live authorization.
PENDING_WRITE_TTL_SECONDS = 120


HOUSTON_TZ = ZoneInfo("America/Chicago")

# --- WhatsApp group notifications (Green API) ------------------------------
# Set these three in your environment (Railway variables, etc.) once you have
# a Green API instance linked to a WhatsApp number that's in the procurement
# group. Left blank, notifications are silently skipped (logged to console)
# so nothing breaks if they're not configured yet.
# --- WhatsApp group notifications (Ultramsg) --------------------------------
# Set these three (four, counting the second group) in your environment
# (Railway variables, etc.) once you have an Ultramsg instance linked to a
# WhatsApp number that's in both the procurement and SitePulse groups. Left
# blank, notifications are silently skipped (logged to console) so nothing
# breaks if they're not configured yet.
ULTRAMSG_INSTANCE_ID = os.environ.get("ULTRAMSG_INSTANCE_ID", "")
ULTRAMSG_TOKEN = os.environ.get("ULTRAMSG_TOKEN", "")
ULTRAMSG_GROUP_CHAT_ID = os.environ.get("ULTRAMSG_GROUP_CHAT_ID", "")  # e.g. "123456789-987654321@g.us" -- procurement
ULTRAMSG_SITEPULSE_GROUP_CHAT_ID = os.environ.get("ULTRAMSG_SITEPULSE_GROUP_CHAT_ID", "")  # equipment/SitePulse group


def send_whatsapp_group_message(text, chat_id=None):
    """Post a message into a WhatsApp group via Ultramsg. Defaults to the
    procurement group if one's configured, otherwise falls back to the
    SitePulse group -- so if only one group is set up (current setup: just
    SitePulse), everything lands there. Never raises -- a WhatsApp hiccup
    should never block someone submitting a request or moving equipment.
    Failures are printed to the server log instead. Returns (ok, detail) so
    callers that want to report success/failure (e.g. the test-message
    button) can, without every other call site needing to check it.
    """
    chat_id = chat_id or ULTRAMSG_GROUP_CHAT_ID or ULTRAMSG_SITEPULSE_GROUP_CHAT_ID
    if not (ULTRAMSG_INSTANCE_ID and ULTRAMSG_TOKEN and chat_id):
        msg = "Ultramsg not configured -- skipping notification:\n" + text
        print("[whatsapp] " + msg)
        return False, "WhatsApp isn't configured yet (missing instance ID, token, or chat ID)."
    url = f"https://api.ultramsg.com/{ULTRAMSG_INSTANCE_ID}/messages/chat"
    try:
        resp = requests.post(url, data={"token": ULTRAMSG_TOKEN, "to": chat_id, "body": text}, timeout=10)
        if resp.status_code >= 300:
            print(f"[whatsapp] Ultramsg returned {resp.status_code}: {resp.text}")
            return False, f"Ultramsg returned an error ({resp.status_code})."
        return True, "Sent."
    except requests.RequestException as e:
        print(f"[whatsapp] failed to send notification: {e}")
        return False, f"Network error reaching Ultramsg: {e}"


def whatsapp_chat_id_for_site(*texts):
    """Match project/job/location text against the configured per-site
    WhatsApp groups (keyword is a case-insensitive substring match against
    any of the given texts, e.g. project="Peninsula Job #4" matches
    keyword="peninsula"). Falls back to the default SitePulse group if
    nothing matches or no site groups are configured yet.
    """
    db = get_db()
    rows = db.execute("SELECT keyword, chat_id FROM whatsapp_site_groups").fetchall()
    haystack = " ".join(t for t in texts if t).lower()
    for row in rows:
        if row["keyword"].lower() in haystack:
            return row["chat_id"]
    return None  # let send_whatsapp_group_message fall back to the default


def send_whatsapp_document(pdf_bytes, filename, chat_id=None, caption=None):
    """Post a PDF (or any small file) into a WhatsApp group via Ultramsg,
    sent as base64 directly in the request -- no public URL/hosting
    needed. Same never-raises, (ok, detail) contract as
    send_whatsapp_group_message. Ultramsg's base64 limit is ~6.5MB of
    encoded text, plenty for a one-page order summary.
    """
    chat_id = chat_id or ULTRAMSG_GROUP_CHAT_ID or ULTRAMSG_SITEPULSE_GROUP_CHAT_ID
    if not (ULTRAMSG_INSTANCE_ID and ULTRAMSG_TOKEN and chat_id):
        print(f"[whatsapp] Ultramsg not configured -- skipping document send: {filename}")
        return False, "WhatsApp isn't configured yet (missing instance ID, token, or chat ID)."
    url = f"https://api.ultramsg.com/{ULTRAMSG_INSTANCE_ID}/messages/document"
    b64 = base64.b64encode(pdf_bytes).decode("ascii")
    data = {"token": ULTRAMSG_TOKEN, "to": chat_id, "document": b64, "filename": filename}
    if caption:
        data["caption"] = caption
    try:
        resp = requests.post(url, data=data, timeout=20)
        if resp.status_code >= 300:
            print(f"[whatsapp] Ultramsg document send returned {resp.status_code}: {resp.text}")
            return False, f"Ultramsg returned an error ({resp.status_code})."
        return True, "Sent."
    except requests.RequestException as e:
        print(f"[whatsapp] failed to send document: {e}")
        return False, f"Network error reaching Ultramsg: {e}"


def _pdf_write_wrapped(c, text, x, y, max_width, font="Helvetica", size=10, leading=14):
    """Write text to a reportlab canvas, wrapping at max_width. Returns the
    y position after the last line, so callers can keep stacking sections."""
    from reportlab.pdfbase.pdfmetrics import stringWidth
    c.setFont(font, size)
    words = text.split(" ")
    line = ""
    for word in words:
        trial = (line + " " + word).strip()
        if stringWidth(trial, font, size) > max_width and line:
            c.drawString(x, y, line)
            y -= leading
            line = word
        else:
            line = trial
    if line:
        c.drawString(x, y, line)
        y -= leading
    return y


def build_concrete_order_pdf(r):
    """One-page PDF summary of a placed concrete order -- project, pour
    details, and every vendor/contact, for attaching to the WhatsApp
    notification and for anyone who wants a printable copy.
    """
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=letter)
    width, height = letter
    navy = colors.HexColor("#0B1220")
    gold = colors.HexColor("#D4A537")
    x = 0.75 * inch
    y = height - 0.9 * inch

    c.setFillColor(navy)
    c.rect(0, height - 1.1 * inch, width, 1.1 * inch, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(x, height - 0.65 * inch, "Concrete Order Confirmation")
    c.setFont("Helvetica", 10)
    c.drawString(x, height - 0.9 * inch, f"Darycet International  |  Order placed {date.today().isoformat()}")

    y = height - 1.5 * inch
    c.setFillColor(navy)

    def section(title):
        nonlocal y
        y -= 6
        c.setFillColor(gold)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x, y, title)
        c.setFillColor(navy)
        y -= 18

    def line(label, value):
        nonlocal y
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x, y, f"{label}:")
        c.setFont("Helvetica", 10)
        c.drawString(x + 1.6 * inch, y, str(value) if value else "\u2014")
        y -= 16

    section("Project")
    line("Project", r["project"])
    line("Job Site Address", r["job_site_address"])
    line("Pour Date", r["pour_date"] + (f" at {r['pour_time']}" if r["pour_time"] else "") if r["pour_date"] else "")
    line("Amount / Mix", " ".join(v for v in [r["concrete_amount"], f"{r['mix_design_psi']} PSI" if r["mix_design_psi"] else ""] if v))

    section("Concrete")
    line("Company", r["concrete_company"])
    line("Phone", r["concrete_company_phone"])
    line("Arrival Time", r["concrete_arrival_time"] or r["pour_time"])

    if r["pump_company"] or r["pump_size"]:
        section(r["pump_type"] if r["pump_type"] else "Pump")
        line("Type", r["pump_type"])
        line("Size", r["pump_size"])
        line("Contact", r["pump_company"])
        line("Phone", r["pump_company_phone"])
        line("Arrival Time", r["pump_arrival_time"])

    if r["lab_required"] == "Yes":
        section("Lab")
        line("Company", r["lab_company"])
        line("Time", r["lab_time"])

    if r["drilling_required"] == "Yes":
        section("Drilling")
        line("Company", r["drilling_company"])
        line("Phone", r["drilling_company_phone"])
        line("Time", r["drilling_time"])

    section("Ordered By")
    line("Name", r["ordered_by"])
    line("Date", r["ordered_date"])

    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.HexColor("#888888"))
    c.drawString(x, 0.6 * inch, "Generated automatically by BuildIQ / SitePulse")
    c.save()
    buf.seek(0)
    return buf.read()


def build_purchase_order_pdf(r, items):
    """One-page PDF summary of a placed purchase order."""
    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=letter)
    width, height = letter
    navy = colors.HexColor("#0B1220")
    gold = colors.HexColor("#D4A537")
    x = 0.75 * inch

    c.setFillColor(navy)
    c.rect(0, height - 1.1 * inch, width, 1.1 * inch, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 18)
    c.drawString(x, height - 0.65 * inch, "Purchase Order Confirmation")
    c.setFont("Helvetica", 10)
    c.drawString(x, height - 0.9 * inch, f"Darycet International  |  Order placed {friendly_date(date.today().isoformat())}")

    y = height - 1.5 * inch
    c.setFillColor(navy)

    def section(title):
        nonlocal y
        y -= 6
        c.setFillColor(gold)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x, y, title)
        c.setFillColor(navy)
        y -= 18

    def line(label, value):
        nonlocal y
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x, y, f"{label}:")
        c.setFont("Helvetica", 10)
        c.drawString(x + 1.6 * inch, y, str(value) if value else "\u2014")
        y -= 16

    section("Job")
    line("Job Name", r["job_name"])
    line("Location", r["location_description"])
    line("PR Number", r["pr_number"])
    line("Needed By", friendly_date(r["needed_on"]))

    section("Vendor")
    line("Company", r["vendor_company"])
    line("Phone", r["vendor_company_phone"])

    if items:
        section("Items")
        for it in items:
            desc = " \u2014 ".join(v for v in [it["item"], it["description"]] if v)
            qty = f" ({it['qty']}{' ' + it['unit'] if it['unit'] else ''})" if it["qty"] else ""
            y = _pdf_write_wrapped(c, f"\u2022 {desc}{qty}", x, y, width - 1.5 * inch, size=10)

    section("Ordered By")
    line("Name", r["ordered_by"])
    line("Date", friendly_date(r["ordered_date"]))

    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(colors.HexColor("#888888"))
    c.drawString(x, 0.6 * inch, "Generated automatically by BuildIQ / SitePulse")
    c.save()
    buf.seek(0)
    return buf.read()


@app.template_filter("friendly_dt")
def friendly_dt(iso_str):
    """'2026-08-18T19:36:00' (stored UTC) -> 'August 18, 2026 at 2:36 PM'
    (converted to Houston/Central time, DST-aware)."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str).replace(tzinfo=ZoneInfo("UTC")).astimezone(HOUSTON_TZ)
        return dt.strftime("%B %-d, %Y at %-I:%M %p")
    except ValueError:
        return iso_str


@app.template_filter("friendly_date")
def friendly_date(date_str):
    """'2026-08-28' -> '08/28/2026'. For plain date fields (no time
    component) -- pour dates, requested dates, ordered dates, etc.
    Leaves anything that doesn't parse as a bare YYYY-MM-DD unchanged,
    rather than erroring, since some callers may pass already-formatted
    or blank values."""
    if not date_str:
        return ""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%m/%d/%Y")
    except ValueError:
        return date_str


@app.template_filter("friendly_short_dt")
def friendly_short_dt(iso_str):
    """'2026-08-19T17:46:00' (stored UTC) -> '08/19/2026 5:46 PM' (Houston
    time). Same MM/DD/YYYY convention as friendly_date, just for the
    "Submitted ..." timestamp headers, which also carry a time -- kept
    separate from friendly_dt (which spells out the month name) so
    existing month-name displays elsewhere aren't affected."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str).replace(tzinfo=ZoneInfo("UTC")).astimezone(HOUSTON_TZ)
        return dt.strftime("%m/%d/%Y %-I:%M %p")
    except ValueError:
        return iso_str


def friendly_time(hhmm):
    """'07:00' -> '7:00 AM'. Used for WhatsApp notifications (not a
    Jinja filter -- those render in a template, this builds plain text
    strings), so every notification's time is formatted consistently
    with what's shown in the app itself."""
    if not hhmm:
        return ""
    try:
        return datetime.strptime(hhmm, "%H:%M").strftime("%-I:%M %p")
    except ValueError:
        return hhmm


DB_DIR = os.environ.get("DATA_DIR", ".")
DB_PATH = os.path.join(DB_DIR, "buildiq.db")
UPLOAD_DIR = os.path.join(DB_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
ALLOWED_PHOTO_EXTENSIONS = {"png", "jpg", "jpeg", "heic", "webp"}
MAX_PHOTO_SIZE_MB = 10

ADMIN_EMAILS = ["ayoub@darycet.com", "rebecca@darycet.com"]

# ---------------------------------------------------------------------------
# TEMPORARY TEST ADMIN -- remove when testing is done.
# To remove: delete this block, delete the user row (see
# scripts/manage_temp_admin.py's --remove command), and remove this email
# from the nav dropdown checks in templates/base.html.
# ---------------------------------------------------------------------------
TEMP_TEST_ADMIN_EMAIL = "test-admin@darycet.com"
ADMIN_EMAILS.append(TEMP_TEST_ADMIN_EMAIL)

# Domains allowed to sign up. Being on this list only grants basic access
# (Equipment Center / SitePulse) -- Project Hunt and admin tooling stay
# gated per-email below, never per-domain.
ALLOWED_SIGNUP_DOMAINS = ["@darycet.com", "@nomaengineering.com"]
# Extra individual emails allowed to sign up even though they're outside
# every allowed domain above.
EXTRA_ALLOWED_SIGNUP_EMAILS = set()
# Full access to every section, including Project Hunt -- named
# individuals only, never a whole domain.
FULL_ACCESS_EMAILS = {"ayoub@darycet.com", "rebecca@darycet.com", "marilu@darycet.com", "hghuneim@nomaengineering.com", TEMP_TEST_ADMIN_EMAIL}
# Atlas (voice assistant) access -- separate from Project Hunt so someone
# can get Atlas without also getting Project Hunt. Everyone in
# FULL_ACCESS_EMAILS gets it too, plus anyone listed here individually.
ATLAS_ACCESS_EMAILS = FULL_ACCESS_EMAILS | {"rebecca@nomaengineering.com"}
# Only these can actually place a concrete/material order -- everyone else
# can submit a request, but "Scheduled/Ordered" plus the vendor/contact
# details is procurement's call.
PROCUREMENT_EMAILS = {"ayoub@darycet.com", "rebecca@darycet.com", "marilu@darycet.com", TEMP_TEST_ADMIN_EMAIL}
# Who can manage the WhatsApp site-group routing -- narrower than full
# admin, but wider than just Ayoub.
WHATSAPP_ADMIN_EMAILS = {"ayoub@darycet.com", "rebecca@darycet.com", TEMP_TEST_ADMIN_EMAIL}


def is_project_hunt_allowed():
    # RUNTIME AUTHORIZATION -- resolves purely through user_has_permission(),
    # which itself implements explicit-deny > explicit-grant > role > false.
    # FULL_ACCESS_EMAILS is never consulted here; it is read-only migration/
    # backfill data (see _backfill_user_roles()). A legacy-listed person's
    # actual access comes from the role that was backfilled onto their
    # account -- if that grant is ever explicitly denied in the Permissions
    # Center, this function must return False even though their email is
    # still in the legacy list.
    if not current_user.is_authenticated:
        return False
    return user_has_permission(current_user, "module:project_hunt:view")


def is_atlas_allowed():
    # RUNTIME AUTHORIZATION -- same as is_project_hunt_allowed() above:
    # user_has_permission() alone, no legacy-list fallback.
    if not current_user.is_authenticated:
        return False
    return user_has_permission(current_user, "module:atlas:view")


def is_procurement():
    # RUNTIME AUTHORIZATION -- same as is_project_hunt_allowed() above.
    if not current_user.is_authenticated:
        return False
    return user_has_permission(current_user, "action:sitepulse:place_order")


def is_whatsapp_admin():
    # RUNTIME AUTHORIZATION -- same as is_project_hunt_allowed() above.
    if not current_user.is_authenticated:
        return False
    return user_has_permission(current_user, "action:team_admin:manage_whatsapp")


def is_product_request_approver():
    # RUNTIME AUTHORIZATION -- same as is_project_hunt_allowed() above.
    # Item 3: whoever holds action:product_intelligence:approve_requests
    # (via role or direct grant -- see ROLE_DEFAULT_PERMISSIONS above) is
    # the procurement approver. Nothing here names a specific person; the
    # authority moves if the permission grant moves.
    if not current_user.is_authenticated:
        return False
    return user_has_permission(current_user, "action:product_intelligence:approve_requests")

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        # Concurrency fix (Product Intelligence approval-gate atomicity
        # review): with no busy_timeout, two genuinely simultaneous
        # writers on two separate connections/threads can hit SQLite's
        # own "database is locked" error immediately rather than one of
        # them briefly waiting -- turning an ordinary race into an
        # unhandled exception instead of a clean one-winner outcome.
        # This makes a losing writer wait briefly for the winner's
        # short transaction to finish, then proceed normally (and, for
        # the approval transition below, correctly see rowcount==0
        # because the winner already committed) instead of erroring.
        # Connection-level setting only -- no schema change.
        g.db.execute("PRAGMA busy_timeout = 5000")
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def is_admin():
    """MIGRATION/BACKFILL ONLY -- not called anywhere in the runtime
    authorization path. Reads the legacy list directly; used only by
    _backfill_user_roles() (to decide which role to grant) and
    _is_protected_admin_account() (a data-protection rule, not an
    access grant -- see its docstring). If you're adding a new route
    check, use _authorized()/user_has_permission() instead, never this."""
    return current_user.is_authenticated and current_user.email in ADMIN_EMAILS


def _authorized(permission_key):
    """RUNTIME AUTHORIZATION for every migrated route below. Resolves
    purely through user_has_permission() -- explicit deny > explicit
    grant > inherited role > false. Does NOT consult is_admin() or any
    other legacy list. A legacy admin's access exists only because
    _backfill_user_roles() granted them the Administrator role (or
    equivalent); if that grant is ever explicitly denied for a specific
    permission in the Permissions Center, this returns False for that
    permission even for someone whose email is still in ADMIN_EMAILS."""
    return user_has_permission(current_user, permission_key)



def _is_protected_admin_account(user_row):
    """A target account is protected from deletion if it's in the legacy
    ADMIN_EMAILS list OR holds the Administrator role in the new system --
    covers an admin created purely through the new system too, not just
    the hardcoded list."""
    if user_row["email"] in ADMIN_EMAILS:
        return True
    db = get_db()
    row = db.execute(
        "SELECT 1 FROM user_roles ur JOIN roles r ON r.id = ur.role_id "
        "WHERE ur.user_id = ? AND r.name = 'Administrator' LIMIT 1",
        (user_row["id"],)
    ).fetchone()
    return row is not None


@app.context_processor
def inject_permissions():
    return {
        "has_project_hunt_access": is_project_hunt_allowed(),
        "has_atlas_access": is_atlas_allowed(),
        "is_procurement": is_procurement(),
        # is_admin_user is now purely permission-based (module:team_admin:view),
        # not the legacy is_admin() email check -- kept for template
        # backward-compatibility (nothing currently reads it after the nav
        # cleanup, but it's cheap to keep correct rather than delete).
        "is_admin_user": _authorized("module:team_admin:view"),
        # Real nav-visibility flags, each mirroring the exact permission
        # its route enforces server-side -- pure user_has_permission(),
        # no legacy-list fallback. Backend remains the source of truth;
        # these only decide what's shown.
        "has_product_intelligence_access": _authorized("module:product_intelligence:view"),
        "has_team_admin_access": _authorized("module:team_admin:view"),
        "has_whatsapp_admin_access": is_whatsapp_admin(),
        "has_system_data_access": _authorized("action:system_data:manage"),
        "has_activity_log_access": _authorized("action:activity_log:view"),
        "has_manage_users_access": _authorized("action:team_admin:manage_users"),
        "has_manage_inventory_access": _authorized("action:sitepulse:manage_inventory"),
        "is_product_request_approver": is_product_request_approver(),
        # Explicit, dedicated flag for the temporary Atlas Voice
        # Diagnostics panel (templates/assistant.html) -- see
        # ATLAS_VOICE_DIAGNOSTICS above. Independent of APP_ENV; defaults
        # false; must be explicitly enabled via env var.
        "atlas_voice_diagnostics_enabled": ATLAS_VOICE_DIAGNOSTICS,
    }


@app.before_request
def restrict_project_hunt():
    """Project Hunt (Bid Tracker) is limited to a specific list of people --
    everyone else gets Equipment Center and SitePulse only. Checked once
    here for every /tracker/* request rather than per-route, so a new route
    added later can't accidentally skip this."""
    if request.path.startswith("/tracker") and current_user.is_authenticated:
        if not is_project_hunt_allowed():
            flash("You don't have access to Project Hunt.", "error")
            return redirect(url_for("home"))


def save_photo(file_storage):
    if not file_storage or not file_storage.filename:
        return None
    ext = file_storage.filename.rsplit(".", 1)[-1].lower() if "." in file_storage.filename else ""
    if ext not in ALLOWED_PHOTO_EXTENSIONS:
        flash(f"Photo not saved -- unsupported file type ({ext or 'unknown'}). Use JPG, PNG, HEIC, or WEBP.", "error")
        return None
    file_storage.seek(0, os.SEEK_END)
    size_mb = file_storage.tell() / (1024 * 1024)
    file_storage.seek(0)
    if size_mb > MAX_PHOTO_SIZE_MB:
        flash(f"Photo not saved -- file too large ({size_mb:.1f}MB, max {MAX_PHOTO_SIZE_MB}MB).", "error")
        return None
    filename = f"{uuid.uuid4().hex}.{ext}"
    file_storage.save(os.path.join(UPLOAD_DIR, secure_filename(filename)))
    return filename


@app.route("/uploads/<filename>")
@login_required
def uploaded_photo(filename):
    return send_from_directory(UPLOAD_DIR, secure_filename(filename))


def log_activity(section, entity_type, entity_id, action, asset_id=None, field=None, old_value=None, new_value=None):
    db = get_db()
    user_email = current_user.email if current_user.is_authenticated else "system"
    db.execute(
        """INSERT INTO activity_log (section, asset_id, entity_type, entity_id, action, field,
           old_value, new_value, user_email, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (section, asset_id, entity_type, entity_id, action, field,
         str(old_value) if old_value is not None else None,
         str(new_value) if new_value is not None else None,
         user_email, datetime.utcnow().isoformat())
    )


def _apply_move(db, move, moved_by, is_auto=False):
    """Shared logic for completing a movement -- whether it's an immediate
    move being recorded right now, or a previously Scheduled one becoming
    active. Snapshots the asset's CURRENT status/hours onto the history
    row (never the value from whenever it was scheduled), moves the asset,
    and stamps who/when it actually happened.
    """
    asset = db.execute("SELECT name, status, hours_mileage FROM sitepulse_assets WHERE id = ?", (move["asset_id"],)).fetchone()
    now = datetime.utcnow().isoformat()
    mover_label = f"{moved_by} (auto, scheduled by {move['created_by']})" if is_auto and move["created_by"] else moved_by
    db.execute("UPDATE sitepulse_assets SET location=?, updated_at=? WHERE id=?",
               (move["to_location"], now, move["asset_id"]))
    db.execute(
        """UPDATE sitepulse_usage_log SET move_status='Applied', status_at_move=?, mileage_hours=?,
           moved_by=?, applied_at=?, out_date=? WHERE id=?""",
        (asset["status"], asset["hours_mileage"], mover_label, now, date.today().isoformat(), move["id"])
    )
    log_activity("sitepulse", "move", move["id"], "applied", asset_id=move["asset_id"],
                 field="location", old_value=move["from_location"], new_value=move["to_location"])
    send_whatsapp_group_message(
        f"📍 {asset['name']} moved{' (scheduled)' if move['scheduled_date'] else ''}\n"
        f"{move['from_location'] or '—'} → {move['to_location']}\n"
        f"Hours/Mileage: {asset['hours_mileage'] or '—'}\n"
        f"{'Auto-applied, scheduled by ' + move['created_by'] if is_auto and move['created_by'] else 'By: ' + moved_by}",
        chat_id=whatsapp_chat_id_for_site(move["to_location"], move["from_location"]) or ULTRAMSG_SITEPULSE_GROUP_CHAT_ID
    )


def apply_due_scheduled_moves(asset_id=None):
    """Scheduled location moves apply themselves once their date arrives --
    there's no cron here, so we check for anything due whenever an asset
    page (or the dashboard) loads and apply it right then. asset_id=None
    checks every asset (used on the dashboard); a specific id scopes it to
    one asset's page load.
    """
    db = get_db()
    today = date.today().isoformat()
    mover = current_user.name or current_user.email if current_user.is_authenticated else "System"
    if asset_id is not None:
        due = db.execute(
            "SELECT * FROM sitepulse_usage_log WHERE asset_id = ? AND entry_kind='move' AND move_status='Scheduled' AND scheduled_date <= ?",
            (asset_id, today)
        ).fetchall()
    else:
        due = db.execute(
            "SELECT * FROM sitepulse_usage_log WHERE entry_kind='move' AND move_status='Scheduled' AND scheduled_date <= ?",
            (today,)
        ).fetchall()
    for move in due:
        _apply_move(db, move, mover, is_auto=True)
    if due:
        db.commit()


def _clean_project_id(db, raw_value):
    """Validate a project_id coming from a form before it's ever stored.
    Returns a real int id if it refers to an actual tracker_projects row,
    otherwise None -- never stores a stray/tampered/stale value. Also
    used by edit forms: if the field wasn't in the submission at all
    (None), the caller is expected to fall back to the existing value
    itself rather than call this, so an omitted field can never silently
    clear a link.
    """
    if not raw_value:
        return None
    try:
        pid = int(raw_value)
    except (TypeError, ValueError):
        return None
    row = db.execute("SELECT id FROM tracker_projects WHERE id = ?", (pid,)).fetchone()
    return pid if row else None


def _backfill_project_links(db):
    """Phase 1 project-identity backfill. Runs every startup (cheap,
    idempotent) and only ever touches rows where project_id is still
    NULL -- once a row is linked (by this or by a person), it's never
    revisited or overwritten. Exact-match only: never guesses between
    multiple candidates, those go to project_link_review instead.
    """
    now = datetime.utcnow().isoformat()
    targets = [
        ("inventory_concrete_requests", "project"),
        ("inventory_purchase_requests", "job_name"),
        ("sitepulse_usage_log", "job_name"),
        ("sitepulse_rentals", "job_name"),
    ]
    for table, text_col in targets:
        rows = db.execute(
            f"SELECT id, {text_col} FROM {table} WHERE project_id IS NULL AND {text_col} IS NOT NULL AND TRIM({text_col}) != ''"
        ).fetchall()
        for row_id, raw_text in rows:
            free_text = raw_text.strip()
            matches = db.execute(
                "SELECT id FROM tracker_projects WHERE LOWER(TRIM(name)) = LOWER(?)", (free_text,)
            ).fetchall()
            if len(matches) == 1:
                db.execute(f"UPDATE {table} SET project_id = ? WHERE id = ?", (matches[0][0], row_id))
            elif len(matches) > 1:
                already_flagged = db.execute(
                    "SELECT id FROM project_link_review WHERE source_table = ? AND source_id = ? AND resolved = 0",
                    (table, row_id)
                ).fetchone()
                if not already_flagged:
                    db.execute(
                        "INSERT INTO project_link_review (source_table, source_id, free_text_value, reason, candidate_project_ids, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                        (table, row_id, free_text, "ambiguous: multiple projects share this name",
                         ",".join(str(m[0]) for m in matches), now)
                    )
            # 0 matches: left alone, no review needed -- most likely a
            # real job that simply isn't (yet) a Project Hunt bid.


PERMISSION_CATALOG = [
    # (key, category, label)
    # -- module access --
    ("module:project_hunt:view", "module", "Project Hunt"),
    ("module:equipment_center:view", "module", "Equipment Center"),
    ("module:sitepulse:view", "module", "SitePulse"),
    ("module:product_intelligence:view", "module", "Product Intelligence"),
    ("module:atlas:view", "module", "Atlas"),
    ("module:bidflow:view", "module", "BidFlow (not yet built)"),
    ("module:engineering:view", "module", "Engineering (not yet built)"),
    ("module:finance:view", "module", "Finance (not yet built)"),
    ("module:team_admin:view", "module", "Team / Admin"),
    # -- actions --
    ("action:project_hunt:manage", "action", "Manage Projects"),
    ("action:equipment_center:manage", "action", "Manage Equipment"),
    ("action:sitepulse:manage", "action", "Manage SitePulse Requests"),
    ("action:sitepulse:place_order", "action", "Place Purchase Orders"),
    ("action:product_intelligence:manage", "action", "Manage Product Intelligence"),
    # Procurement approval gate (item 3): deliberately a SEPARATE
    # permission from action:product_intelligence:manage above -- the
    # procurement approver (e.g. a Procurement Manager) is not
    # necessarily the same person as whoever manages the dev pipeline
    # (Ayoub), and shouldn't need to be granted full PI-manage rights
    # just to approve/return incoming requests. No name/email is
    # hardcoded anywhere -- whoever holds this permission (via role or
    # direct grant) is the approver, and that can change with zero code
    # changes.
    ("action:product_intelligence:approve_requests", "action", "Approve Product Requests (Procurement)"),
    ("action:team_admin:manage_users", "action", "Manage Users"),
    ("action:team_admin:manage_whatsapp", "action", "Manage WhatsApp Groups"),
    # -- Phase 3A additions: these three existed as gaps in the Phase 2
    # audit (is_admin()-gated routes with no specific matching key). Not
    # wired into any route yet -- that migration is a later, separate
    # step. Added now, seeded, and backfilled onto Administrator so the
    # model is ready when that migration happens, with zero behavior
    # change today.
    ("action:system_data:manage", "action", "Manage System Data (Backup / Restore / Import / Export)"),
    ("action:activity_log:view", "action", "View Activity Logs"),
    ("action:sitepulse:manage_inventory", "action", "Manage SitePulse Inventory (incl. deletion)"),
    # -- Atlas-specific: separate from the manual action permissions above
    # on purpose (a person can be allowed to do something manually
    # without allowing Atlas to do it on their behalf, or vice versa) --
    ("atlas:view_business_data", "atlas", "Atlas: View Business Data"),
    ("atlas:create_requests", "atlas", "Atlas: Create Requests"),
]

# name -> set of permission keys granted by that role
ROLE_DEFAULT_PERMISSIONS = {
    "Administrator": [key for key, _, _ in PERMISSION_CATALOG],  # everything
    "Project Manager": [
        "module:project_hunt:view", "action:project_hunt:manage",
        "module:equipment_center:view", "action:equipment_center:manage",
        "module:sitepulse:view", "action:sitepulse:manage",
        "module:atlas:view", "atlas:view_business_data", "atlas:create_requests",
    ],
    "Procurement": [
        "module:sitepulse:view", "action:sitepulse:manage", "action:sitepulse:place_order",
        "module:equipment_center:view", "action:equipment_center:manage",
        # Item 3 (procurement approval gate): Procurement approvers need
        # to see the Product Intelligence "Pending Approval" queue and
        # act on it -- module:product_intelligence:view (see) plus the
        # dedicated approve_requests action (act), NOT the broader
        # action:product_intelligence:manage (that stays scoped to
        # whoever runs the dev pipeline, e.g. Ayoub).
        "module:product_intelligence:view", "action:product_intelligence:approve_requests",
    ],
    "Estimator": [
        "module:project_hunt:view",
        "module:equipment_center:view", "action:equipment_center:manage",
        "module:sitepulse:view", "action:sitepulse:manage",
    ],
    "Operations": [
        "module:equipment_center:view", "action:equipment_center:manage",
        "module:sitepulse:view", "action:sitepulse:manage",
    ],
    "Employee": [
        "module:equipment_center:view", "action:equipment_center:manage",
        "module:sitepulse:view", "action:sitepulse:manage",
    ],
}

# Product Intelligence 2.0 role simplification: going forward, only these
# three roles are offered in the Users & Permissions assignment UI.
# "Project Manager", "Estimator", and "Employee" remain fully defined
# above (their permission bundles, seeding, and every existing
# user_roles row referencing them keep working exactly as before) --
# they're just no longer offered as a NEW assignment choice. This is a
# UI-layer restriction only: user_has_permission() still resolves
# through role_permissions/user_permission_overrides exactly as before,
# nothing here changes what a role grants or how permissions resolve.
ASSIGNABLE_ROLES = ["Administrator", "Procurement", "Operations"]


def _seed_roles_and_permissions(db):
    """Idempotent: inserts each role/permission once, never overwrites an
    existing row (so any hand-edited role_permissions from the admin UI
    survive every restart untouched)."""
    now = datetime.utcnow().isoformat()
    for name in ROLE_DEFAULT_PERMISSIONS:
        db.execute("INSERT OR IGNORE INTO roles (name, created_at) VALUES (?, ?)", (name, now))
    for key, category, label in PERMISSION_CATALOG:
        db.execute("INSERT OR IGNORE INTO permissions (key, category, label) VALUES (?, ?, ?)", (key, category, label))

    # Only wire up role_permissions the FIRST time a role is seeded (i.e.
    # if it currently has zero permission rows) -- once an admin has
    # edited a role's permissions from the UI, this must never silently
    # re-apply the defaults over their changes.
    for role_name, perm_keys in ROLE_DEFAULT_PERMISSIONS.items():
        role_row = db.execute("SELECT id FROM roles WHERE name = ?", (role_name,)).fetchone()
        if not role_row:
            continue
        role_id = role_row[0]
        existing = db.execute("SELECT COUNT(*) FROM role_permissions WHERE role_id = ?", (role_id,)).fetchone()[0]
        if existing > 0:
            continue
        for key in perm_keys:
            perm_row = db.execute("SELECT id FROM permissions WHERE key = ?", (key,)).fetchone()
            if perm_row:
                db.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_id) VALUES (?, ?)", (role_id, perm_row[0]))


def _grant_administrator_new_permissions(db, keys):
    """_seed_roles_and_permissions only wires up a role's permissions the
    FIRST time that role has zero rows -- intentional, so an admin's
    hand-edits from the UI are never silently overwritten. That means a
    permission key added to the catalog after a role was first seeded
    (like the three Phase 3A additions above) would never actually reach
    Administrator without this. Idempotent, additive-only, and scoped to
    exactly the keys passed in -- it can't remove or change anything an
    admin has already configured, and every existing admin keeps every
    permission they already had."""
    role_row = db.execute("SELECT id FROM roles WHERE name = 'Administrator'").fetchone()
    if not role_row:
        return
    role_id = role_row[0]
    for key in keys:
        perm_row = db.execute("SELECT id FROM permissions WHERE key = ?", (key,)).fetchone()
        if perm_row:
            db.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_id) VALUES (?, ?)", (role_id, perm_row[0]))


def _grant_role_new_permissions(db, role_name, keys):
    """Same idea as _grant_administrator_new_permissions, generalized to
    any role -- used for item 3 (procurement approval gate) so existing
    Procurement role holders actually receive the new
    action:product_intelligence:approve_requests permission (and the
    module:product_intelligence:view it depends on) without their
    role's other, possibly hand-edited, permissions being touched.
    Idempotent (INSERT OR IGNORE) and purely additive."""
    role_row = db.execute("SELECT id FROM roles WHERE name = ?", (role_name,)).fetchone()
    if not role_row:
        return
    role_id = role_row[0]
    for key in keys:
        perm_row = db.execute("SELECT id FROM permissions WHERE key = ?", (key,)).fetchone()
        if perm_row:
            db.execute("INSERT OR IGNORE INTO role_permissions (role_id, permission_id) VALUES (?, ?)", (role_id, perm_row[0]))


def _backfill_user_roles(db):
    """One-time-per-user backfill: gives every existing account whose
    email appears in the legacy hardcoded lists an equivalent role (or,
    where no single role fits, a role plus explicit overrides). Only
    runs for a user if they have ZERO roles and ZERO overrides already --
    never touches an account an admin has since configured by hand.

    This does not remove or replace ADMIN_EMAILS / FULL_ACCESS_EMAILS /
    ATLAS_ACCESS_EMAILS / PROCUREMENT_EMAILS / WHATSAPP_ADMIN_EMAILS --
    those keep gating the existing routes exactly as before. This backfill
    only populates the NEW system so it can be verified against the old
    one before anything old is ever removed.
    """
    now = datetime.utcnow().isoformat()
    role_id_by_name = {name: rid for rid, name in db.execute("SELECT id, name FROM roles").fetchall()}
    perm_id_by_key = {key: pid for pid, key in db.execute("SELECT id, key FROM permissions").fetchall()}

    def already_configured(user_id):
        has_role = db.execute("SELECT 1 FROM user_roles WHERE user_id = ?", (user_id,)).fetchone()
        has_override = db.execute("SELECT 1 FROM user_permission_overrides WHERE user_id = ?", (user_id,)).fetchone()
        return bool(has_role or has_override)

    def assign_role(user_id, role_name):
        rid = role_id_by_name.get(role_name)
        if rid:
            db.execute("INSERT OR IGNORE INTO user_roles (user_id, role_id) VALUES (?, ?)", (user_id, rid))

    def grant_override(user_id, perm_key):
        pid = perm_id_by_key.get(perm_key)
        if pid:
            db.execute(
                "INSERT OR IGNORE INTO user_permission_overrides (user_id, permission_id, state, granted_by, updated_at) VALUES (?, ?, 'grant', 'phase2_backfill', ?)",
                (user_id, pid, now)
            )

    users = db.execute("SELECT id, email FROM users").fetchall()
    for uid, email in users:
        if already_configured(uid):
            continue  # an admin (or a prior run) already set this user up -- never overwrite

        if email in ADMIN_EMAILS:
            assign_role(uid, "Administrator")
        elif email in FULL_ACCESS_EMAILS:
            assign_role(uid, "Project Manager")
            if email in PROCUREMENT_EMAILS:
                assign_role(uid, "Procurement")
            if email in ATLAS_ACCESS_EMAILS:
                pass  # Project Manager role already grants Atlas access
        elif email in ATLAS_ACCESS_EMAILS:
            # In today's hardcoded lists this is only
            # rebecca@nomaengineering.com -- Atlas access without full
            # Project Hunt access. No single role fits that shape, so:
            # baseline Employee role + explicit Atlas overrides. This is
            # exactly the case the override system exists for.
            assign_role(uid, "Employee")
            grant_override(uid, "module:atlas:view")
            grant_override(uid, "atlas:view_business_data")
            grant_override(uid, "atlas:create_requests")
        elif email in PROCUREMENT_EMAILS:
            assign_role(uid, "Procurement")
        else:
            assign_role(uid, "Employee")


def user_has_permission(user, key):
    """The single resolver every permission check in Phase 2 goes through.
    Explicit deny always wins over everything. Explicit grant wins over
    role membership. No override at all falls back to whatever the
    user's role(s) grant. An unknown permission key is a deny, not a
    crash -- a typo in a tool's `permission` field fails closed. A
    missing/empty key is ALSO a deny, not an automatic grant -- fail
    closed always, never fail open. (This was a real bug: `if not key:
    return True` let a None/empty key silently authorize anything that
    called this resolver with one -- fixed after CTO audit.)"""
    if not key:
        return False
    if not (user and getattr(user, "is_authenticated", False)):
        return False
    db = get_db()
    perm_row = db.execute("SELECT id FROM permissions WHERE key = ?", (key,)).fetchone()
    if not perm_row:
        return False
    pid = perm_row["id"]
    override = db.execute(
        "SELECT state FROM user_permission_overrides WHERE user_id = ? AND permission_id = ?",
        (user.id, pid)
    ).fetchone()
    if override:
        return override["state"] == "grant"
    role_grant = db.execute(
        "SELECT 1 FROM user_roles ur JOIN role_permissions rp ON rp.role_id = ur.role_id "
        "WHERE ur.user_id = ? AND rp.permission_id = ? LIMIT 1",
        (user.id, pid)
    ).fetchone()
    return role_grant is not None


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            created_at TEXT
        );

        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section TEXT NOT NULL,
            asset_id INTEGER,
            entity_type TEXT NOT NULL,
            entity_id INTEGER,
            action TEXT NOT NULL,
            field TEXT,
            old_value TEXT,
            new_value TEXT,
            user_email TEXT,
            created_at TEXT NOT NULL
        );

        -- Per-site WhatsApp group routing (item 16) -- e.g. anything
        -- mentioning "Peninsula" posts to the Peninsula group instead of
        -- the default SitePulse group. Managed from an admin page so new
        -- sites/groups can be added without a code change.
        CREATE TABLE IF NOT EXISTS whatsapp_site_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            keyword TEXT NOT NULL UNIQUE,
            chat_id TEXT NOT NULL,
            created_at TEXT NOT NULL
        );

        -- SitePulse: equipment + rentals only. Concrete/materials moved to
        -- Site Inventory below -- this is the "trim SitePulse down" ask.
        CREATE TABLE IF NOT EXISTS sitepulse_assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL, description TEXT, year TEXT, serial_number TEXT,
            value TEXT, daily_rate TEXT, weekly_rate TEXT, monthly_rate TEXT,
            status TEXT DEFAULT 'Available', location TEXT, hours_mileage TEXT,
            created_at TEXT, updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS sitepulse_usage_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, asset_id INTEGER NOT NULL,
            usage_type TEXT DEFAULT 'Internal Job', job_name TEXT, project_id INTEGER, job_address TEXT,
            client TEXT, out_date TEXT, duration_unit TEXT, return_date TEXT, notes TEXT,
            photo_filename TEXT, created_at TEXT,
            entry_kind TEXT DEFAULT 'usage', from_location TEXT, to_location TEXT,
            mileage_hours TEXT, move_status TEXT DEFAULT 'Applied', scheduled_date TEXT,
            scheduled_time TEXT, move_reason TEXT, status_at_move TEXT, moved_by TEXT,
            applied_at TEXT, created_by TEXT,
            FOREIGN KEY (asset_id) REFERENCES sitepulse_assets (id)
        );
        CREATE TABLE IF NOT EXISTS sitepulse_maintenance_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, asset_id INTEGER NOT NULL,
            entry_date TEXT, work_done TEXT, parts TEXT, hours_at_service TEXT,
            reported_by TEXT, resolved INTEGER DEFAULT 0, photo_filename TEXT, created_at TEXT,
            FOREIGN KEY (asset_id) REFERENCES sitepulse_assets (id)
        );
        CREATE TABLE IF NOT EXISTS sitepulse_mileage_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT, asset_id INTEGER NOT NULL,
            reading_date TEXT NOT NULL, mileage TEXT NOT NULL, notes TEXT, created_at TEXT,
            FOREIGN KEY (asset_id) REFERENCES sitepulse_assets (id)
        );
        CREATE TABLE IF NOT EXISTS sitepulse_rentals (
            id INTEGER PRIMARY KEY AUTOINCREMENT, vendor TEXT NOT NULL,
            equipment_description TEXT NOT NULL, job_name TEXT, project_id INTEGER, rate_amount TEXT,
            rate_period TEXT DEFAULT 'Daily', rented_date TEXT NOT NULL, due_date TEXT,
            returned_date TEXT, notes TEXT, created_at TEXT, updated_at TEXT
        );

        -- Site Inventory: concrete requests + material inventory, split out
        -- of SitePulse into their own section per today's direction.
        CREATE TABLE IF NOT EXISTS inventory_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT, item_name TEXT NOT NULL,
            site TEXT NOT NULL, quantity TEXT, unit TEXT, shelf_location TEXT,
            notes TEXT, created_at TEXT, updated_at TEXT
        );

        -- Request Center + Product Intelligence. Employees submit a
        -- request (feature_requests); every status change is logged to
        -- feature_request_status_history, which is also what the employee
        -- sees as their timeline. feature_request_intelligence holds the
        -- admin-only fields (notes, solution, testing, feedback) in a
        -- genuinely separate table -- not hidden columns on the same
        -- table -- so no employee-facing query can ever touch it.
        CREATE TABLE IF NOT EXISTS feature_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            requester_email TEXT NOT NULL,
            requester_name TEXT,
            department TEXT,
            original_request TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Submitted',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        -- Departments is its own table (not a hardcoded list in code) so
        -- new departments can be added from the admin UI as the company
        -- grows, with zero code changes or redesign required.
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL
        );
        -- Command Center roadmap: the modules being built, which lane
        -- they're in (now/next/later), and how far along each is.
        CREATE TABLE IF NOT EXISTS roadmap_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            lane TEXT NOT NULL DEFAULT 'later',
            note TEXT,
            progress_pct INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            updated_at TEXT NOT NULL
        );
        -- Phase 1 project-identity migration: when a free-text project
        -- name matches more than one tracker_projects row (or needs a
        -- human to confirm), it's recorded here instead of guessed.
        -- Nothing here is ever auto-resolved; this is purely a review
        -- queue for a person to link manually later.
        CREATE TABLE IF NOT EXISTS project_link_review (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_table TEXT NOT NULL,
            source_id INTEGER NOT NULL,
            free_text_value TEXT NOT NULL,
            reason TEXT NOT NULL,
            candidate_project_ids TEXT,
            resolved INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        );
        -- Phase 2: Roles + Permissions. Additive alongside the existing
        -- hardcoded email lists -- those are NOT removed this phase. See
        -- _seed_roles_and_permissions() / _backfill_user_roles().
        CREATE TABLE IF NOT EXISTS roles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            description TEXT,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            key TEXT NOT NULL UNIQUE,
            category TEXT NOT NULL,
            label TEXT NOT NULL,
            description TEXT
        );
        CREATE TABLE IF NOT EXISTS role_permissions (
            role_id INTEGER NOT NULL,
            permission_id INTEGER NOT NULL,
            PRIMARY KEY (role_id, permission_id)
        );
        CREATE TABLE IF NOT EXISTS user_roles (
            user_id INTEGER NOT NULL,
            role_id INTEGER NOT NULL,
            PRIMARY KEY (user_id, role_id)
        );
        CREATE TABLE IF NOT EXISTS user_permission_overrides (
            user_id INTEGER NOT NULL,
            permission_id INTEGER NOT NULL,
            state TEXT NOT NULL,
            granted_by TEXT,
            updated_at TEXT NOT NULL,
            PRIMARY KEY (user_id, permission_id)
        );
        CREATE TABLE IF NOT EXISTS feature_request_status_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            feature_request_id INTEGER NOT NULL,
            status TEXT NOT NULL,
            release_note TEXT,
            changed_by TEXT,
            changed_at TEXT NOT NULL,
            FOREIGN KEY (feature_request_id) REFERENCES feature_requests(id)
        );
        CREATE TABLE IF NOT EXISTS feature_request_intelligence (
            feature_request_id INTEGER PRIMARY KEY,
            buildiq_module TEXT,
            internal_notes TEXT,
            solution_built TEXT,
            testing_notes TEXT,
            user_feedback TEXT,
            release_date TEXT,
            updated_at TEXT,
            FOREIGN KEY (feature_request_id) REFERENCES feature_requests(id)
        );
        CREATE TABLE IF NOT EXISTS feature_request_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            feature_request_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            uploaded_by TEXT,
            created_at TEXT,
            FOREIGN KEY (feature_request_id) REFERENCES feature_requests(id)
        );
        -- Procurement approval gate (independent dimension from the
        -- feature_requests.status dev-lifecycle column -- see the long
        -- comment above _log_request_status()/product_intelligence()
        -- for why these are deliberately NOT the same field). Mirrors
        -- the existing status/feature_request_status_history pattern:
        -- feature_requests carries a fast "current decision" cache
        -- (columns added via the ALTER block below, since this table
        -- already existed before this feature), this table is the full
        -- audit trail of every decision ever made (not just the latest).
        CREATE TABLE IF NOT EXISTS feature_request_approvals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            feature_request_id INTEGER NOT NULL,
            decision TEXT NOT NULL,
            reason TEXT,
            decided_by TEXT NOT NULL,
            decided_at TEXT NOT NULL,
            FOREIGN KEY (feature_request_id) REFERENCES feature_requests(id)
        );
        -- v4 (employee feedback loop / Update & Resubmit): a Resubmission
        -- is NOT an approval decision -- overloading feature_request_approvals
        -- with a fake "decision" value for it would corrupt that table's
        -- actual meaning (every real row there already means "a
        -- procurement approver made this call", which a resubmission
        -- specifically is not -- the REQUESTER performs it). Smallest
        -- clean addition: its own tiny, purely-additive table, exactly
        -- mirroring feature_request_approvals' shape/intent (an
        -- append-only event log), so it can be merged into the same
        -- chronological timeline views without confusing the two kinds
        -- of event.
        CREATE TABLE IF NOT EXISTS feature_request_resubmissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            feature_request_id INTEGER NOT NULL,
            resubmitted_by TEXT NOT NULL,
            resubmitted_at TEXT NOT NULL,
            FOREIGN KEY (feature_request_id) REFERENCES feature_requests(id)
        );

        CREATE TABLE IF NOT EXISTS inventory_concrete_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT, project TEXT NOT NULL,
            project_id INTEGER,
            job_site_address TEXT, area_description TEXT, pour_date TEXT NOT NULL,
            pour_time TEXT, mix_design_psi TEXT, mix_slump TEXT, concrete_amount TEXT,
            truck_spacing TEXT, pump_type TEXT, pump_size TEXT, pump_arrival_time TEXT,
            lab_required TEXT, lab_time TEXT, drilling_required TEXT, drilling_time TEXT,
            requested_by TEXT, requested_signature TEXT, requested_date TEXT,
            ordered_by TEXT, ordered_signature TEXT, ordered_date TEXT,
            concrete_company TEXT, concrete_company_phone TEXT,
            pump_company TEXT, pump_company_phone TEXT,
            lab_company TEXT, drilling_company TEXT, drilling_company_phone TEXT,
            status TEXT DEFAULT 'Submitted',
            created_at TEXT, updated_at TEXT
        );

        -- Purchase Request Form (DIL-24-CON-F-PR-1) -- same paper-replica
        -- approach as Concrete Requests, plus the same Submit/Scheduled/
        -- Completed/Delete workflow.
        CREATE TABLE IF NOT EXISTS inventory_purchase_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pr_number TEXT, request_date TEXT NOT NULL,
            job_name TEXT, project_id INTEGER, location_description TEXT,
            requested_by TEXT, needed_on TEXT, source_of_supply TEXT,
            requestor_signature TEXT, requestor_date TEXT,
            ordered_by TEXT, ordered_date TEXT, vendor_company TEXT, vendor_company_phone TEXT,
            status TEXT DEFAULT 'Submitted',
            created_at TEXT, updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS inventory_purchase_request_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT, purchase_request_id INTEGER NOT NULL,
            item TEXT, description TEXT, supplier TEXT, qty TEXT, unit TEXT,
            FOREIGN KEY (purchase_request_id) REFERENCES inventory_purchase_requests (id)
        );

        -- Bid Tracker: full port of Command Center's schema.
        CREATE TABLE IF NOT EXISTS tracker_projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, client TEXT,
            address TEXT, bid_due_date TEXT, estimated_value TEXT,
            status TEXT DEFAULT 'In Progress', assigned_to TEXT, notes TEXT,
            created_at TEXT, updated_at TEXT
        );
        CREATE TABLE IF NOT EXISTS tracker_quotes (
            id INTEGER PRIMARY KEY AUTOINCREMENT, project_id INTEGER NOT NULL,
            trade TEXT NOT NULL, vendor_name TEXT, vendor_contact TEXT, vendor_email TEXT,
            vendor_phone TEXT, rfq_sent_date TEXT, status TEXT DEFAULT 'Not Sent',
            is_submit_blocking INTEGER DEFAULT 0, amount TEXT, notes TEXT,
            follow_up_email TEXT, rfq_email TEXT, attachment_filename TEXT,
            attachment_original_name TEXT, created_at TEXT, updated_at TEXT,
            FOREIGN KEY (project_id) REFERENCES tracker_projects (id)
        );
        CREATE TABLE IF NOT EXISTS tracker_docs (
            id INTEGER PRIMARY KEY AUTOINCREMENT, project_id INTEGER NOT NULL,
            doc_name TEXT, doc_type TEXT, status TEXT DEFAULT 'Needed',
            notes TEXT, link TEXT, created_at TEXT,
            FOREIGN KEY (project_id) REFERENCES tracker_projects (id)
        );
        CREATE TABLE IF NOT EXISTS tracker_unit_prices (
            id INTEGER PRIMARY KEY AUTOINCREMENT, category TEXT, item TEXT NOT NULL,
            unit TEXT, price TEXT, notes TEXT, updated_at TEXT
        );
    """)

    # Safe migration: adds the duration_unit column if this database
    # already existed before item 8 was built (e.g. real signups already
    # on Railway). CREATE TABLE IF NOT EXISTS above only handles brand-new
    # databases -- an existing table needs this explicit ALTER instead,
    # or every INSERT into sitepulse_usage_log fails with "no column
    # named duration_unit" once the code expects that column to exist.
    try:
        db.execute("ALTER TABLE sitepulse_usage_log ADD COLUMN duration_unit TEXT")
    except sqlite3.OperationalError:
        pass  # column already exists -- nothing to do

    # Same situation for the "status" column on concrete/purchase requests --
    # added when Submit/Scheduled/Completed was built, after some databases
    # may already have existed without it.
    try:
        db.execute("ALTER TABLE inventory_concrete_requests ADD COLUMN status TEXT DEFAULT 'Submitted'")
    except sqlite3.OperationalError:
        pass
    try:
        db.execute("ALTER TABLE inventory_purchase_requests ADD COLUMN status TEXT DEFAULT 'Submitted'")
    except sqlite3.OperationalError:
        pass

    # Order-placement fields added for item 14 -- procurement records who
    # ordered, when, and the vendor/contact for each piece (concrete truck,
    # pump, lab, drilling; vendor for purchase requests). Same safe-migration
    # pattern as above for databases created before this existed.
    for column_sql in [
        "ALTER TABLE inventory_concrete_requests ADD COLUMN concrete_company TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN concrete_company_phone TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN pump_company TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN pump_company_phone TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN lab_company TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN drilling_company TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN drilling_company_phone TEXT",
        "ALTER TABLE inventory_purchase_requests ADD COLUMN ordered_by TEXT",
        "ALTER TABLE inventory_purchase_requests ADD COLUMN ordered_date TEXT",
        "ALTER TABLE inventory_purchase_requests ADD COLUMN vendor_company TEXT",
        "ALTER TABLE inventory_purchase_requests ADD COLUMN vendor_company_phone TEXT",
        "ALTER TABLE inventory_purchase_requests ADD COLUMN expected_delivery_date TEXT",
        # Location-move auto-tracking (item 15) -- Status & Location now
        # writes movement entries straight into the usage log, including
        # ones scheduled for a future date.
        "ALTER TABLE sitepulse_usage_log ADD COLUMN entry_kind TEXT DEFAULT 'usage'",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN from_location TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN to_location TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN mileage_hours TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN move_status TEXT DEFAULT 'Applied'",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN scheduled_date TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN scheduled_time TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN move_reason TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN status_at_move TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN moved_by TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN applied_at TEXT",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN created_by TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN pump_type TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN concrete_arrival_time TEXT",
        "ALTER TABLE inventory_concrete_requests ADD COLUMN full_reminder_sent_at TEXT",
        "ALTER TABLE users ADD COLUMN department TEXT",
        "ALTER TABLE inventory_purchase_request_items ADD COLUMN unit TEXT",
        # Phase 1: project identity columns (see project_link_review above).
        # Nullable and additive -- existing free-text values are untouched.
        "ALTER TABLE inventory_concrete_requests ADD COLUMN project_id INTEGER",
        "ALTER TABLE inventory_purchase_requests ADD COLUMN project_id INTEGER",
        "ALTER TABLE sitepulse_usage_log ADD COLUMN project_id INTEGER",
        "ALTER TABLE sitepulse_rentals ADD COLUMN project_id INTEGER",
        # Procurement approval gate columns (see feature_request_approvals
        # above). Deliberately NO SQL-level DEFAULT here -- rows that
        # existed before this migration must come through as NULL so the
        # one-time backfill below (and only that backfill) can mark them
        # 'Approved' explicitly, once, conservatively. Every new insert
        # after this point sets approval_status itself (in request_center()
        # below) -- it never relies on a column default.
        "ALTER TABLE feature_requests ADD COLUMN approval_status TEXT",
        "ALTER TABLE feature_requests ADD COLUMN approval_decided_by TEXT",
        "ALTER TABLE feature_requests ADD COLUMN approval_decided_at TEXT",
        "ALTER TABLE feature_requests ADD COLUMN approval_reason TEXT",
    ]:
        try:
            db.execute(column_sql)
        except sqlite3.OperationalError:
            pass

    # One-time, idempotent backfill for the procurement approval gate:
    # any row that predates this migration has approval_status IS NULL
    # (see the ALTER above -- no SQL default on purpose). Those requests
    # were never subject to an approval gate at all, so treating them as
    # "not yet approved" would silently block/reclassify real historical
    # work that already went through Building/Testing/Released under the
    # old rules. Conservative choice: mark them Approved, using their own
    # created_at as the decision time and a clearly-labeled system actor
    # (never a real person's name, so this never looks like someone
    # secretly approved old requests). Idempotent: only ever touches rows
    # still NULL, so running this on every app start is a no-op after the
    # first time. New requests (see request_center()) always set
    # approval_status='Pending' explicitly at insert time and therefore
    # never match this WHERE clause.
    _now_backfill = datetime.utcnow().isoformat()
    db.execute(
        """UPDATE feature_requests SET approval_status = 'Approved',
           approval_decided_by = 'system (predates approval gate)',
           approval_decided_at = COALESCE(created_at, ?)
           WHERE approval_status IS NULL""",
        (_now_backfill,)
    )
    db.commit()

    # Belt-and-suspenders: project_link_review table, same pattern as
    # departments below.
    try:
        db.execute(
            """CREATE TABLE IF NOT EXISTS project_link_review (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_table TEXT NOT NULL,
                source_id INTEGER NOT NULL,
                free_text_value TEXT NOT NULL,
                reason TEXT NOT NULL,
                candidate_project_ids TEXT,
                resolved INTEGER DEFAULT 0,
                created_at TEXT NOT NULL
            )"""
        )
    except sqlite3.OperationalError:
        pass

    # Indexes on the new project_id columns -- safe to run every startup,
    # CREATE INDEX IF NOT EXISTS is a no-op once they exist.
    for index_sql in [
        "CREATE INDEX IF NOT EXISTS idx_concrete_project_id ON inventory_concrete_requests(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_purchase_project_id ON inventory_purchase_requests(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_usage_log_project_id ON sitepulse_usage_log(project_id)",
        "CREATE INDEX IF NOT EXISTS idx_rentals_project_id ON sitepulse_rentals(project_id)",
    ]:
        try:
            db.execute(index_sql)
        except sqlite3.OperationalError:
            pass

    _backfill_project_links(db)

    # Phase 2: same belt-and-suspenders pattern for the roles/permissions
    # tables, plus their indexes.
    for table_sql in [
        """CREATE TABLE IF NOT EXISTS roles (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE,
            description TEXT, created_at TEXT NOT NULL)""",
        """CREATE TABLE IF NOT EXISTS permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT, key TEXT NOT NULL UNIQUE,
            category TEXT NOT NULL, label TEXT NOT NULL, description TEXT)""",
        """CREATE TABLE IF NOT EXISTS role_permissions (
            role_id INTEGER NOT NULL, permission_id INTEGER NOT NULL,
            PRIMARY KEY (role_id, permission_id))""",
        """CREATE TABLE IF NOT EXISTS user_roles (
            user_id INTEGER NOT NULL, role_id INTEGER NOT NULL,
            PRIMARY KEY (user_id, role_id))""",
        """CREATE TABLE IF NOT EXISTS user_permission_overrides (
            user_id INTEGER NOT NULL, permission_id INTEGER NOT NULL, state TEXT NOT NULL,
            granted_by TEXT, updated_at TEXT NOT NULL, PRIMARY KEY (user_id, permission_id))""",
    ]:
        try:
            db.execute(table_sql)
        except sqlite3.OperationalError:
            pass

    for index_sql in [
        "CREATE INDEX IF NOT EXISTS idx_role_permissions_role ON role_permissions(role_id)",
        "CREATE INDEX IF NOT EXISTS idx_user_roles_user ON user_roles(user_id)",
        "CREATE INDEX IF NOT EXISTS idx_user_overrides_user ON user_permission_overrides(user_id)",
    ]:
        try:
            db.execute(index_sql)
        except sqlite3.OperationalError:
            pass

    _seed_roles_and_permissions(db)
    _grant_administrator_new_permissions(db, [
        "action:system_data:manage",
        "action:activity_log:view",
        "action:sitepulse:manage_inventory",
        # Fix 1 (Administrator approval-permission bootstrap dead-end):
        # Administrator's role-default permission set already includes
        # EVERY key in PERMISSION_CATALOG (see ROLE_DEFAULT_PERMISSIONS
        # above -- "everything"), so a brand-new database seeds this
        # correctly the first time. The gap is existing databases: an
        # Administrator role row that already had its permissions wired
        # up (zero-rows check in _seed_roles_and_permissions) BEFORE
        # action:product_intelligence:approve_requests existed in the
        # catalog never automatically receives new keys added later --
        # exactly the class of gap this backfill helper exists for.
        # Without this, an existing Administrator can see the permission
        # toggle but not grant it to themselves (self-permission-
        # modification is intentionally blocked), a real dead end.
        "action:product_intelligence:approve_requests",
    ])
    # Item 3: existing Procurement role holders get the new approval
    # permission (and its view prerequisite) without any other part of
    # their role -- or any hand-edits an admin already made to it --
    # being touched.
    _grant_role_new_permissions(db, "Procurement", [
        "module:product_intelligence:view",
        "action:product_intelligence:approve_requests",
    ])
    _backfill_user_roles(db)

    # Belt-and-suspenders: create departments here too, as its own
    # standalone statement, in case it didn't take earlier (e.g. an
    # existing production DB that predates this table and whose
    # executescript run stopped short of it for any reason).
    try:
        db.execute(
            """CREATE TABLE IF NOT EXISTS departments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL
            )"""
        )
    except sqlite3.OperationalError:
        pass

    # Seed the initial department list once. After this, departments are
    # managed entirely from the admin UI (Users & Departments page) --
    # adding a new one is a row insert, not a code change.
    existing_dept_count = db.execute("SELECT COUNT(*) FROM departments").fetchone()[0]
    if existing_dept_count == 0:
        now = datetime.utcnow().isoformat()
        for dept_name in ["Estimating", "Procurement", "Operations"]:
            db.execute("INSERT OR IGNORE INTO departments (name, created_at) VALUES (?, ?)", (dept_name, now))

    # Same belt-and-suspenders pattern as departments: create the table as
    # its own standalone statement too, then seed once.
    try:
        db.execute(
            """CREATE TABLE IF NOT EXISTS roadmap_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                lane TEXT NOT NULL DEFAULT 'later',
                note TEXT,
                progress_pct INTEGER DEFAULT 0,
                sort_order INTEGER DEFAULT 0,
                updated_at TEXT NOT NULL
            )"""
        )
    except sqlite3.OperationalError:
        pass
    existing_roadmap_count = db.execute("SELECT COUNT(*) FROM roadmap_items").fetchone()[0]
    if existing_roadmap_count == 0:
        now = datetime.utcnow().isoformat()
        roadmap_seed = [
            # progress_pct is 0 for every freshly-seeded item deliberately --
            # there is no defensible measurable source for a completion
            # percentage on any of these, so none is invented. progress_pct
            # remains a real, admin-editable field (see roadmap_item_update)
            # for backward compatibility; it simply starts truthful (0)
            # instead of claiming unearned completion.
            ("Product Core", "now", "Canonical Project Identity foundation is complete -- concrete, purchase, and rental records link to real projects. Currently extending that connectivity into more of Project Hunt/SitePulse.", 0, 1),
            ("Product Intelligence", "now", "Command Center experience refinement -- visual hierarchy, real-data intelligence, and honest empty states.", 0, 2),
            ("Project Connectivity", "next", "Turning canonical Project Identity into useful connected project intelligence across modules.", 0, 3),
            ("Atlas", "evolving", "BuildIQ's intelligence and action layer -- read tools shipped; continuously gaining capability rather than reaching a fixed 100%.", 0, 4),
            ("BidFlow", "later", "Takeoff + bid system. Parked until the real estimating workflow/Excel sheets are available.", 0, 5),
            ("Redline", "later", "Parked intentionally.", 0, 6),
            ("Finance", "later", "Parked intentionally.", 0, 7),
        ]
        for name, lane, note, pct, order in roadmap_seed:
            db.execute(
                "INSERT INTO roadmap_items (name, lane, note, progress_pct, sort_order, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
                (name, lane, note, pct, order, now)
            )
    # NOTE: a non-empty roadmap_items table is NEVER auto-rewritten during
    # normal application startup. An earlier version of this function
    # called _correct_stale_roadmap_seed() here unconditionally -- that
    # ran against ANY existing database (including, eventually, a real
    # production one), silently deleting and replacing roadmap rows on
    # every boot. Removed entirely per CTO audit. If a legacy-seed
    # database genuinely needs upgrading to the new roadmap story, run
    # scripts/upgrade_roadmap_seed.py explicitly and deliberately --
    # never as a side effect of `import app`. Administrator-edited
    # roadmap data is never touched by ordinary startup.

    db.commit()
    db.close()


init_db()


class User(UserMixin):
    def __init__(self, row):
        self.id = row["id"]
        self.name = row["name"]
        self.email = row["email"]


@login_manager.user_loader
def load_user(user_id):
    db = get_db()
    row = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return User(row) if row else None


# ---------------------------------------------------------------------------
# Auth + Home
# ---------------------------------------------------------------------------

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        if not any(email.endswith(d) for d in ALLOWED_SIGNUP_DOMAINS) and email not in EXTRA_ALLOWED_SIGNUP_EMAILS:
            flash(f"Sign up with your {' or '.join(ALLOWED_SIGNUP_DOMAINS)} email.", "error")
            return redirect(url_for("signup"))
        db = get_db()
        existing = db.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone()
        if existing:
            flash("An account with that email already exists.", "error")
            return redirect(url_for("signup"))
        db.execute(
            "INSERT INTO users (name, email, password_hash, created_at) VALUES (?, ?, ?, ?)",
            (request.form.get("name", ""), email, generate_password_hash(request.form["password"]),
             datetime.utcnow().isoformat())
        )
        db.commit()
        # Runtime authorization no longer consults the legacy lists at
        # all (see is_admin()/_authorized() etc below) -- so a
        # legacy-listed person must be backfilled into a real role the
        # moment their account exists, not just at the next server
        # restart. _backfill_user_roles() is idempotent and only ever
        # touches a user with zero roles/overrides, so calling it here
        # is safe and cannot re-run for anyone already configured.
        _backfill_user_roles(db)
        db.commit()
        flash("Account created. Log in below.")
        return redirect(url_for("login"))
    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        db = get_db()
        row = db.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
        if row and check_password_hash(row["password_hash"], request.form["password"]):
            login_user(User(row))
            return redirect(url_for("home"))
        flash("Invalid email or password.", "error")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("home"))


@app.route("/team")
@login_required
def team_list():
    if not _authorized("module:team_admin:view"):
        return redirect(url_for("home"))
    db = get_db()
    users = db.execute("SELECT * FROM users ORDER BY created_at ASC").fetchall()
    return render_template("team.html", users=users)


@app.route("/team/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_team_member(user_id):
    if not _authorized("action:team_admin:manage_users"):
        return redirect(url_for("home"))
    db = get_db()
    target = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target:
        flash("User not found.", "error")
        return redirect(url_for("team_list"))
    if _is_protected_admin_account(target):
        flash("Can't remove an admin account this way.", "error")
        return redirect(url_for("team_list"))
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    flash(f"Removed {target['email']} -- they can sign up again with the same email.")
    return redirect(url_for("team_list"))


@app.route("/whatsapp-groups")
@login_required
def whatsapp_site_groups_list():
    if not is_whatsapp_admin():
        return redirect(url_for("home"))
    db = get_db()
    groups = db.execute("SELECT * FROM whatsapp_site_groups ORDER BY keyword ASC").fetchall()
    return render_template("whatsapp_site_groups.html", groups=groups,
                            default_chat_id=ULTRAMSG_SITEPULSE_GROUP_CHAT_ID or ULTRAMSG_GROUP_CHAT_ID)


@app.route("/whatsapp-groups/new", methods=["POST"])
@login_required
def whatsapp_site_groups_new():
    if not is_whatsapp_admin():
        return redirect(url_for("home"))
    db = get_db()
    keyword = request.form.get("keyword", "").strip()
    chat_id = request.form.get("chat_id", "").strip()
    if not keyword or not chat_id:
        flash("Both a site keyword and a chat ID are required.", "error")
        return redirect(url_for("whatsapp_site_groups_list"))
    try:
        db.execute("INSERT INTO whatsapp_site_groups (keyword, chat_id, created_at) VALUES (?, ?, ?)",
                   (keyword, chat_id, datetime.utcnow().isoformat()))
        db.commit()
        flash(f'Anything mentioning "{keyword}" now routes to that group.')
    except sqlite3.IntegrityError:
        flash(f'"{keyword}" is already mapped to a group -- delete it first if you want to change it.', "error")
    return redirect(url_for("whatsapp_site_groups_list"))


@app.route("/whatsapp-groups/<int:group_id>/delete", methods=["POST"])
@login_required
def whatsapp_site_groups_delete(group_id):
    if not is_whatsapp_admin():
        return redirect(url_for("home"))
    db = get_db()
    db.execute("DELETE FROM whatsapp_site_groups WHERE id = ?", (group_id,))
    db.commit()
    flash("Removed -- that site's notifications will fall back to the default group.")
    return redirect(url_for("whatsapp_site_groups_list"))


@app.route("/whatsapp-groups/test", methods=["POST"])
@login_required
def whatsapp_site_groups_test():
    if not is_whatsapp_admin():
        return redirect(url_for("home"))
    chat_id = request.form.get("chat_id", "").strip()
    label = request.form.get("label", "").strip() or "this group"
    if not chat_id:
        flash("No chat ID given to test.", "error")
        return redirect(url_for("whatsapp_site_groups_list"))
    ok, detail = send_whatsapp_group_message(
        f"\U0001F9EA Test notification\n"
        f"This confirms the group is receiving BuildIQ alerts correctly.\n"
        f"Triggered by: {current_user.name or current_user.email}",
        chat_id=chat_id
    )
    if ok:
        flash(f"Test message sent to {label} -- check WhatsApp to confirm it landed.")
    else:
        flash(f"Couldn't send to {label}: {detail}", "error")
    return redirect(url_for("whatsapp_site_groups_list"))


@app.route("/")
def home():
    return render_template("home.html")


# ---------------------------------------------------------------------------
# SitePulse -- Equipment
# ---------------------------------------------------------------------------

SP_STATUS_OPTIONS = ["Available", "Out on Job", "In Maintenance", "Sold", "Stolen"]
SP_STATUS_BADGE = {
    "Available": "status-awarded", "Out on Job": "status-inprogress",
    "In Maintenance": "status-pending", "Sold": "status-submitted", "Stolen": "status-submitted",
}


@app.template_filter("sp_statusclass")
def sp_statusclass(status):
    return SP_STATUS_BADGE.get(status, "status-pending")


@app.route("/sitepulse/")
@login_required
def sitepulse_dashboard():
    if not _authorized("module:equipment_center:view"):
        flash("You don't have access to Equipment Center.", "error")
        return redirect(url_for("home"))
    apply_due_scheduled_moves()
    db = get_db()
    status_filter = request.args.get("status", "")
    location_filter = request.args.get("location", "")

    conditions = []
    params = []
    if status_filter:
        conditions.append("status = ?")
        params.append(status_filter)
    else:
        conditions.append("status NOT IN ('Sold', 'Stolen')")
    if location_filter:
        conditions.append("location = ?")
        params.append(location_filter)
    query = "SELECT * FROM sitepulse_assets WHERE " + " AND ".join(conditions) + " ORDER BY name ASC"
    asset_rows = db.execute(query, params).fetchall()

    available_count = db.execute("SELECT COUNT(*) as c FROM sitepulse_assets WHERE status = 'Available'").fetchone()["c"]
    out_count = db.execute("SELECT COUNT(*) as c FROM sitepulse_assets WHERE status = 'Out on Job'").fetchone()["c"]
    maint_count = db.execute("SELECT COUNT(*) as c FROM sitepulse_assets WHERE status = 'In Maintenance'").fetchone()["c"]

    today_str = date.today().isoformat()
    active_rentals_count = db.execute(
        "SELECT COUNT(*) as c FROM sitepulse_rentals WHERE returned_date IS NULL OR returned_date = ''"
    ).fetchone()["c"]
    overdue_rentals_count = db.execute(
        "SELECT COUNT(*) as c FROM sitepulse_rentals WHERE (returned_date IS NULL OR returned_date = '') "
        "AND due_date IS NOT NULL AND due_date != '' AND due_date < ?", (today_str,)
    ).fetchone()["c"]

    locations = [r["location"] for r in db.execute(
        "SELECT DISTINCT location FROM sitepulse_assets WHERE location IS NOT NULL AND location != '' ORDER BY location ASC"
    ).fetchall()]

    # Latest activity per asset -- same "fold onto the asset row" approach as
    # real SitePulse: most recent usage vs. most recent maintenance, whichever
    # is newer wins, shown right in the dashboard row instead of a separate feed.
    latest_usage = {}
    for r in db.execute("SELECT * FROM sitepulse_usage_log ORDER BY created_at DESC").fetchall():
        if r["asset_id"] not in latest_usage:
            latest_usage[r["asset_id"]] = r
    latest_maint = {}
    for r in db.execute("SELECT * FROM sitepulse_maintenance_log ORDER BY created_at DESC").fetchall():
        if r["asset_id"] not in latest_maint:
            latest_maint[r["asset_id"]] = r

    assets = []
    for a in asset_rows:
        a_dict = dict(a)
        u = latest_usage.get(a["id"])
        m = latest_maint.get(a["id"])
        chosen, chosen_type = None, None
        if u and m:
            chosen, chosen_type = (u, "Usage") if u["created_at"] >= m["created_at"] else (m, "Maintenance")
        elif u:
            chosen, chosen_type = u, "Usage"
        elif m:
            chosen, chosen_type = m, "Maintenance"
        if chosen_type == "Usage":
            a_dict["latest_activity"] = chosen["job_name"] or chosen["client"] or chosen["usage_type"]
            a_dict["latest_activity_date"] = chosen["out_date"] or chosen["created_at"][:10]
        elif chosen_type == "Maintenance":
            status_tag = "Resolved" if chosen["resolved"] else "Open"
            a_dict["latest_activity"] = f"{chosen['work_done']} ({status_tag})" if chosen["work_done"] else f"Maintenance ({status_tag})"
            a_dict["latest_activity_date"] = chosen["entry_date"] or chosen["created_at"][:10]
        else:
            a_dict["latest_activity"] = None
            a_dict["latest_activity_date"] = None
        a_dict["latest_activity_type"] = chosen_type
        assets.append(a_dict)

    return render_template("sitepulse/dashboard.html", assets=assets, status_options=SP_STATUS_OPTIONS,
                            available_count=available_count, out_count=out_count, maint_count=maint_count,
                            active_rentals_count=active_rentals_count, overdue_rentals_count=overdue_rentals_count,
                            current_filter=status_filter, current_location=location_filter, locations=locations)


@app.route("/sitepulse/asset/new", methods=["GET", "POST"])
@login_required
def sitepulse_new_asset():
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    if request.method == "POST":
        db = get_db()
        now = datetime.utcnow().isoformat()
        cur = db.execute(
            """INSERT INTO sitepulse_assets (name, description, year, serial_number, value, daily_rate,
               weekly_rate, monthly_rate, status, location, hours_mileage, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (request.form["name"], request.form.get("description", ""), request.form.get("year", ""),
             request.form.get("serial_number", ""), request.form.get("value", ""),
             request.form.get("daily_rate", ""), request.form.get("weekly_rate", ""),
             request.form.get("monthly_rate", ""), "Available", request.form.get("location", ""),
             request.form.get("hours_mileage", ""), now, now)
        )
        log_activity("sitepulse", "asset", cur.lastrowid, "created", new_value=request.form["name"])
        db.commit()
        flash("Equipment added.")
        return redirect(url_for("sitepulse_dashboard"))
    return render_template("sitepulse/new_asset.html")


@app.route("/sitepulse/asset/<int:asset_id>")
@login_required
def sitepulse_view_asset(asset_id):
    if not _authorized("module:equipment_center:view"):
        flash("You don't have access to Equipment Center.", "error")
        return redirect(url_for("home"))
    apply_due_scheduled_moves(asset_id)
    db = get_db()
    a = db.execute("SELECT * FROM sitepulse_assets WHERE id = ?", (asset_id,)).fetchone()
    if not a:
        flash("Asset not found.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    usage = db.execute("SELECT * FROM sitepulse_usage_log WHERE asset_id = ? AND move_status != 'Scheduled' ORDER BY COALESCE(applied_at, out_date, created_at) DESC", (asset_id,)).fetchall()
    scheduled_moves = db.execute(
        "SELECT * FROM sitepulse_usage_log WHERE asset_id = ? AND entry_kind='move' AND move_status='Scheduled' ORDER BY scheduled_date ASC",
        (asset_id,)
    ).fetchall()
    maintenance = db.execute("SELECT * FROM sitepulse_maintenance_log WHERE asset_id = ? ORDER BY entry_date DESC", (asset_id,)).fetchall()

    mileage_entries = db.execute(
        "SELECT * FROM sitepulse_mileage_log WHERE asset_id = ? ORDER BY reading_date DESC", (asset_id,)
    ).fetchall()
    # Group readings by month (YYYY-MM), keeping the highest reading seen
    # in each month -- that's the odometer's value at that point, which is
    # what "mileage this month" means in practice (a running total, not
    # separate trip counts).
    monthly_totals = {}
    for m in mileage_entries:
        try:
            month_key = m["reading_date"][:7]
            reading_num = float(m["mileage"])
            if month_key not in monthly_totals or reading_num > monthly_totals[month_key]:
                monthly_totals[month_key] = reading_num
        except (ValueError, TypeError):
            pass
    monthly_totals_sorted = sorted(monthly_totals.items(), reverse=True)

    tracker_projects = db.execute(
        "SELECT id, name, client FROM tracker_projects WHERE status NOT IN ('Archived','Cancelled') ORDER BY name"
    ).fetchall()
    return render_template("sitepulse/asset.html", a=a, usage=usage, maintenance=maintenance,
                            scheduled_moves=scheduled_moves,
                            mileage_entries=mileage_entries, monthly_totals=monthly_totals_sorted,
                            status_options=SP_STATUS_OPTIONS, usage_type_options=["Internal Job", "External Rental"],
                            today=date.today().isoformat(), tracker_projects=tracker_projects)


@app.route("/sitepulse/asset/<int:asset_id>/edit-details", methods=["POST"])
@login_required
def sitepulse_edit_asset_details(asset_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    db.execute(
        """UPDATE sitepulse_assets SET name=?, description=?, year=?, serial_number=?, value=?,
           daily_rate=?, weekly_rate=?, monthly_rate=?, updated_at=? WHERE id=?""",
        (request.form["name"], request.form.get("description", ""), request.form.get("year", ""),
         request.form.get("serial_number", ""), request.form.get("value", ""),
         request.form.get("daily_rate", ""), request.form.get("weekly_rate", ""),
         request.form.get("monthly_rate", ""), datetime.utcnow().isoformat(), asset_id)
    )
    log_activity("sitepulse", "asset", asset_id, "updated", asset_id=asset_id, field="details", new_value=request.form["name"])
    db.commit()
    flash("Asset details updated.")
    return redirect(url_for("sitepulse_view_asset", asset_id=asset_id))


@app.route("/sitepulse/asset/<int:asset_id>/update", methods=["POST"])
@login_required
def sitepulse_update_asset(asset_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    old = db.execute("SELECT name, status, location FROM sitepulse_assets WHERE id = ?", (asset_id,)).fetchone()
    asset_name = old["name"]
    new_status = request.form["status"]
    new_location = request.form.get("location", "").strip()
    hours_mileage = request.form.get("hours_mileage", "")
    schedule_date = request.form.get("schedule_date", "").strip()
    schedule_time = request.form.get("schedule_time", "").strip()
    move_reason = request.form.get("move_reason", "").strip()
    now = datetime.utcnow().isoformat()
    today = date.today().isoformat()
    old_location = old["location"] or ""
    mover = current_user.name or current_user.email

    location_changed = new_location != old_location and new_location != ""
    # A date in the future schedules the move for later. A date that's
    # today or in the past backdates an already-happened move to that
    # actual date, instead of always stamping it with today. Blank means
    # "happening right now" -- unchanged from before.
    effective_date = schedule_date or today

    if location_changed and schedule_date and schedule_date > today:
        # Future move: don't touch the asset's location yet -- park it as a
        # Scheduled entry in the usage log. apply_due_scheduled_moves()
        # (or manually completing it) fills in the actual status/hours/mover
        # snapshot once it really happens, not at scheduling time.
        db.execute("UPDATE sitepulse_assets SET status=?, hours_mileage=?, updated_at=? WHERE id=?",
                   (new_status, hours_mileage, now, asset_id))
        cur = db.execute(
            """INSERT INTO sitepulse_usage_log (asset_id, entry_kind, from_location, to_location,
               move_status, scheduled_date, scheduled_time, move_reason, created_by, created_at)
               VALUES (?, 'move', ?, ?, 'Scheduled', ?, ?, ?, ?, ?)""",
            (asset_id, old_location, new_location, schedule_date, schedule_time, move_reason, mover, now)
        )
        log_activity("sitepulse", "move", cur.lastrowid, "scheduled", asset_id=asset_id,
                     field="location", old_value=old_location, new_value=new_location)
        db.commit()
        send_whatsapp_group_message(
            f"📅 Move scheduled: {asset_name}\n"
            f"{old_location or '—'} → {new_location}\n"
            f"Date: {schedule_date}{' at ' + schedule_time if schedule_time else ''}\n"
            + (f"Reason: {move_reason}\n" if move_reason else "")
            + f"Scheduled by: {mover}",
            chat_id=whatsapp_chat_id_for_site(new_location, old_location) or ULTRAMSG_SITEPULSE_GROUP_CHAT_ID
        )
        flash(f"Move to {new_location} scheduled for {schedule_date}.")
        return redirect(url_for("sitepulse_view_asset", asset_id=asset_id))

    db.execute("UPDATE sitepulse_assets SET status=?, location=?, hours_mileage=?, updated_at=? WHERE id=?",
               (new_status, new_location, hours_mileage, now, asset_id))

    if location_changed:
        cur = db.execute(
            """INSERT INTO sitepulse_usage_log (asset_id, entry_kind, from_location, to_location,
               mileage_hours, move_status, status_at_move, moved_by, scheduled_date, applied_at,
               out_date, created_by, created_at)
               VALUES (?, 'move', ?, ?, ?, 'Applied', ?, ?, ?, ?, ?, ?, ?)""",
            (asset_id, old_location, new_location, hours_mileage, new_status, mover, effective_date, now, effective_date, mover, now)
        )
        log_activity("sitepulse", "move", cur.lastrowid, "created", asset_id=asset_id,
                     field="location", old_value=old_location, new_value=new_location)
        send_whatsapp_group_message(
            f"📍 {asset_name} moved\n"
            f"{old_location or '—'} → {new_location}\n"
            f"Hours/Mileage: {hours_mileage or '—'}\n"
            f"By: {mover}",
            chat_id=whatsapp_chat_id_for_site(new_location, old_location) or ULTRAMSG_SITEPULSE_GROUP_CHAT_ID
        )

    if old["status"] != new_status:
        log_activity("sitepulse", "asset", asset_id, "updated", field="status", old_value=old["status"], new_value=new_status)
        send_whatsapp_group_message(
            f"🔧 {asset_name} status changed\n"
            f"{old['status']} → {new_status}\n"
            f"By: {mover}",
            chat_id=whatsapp_chat_id_for_site(new_location, old_location) or ULTRAMSG_SITEPULSE_GROUP_CHAT_ID
        )
    db.commit()
    flash("Asset updated.")
    return redirect(url_for("sitepulse_view_asset", asset_id=asset_id))


@app.route("/sitepulse/move/<int:move_id>/complete", methods=["POST"])
@login_required
def sitepulse_complete_scheduled_move(move_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    move = db.execute("SELECT * FROM sitepulse_usage_log WHERE id = ? AND entry_kind='move'", (move_id,)).fetchone()
    if not move:
        flash("Scheduled move not found.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    if move["move_status"] != "Scheduled":
        flash("That move has already been applied.", "error")
        return redirect(url_for("sitepulse_view_asset", asset_id=move["asset_id"]))
    _apply_move(db, move, current_user.name or current_user.email, is_auto=False)
    db.commit()
    flash(f"Marked moved to {move['to_location']}.")
    return redirect(url_for("sitepulse_view_asset", asset_id=move["asset_id"]))


@app.route("/sitepulse/move/<int:move_id>/cancel", methods=["POST"])
@login_required
def sitepulse_cancel_scheduled_move(move_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    move = db.execute("SELECT * FROM sitepulse_usage_log WHERE id = ? AND entry_kind='move'", (move_id,)).fetchone()
    if not move:
        flash("Scheduled move not found.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    if move["move_status"] != "Scheduled":
        flash("That move has already been applied.", "error")
        return redirect(url_for("sitepulse_view_asset", asset_id=move["asset_id"]))
    db.execute("DELETE FROM sitepulse_usage_log WHERE id = ?", (move_id,))
    log_activity("sitepulse", "move", move_id, "cancelled", asset_id=move["asset_id"],
                 field="location", old_value=move["from_location"], new_value=move["to_location"])
    db.commit()
    flash("Scheduled move cancelled.")
    return redirect(url_for("sitepulse_view_asset", asset_id=move["asset_id"]))


@app.route("/sitepulse/asset/<int:asset_id>/status", methods=["POST"])
@login_required
def sitepulse_quick_status(asset_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    new_status = request.form["status"]
    old = db.execute("SELECT status FROM sitepulse_assets WHERE id = ?", (asset_id,)).fetchone()["status"]
    db.execute("UPDATE sitepulse_assets SET status=?, updated_at=? WHERE id=?", (new_status, datetime.utcnow().isoformat(), asset_id))
    log_activity("sitepulse", "asset", asset_id, "updated", field="status", old_value=old, new_value=new_status)
    db.commit()
    return redirect(request.referrer or url_for("sitepulse_dashboard"))


@app.route("/sitepulse/asset/<int:asset_id>/usage/new", methods=["POST"])
@login_required
def sitepulse_new_usage(asset_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    now = datetime.utcnow().isoformat()
    photo_filename = save_photo(request.files.get("photo"))
    cur = db.execute(
        """INSERT INTO sitepulse_usage_log (asset_id, usage_type, job_name, project_id, job_address, client,
           out_date, duration_unit, return_date, notes, photo_filename, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (asset_id, request.form.get("usage_type", "Internal Job"), request.form.get("job_name", ""),
         _clean_project_id(db, request.form.get("project_id")),
         request.form.get("job_address", ""), request.form.get("client", ""), request.form.get("out_date", ""),
         request.form.get("duration_unit", ""), request.form.get("return_date", ""), request.form.get("notes", ""),
         photo_filename, now)
    )
    db.execute("UPDATE sitepulse_assets SET status='Out on Job', updated_at=? WHERE id=?", (now, asset_id))
    log_activity("sitepulse", "usage", cur.lastrowid, "created", asset_id=asset_id, new_value=request.form.get("job_name", ""))
    db.commit()
    flash("Usage logged, asset marked Out on Job.")
    return redirect(url_for("sitepulse_view_asset", asset_id=asset_id))


@app.route("/sitepulse/usage/<int:usage_id>/update", methods=["POST"])
@login_required
def sitepulse_update_usage(usage_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    entry = db.execute("SELECT * FROM sitepulse_usage_log WHERE id = ?", (usage_id,)).fetchone()
    if not entry:
        flash("Usage entry not found.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    return_date = request.form.get("return_date", "")
    new_photo = save_photo(request.files.get("photo"))
    photo_filename = new_photo if new_photo else entry["photo_filename"]
    db.execute(
        """UPDATE sitepulse_usage_log SET usage_type=?, job_name=?, project_id=?, job_address=?, client=?, out_date=?,
           duration_unit=?, return_date=?, notes=?, photo_filename=? WHERE id=?""",
        (request.form.get("usage_type", "Internal Job"), request.form.get("job_name", ""),
         _clean_project_id(db, request.form.get("project_id")),
         request.form.get("job_address", ""), request.form.get("client", ""), request.form.get("out_date", ""),
         request.form.get("duration_unit", ""), return_date, request.form.get("notes", ""), photo_filename, usage_id)
    )
    if return_date:
        db.execute("UPDATE sitepulse_assets SET status='Available', updated_at=? WHERE id=?",
                   (datetime.utcnow().isoformat(), entry["asset_id"]))
    log_activity("sitepulse", "usage", usage_id, "updated", asset_id=entry["asset_id"], new_value=request.form.get("job_name", ""))
    db.commit()
    flash("Usage entry updated.")
    return redirect(url_for("sitepulse_view_asset", asset_id=entry["asset_id"]))


@app.route("/sitepulse/usage/<int:usage_id>/delete", methods=["POST"])
@login_required
def sitepulse_delete_usage(usage_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    entry = db.execute("SELECT * FROM sitepulse_usage_log WHERE id = ?", (usage_id,)).fetchone()
    if not entry:
        flash("Usage entry not found.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    asset_id = entry["asset_id"]
    db.execute("DELETE FROM sitepulse_usage_log WHERE id = ?", (usage_id,))
    log_activity("sitepulse", "usage", usage_id, "deleted", asset_id=asset_id, old_value=entry["job_name"])
    db.commit()
    flash("Usage entry deleted.")
    return redirect(url_for("sitepulse_view_asset", asset_id=asset_id))


@app.route("/sitepulse/asset/<int:asset_id>/maintenance/new", methods=["POST"])
@login_required
def sitepulse_new_maintenance(asset_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    now = datetime.utcnow().isoformat()
    photo_filename = save_photo(request.files.get("photo"))
    cur = db.execute(
        """INSERT INTO sitepulse_maintenance_log (asset_id, entry_date, work_done, parts, hours_at_service,
           reported_by, resolved, photo_filename, created_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (asset_id, request.form.get("entry_date", date.today().isoformat()), request.form.get("work_done", ""),
         request.form.get("parts", ""), request.form.get("hours_at_service", ""),
         current_user.name or current_user.email, 1 if request.form.get("resolved") else 0, photo_filename, now)
    )
    log_activity("sitepulse", "maintenance", cur.lastrowid, "created", asset_id=asset_id, new_value=request.form.get("work_done", ""))
    db.commit()
    asset = db.execute("SELECT name, location FROM sitepulse_assets WHERE id = ?", (asset_id,)).fetchone()
    send_whatsapp_group_message(
        f"🛠️ Maintenance logged: {asset['name']}\n"
        f"Issue: {request.form.get('work_done', '') or '—'}\n"
        + (f"Parts: {request.form.get('parts')}\n" if request.form.get('parts') else "")
        + f"Reported by: {current_user.name or current_user.email}",
        chat_id=whatsapp_chat_id_for_site(asset["location"]) or ULTRAMSG_SITEPULSE_GROUP_CHAT_ID
    )
    flash("Maintenance entry logged.")
    return redirect(url_for("sitepulse_view_asset", asset_id=asset_id))


@app.route("/sitepulse/asset/<int:asset_id>/mileage/new", methods=["POST"])
@login_required
def sitepulse_new_mileage(asset_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    now = datetime.utcnow().isoformat()
    db.execute(
        """INSERT INTO sitepulse_mileage_log (asset_id, reading_date, mileage, notes, created_at)
           VALUES (?, ?, ?, ?, ?)""",
        (asset_id, request.form.get("reading_date", date.today().isoformat()),
         request.form.get("mileage", ""), request.form.get("notes", ""), now)
    )
    log_activity("sitepulse", "mileage", asset_id, "created", asset_id=asset_id,
                 new_value=request.form.get("mileage", ""))
    db.commit()
    flash("Mileage reading logged.")
    return redirect(url_for("sitepulse_view_asset", asset_id=asset_id))


@app.route("/sitepulse/maintenance/<int:entry_id>/update", methods=["POST"])
@login_required
def sitepulse_update_maintenance(entry_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    entry = db.execute("SELECT * FROM sitepulse_maintenance_log WHERE id = ?", (entry_id,)).fetchone()
    if not entry:
        flash("Maintenance entry not found.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    new_photo = save_photo(request.files.get("photo"))
    photo_filename = new_photo if new_photo else entry["photo_filename"]
    db.execute(
        """UPDATE sitepulse_maintenance_log SET entry_date=?, work_done=?, parts=?, hours_at_service=?,
           resolved=?, photo_filename=? WHERE id=?""",
        (request.form.get("entry_date", ""), request.form.get("work_done", ""), request.form.get("parts", ""),
         request.form.get("hours_at_service", ""), 1 if request.form.get("resolved") else 0, photo_filename, entry_id)
    )
    log_activity("sitepulse", "maintenance", entry_id, "updated", asset_id=entry["asset_id"], new_value=request.form.get("work_done", ""))
    db.commit()
    flash("Maintenance entry updated.")
    return redirect(url_for("sitepulse_view_asset", asset_id=entry["asset_id"]))


@app.route("/sitepulse/maintenance/<int:entry_id>/delete", methods=["POST"])
@login_required
def sitepulse_delete_maintenance(entry_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    entry = db.execute("SELECT * FROM sitepulse_maintenance_log WHERE id = ?", (entry_id,)).fetchone()
    if not entry:
        flash("Maintenance entry not found.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    asset_id = entry["asset_id"]
    db.execute("DELETE FROM sitepulse_maintenance_log WHERE id = ?", (entry_id,))
    log_activity("sitepulse", "maintenance", entry_id, "deleted", asset_id=asset_id, old_value=entry["work_done"])
    db.commit()
    flash("Maintenance entry deleted.")
    return redirect(url_for("sitepulse_view_asset", asset_id=asset_id))


@app.route("/sitepulse/activity")
@login_required
def sitepulse_activity_log():
    if not _authorized("action:activity_log:view"):
        flash("Not authorized.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    entries = db.execute(
        """SELECT a.*, ast.name AS asset_name FROM activity_log a
           LEFT JOIN sitepulse_assets ast ON a.asset_id = ast.id
           WHERE a.section = 'sitepulse' ORDER BY a.created_at DESC LIMIT 300"""
    ).fetchall()
    return render_template("sitepulse/activity_log.html", entries=entries)


@app.route("/sitepulse/asset/<int:asset_id>/activity")
@login_required
def sitepulse_asset_activity_log(asset_id):
    if not _authorized("action:activity_log:view"):
        flash("Not authorized.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    asset = db.execute("SELECT * FROM sitepulse_assets WHERE id = ?", (asset_id,)).fetchone()
    if not asset:
        flash("Asset not found.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    entries = db.execute(
        "SELECT * FROM activity_log WHERE section='sitepulse' AND asset_id = ? ORDER BY created_at DESC", (asset_id,)
    ).fetchall()
    return render_template("sitepulse/activity_log.html", entries=entries, asset=asset)


@app.route("/sitepulse/geocode")
@login_required
def sitepulse_geocode():
    if not _authorized("module:equipment_center:view"):
        flash("You don't have access to Equipment Center.", "error")
        return redirect(url_for("home"))
    address = request.args.get("address", "").strip()
    if not address:
        return {"lat": None, "lon": None}
    try:
        resp = requests.get(
            "https://geocoding.geo.census.gov/geocoder/locations/onelineaddress",
            params={"address": address, "benchmark": "2020", "format": "json"}, timeout=8)
        matches = resp.json().get("result", {}).get("addressMatches", [])
        if not matches:
            return {"lat": None, "lon": None}
        coords = matches[0]["coordinates"]
        return {"lat": coords["y"], "lon": coords["x"]}
    except Exception:
        return {"lat": None, "lon": None}


# ---- Rentals ----

@app.route("/sitepulse/rentals")
@login_required
def sitepulse_rentals_list():
    if not _authorized("module:equipment_center:view"):
        flash("You don't have access to Equipment Center.", "error")
        return redirect(url_for("home"))
    db = get_db()
    today_str = date.today().isoformat()
    show = request.args.get("show", "active")
    condition = {"returned": "returned_date IS NOT NULL AND returned_date != ''",
                 "all": "1=1"}.get(show, "returned_date IS NULL OR returned_date = ''")
    rows = db.execute(f"SELECT * FROM sitepulse_rentals WHERE {condition} ORDER BY due_date ASC, rented_date DESC").fetchall()

    rentals = []
    for r in rows:
        rd = dict(r)
        if rd["returned_date"]:
            rd["rental_status"] = "Returned"
        elif rd["due_date"] and rd["due_date"] < today_str:
            rd["rental_status"] = "Overdue"
        else:
            rd["rental_status"] = "Active"
        end = rd["returned_date"] or today_str
        try:
            days = max((date.fromisoformat(end) - date.fromisoformat(rd["rented_date"])).days + 1, 1)
            rate = float(rd["rate_amount"] or 0)
            daily_equiv = {"Weekly": rate / 7, "Monthly": rate / 30}.get(rd["rate_period"], rate)
            rd["running_cost"] = round(days * daily_equiv, 2)
            rd["days_out"] = days
        except (ValueError, TypeError):
            rd["running_cost"] = None
            rd["days_out"] = None
        rentals.append(rd)

    total_cost = round(sum(r["running_cost"] for r in rentals if r["running_cost"]), 2)
    return render_template("sitepulse/rentals.html", rentals=rentals, show=show, total_running_cost=total_cost)


@app.route("/sitepulse/rentals/new", methods=["GET", "POST"])
@login_required
def sitepulse_new_rental():
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    tracker_projects = db.execute(
        "SELECT id, name, client FROM tracker_projects WHERE status NOT IN ('Archived','Cancelled') ORDER BY name"
    ).fetchall()
    if request.method == "POST":
        now = datetime.utcnow().isoformat()
        cur = db.execute(
            """INSERT INTO sitepulse_rentals (vendor, equipment_description, job_name, project_id, rate_amount,
               rate_period, rented_date, due_date, notes, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (request.form["vendor"], request.form["equipment_description"], request.form.get("job_name", ""),
             _clean_project_id(db, request.form.get("project_id")),
             request.form.get("rate_amount", ""), request.form.get("rate_period", "Daily"),
             request.form["rented_date"], request.form.get("due_date", ""), request.form.get("notes", ""), now, now)
        )
        log_activity("sitepulse", "rental", cur.lastrowid, "created", new_value=request.form["equipment_description"])
        db.commit()
        flash("Rental logged.")
        return redirect(url_for("sitepulse_rentals_list"))
    return render_template("sitepulse/new_rental.html", today=date.today().isoformat(), tracker_projects=tracker_projects)


@app.route("/sitepulse/rentals/<int:rental_id>/update", methods=["POST"])
@login_required
def sitepulse_update_rental(rental_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    db.execute(
        """UPDATE sitepulse_rentals SET vendor=?, equipment_description=?, job_name=?, rate_amount=?,
           rate_period=?, rented_date=?, due_date=?, notes=?, updated_at=? WHERE id=?""",
        (request.form["vendor"], request.form["equipment_description"], request.form.get("job_name", ""),
         request.form.get("rate_amount", ""), request.form.get("rate_period", "Daily"),
         request.form["rented_date"], request.form.get("due_date", ""), request.form.get("notes", ""),
         datetime.utcnow().isoformat(), rental_id)
    )
    log_activity("sitepulse", "rental", rental_id, "updated", new_value=request.form["equipment_description"])
    db.commit()
    flash("Rental updated.")
    return redirect(url_for("sitepulse_rentals_list"))


@app.route("/sitepulse/rentals/<int:rental_id>/return", methods=["POST"])
@login_required
def sitepulse_return_rental(rental_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    returned_date = request.form.get("returned_date") or date.today().isoformat()
    db.execute("UPDATE sitepulse_rentals SET returned_date=?, updated_at=? WHERE id=?",
               (returned_date, datetime.utcnow().isoformat(), rental_id))
    log_activity("sitepulse", "rental", rental_id, "returned", field="returned_date", new_value=returned_date)
    db.commit()
    flash("Rental marked returned.")
    return redirect(url_for("sitepulse_rentals_list"))


@app.route("/sitepulse/rentals/<int:rental_id>/delete", methods=["POST"])
@login_required
def sitepulse_delete_rental(rental_id):
    if not _authorized("action:equipment_center:manage"):
        flash("You don't have permission to make changes in Equipment Center.", "error")
        return redirect(url_for("sitepulse_dashboard"))
    db = get_db()
    r = db.execute("SELECT * FROM sitepulse_rentals WHERE id = ?", (rental_id,)).fetchone()
    db.execute("DELETE FROM sitepulse_rentals WHERE id = ?", (rental_id,))
    log_activity("sitepulse", "rental", rental_id, "deleted", old_value=r["equipment_description"] if r else None)
    db.commit()
    flash("Rental deleted.")
    return redirect(url_for("sitepulse_rentals_list"))


# ---------------------------------------------------------------------------
# Site Inventory -- Concrete Requests + Materials
# ---------------------------------------------------------------------------

@app.route("/inventory/")
@login_required
def inventory_home():
    if not _authorized("module:sitepulse:view"):
        flash("You don't have access to SitePulse.", "error")
        return redirect(url_for("home"))
    return render_template("inventory/home.html")


@app.route("/inventory/materials")
@login_required
def inventory_materials_list():
    if not _authorized("module:sitepulse:view"):
        flash("You don't have access to SitePulse.", "error")
        return redirect(url_for("home"))
    db = get_db()
    search = request.args.get("q", "").strip()
    if search:
        like = f"%{search}%"
        rows = db.execute(
            "SELECT * FROM inventory_materials WHERE item_name LIKE ? OR site LIKE ? OR shelf_location LIKE ? "
            "ORDER BY site, item_name",
            (like, like, like)).fetchall()
    else:
        rows = db.execute("SELECT * FROM inventory_materials ORDER BY site, item_name").fetchall()
    return render_template("inventory/materials.html", materials=rows, search=search)


@app.route("/inventory/materials/new", methods=["GET", "POST"])
@login_required
def inventory_new_material():
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    if request.method == "POST":
        db = get_db()
        now = datetime.utcnow().isoformat()
        cur = db.execute(
            """INSERT INTO inventory_materials (item_name, site, quantity, unit, shelf_location, notes,
               created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (request.form["item_name"], request.form["site"], request.form.get("quantity", ""),
             request.form.get("unit", ""), request.form.get("shelf_location", ""),
             request.form.get("notes", ""), now, now)
        )
        log_activity("inventory", "material", cur.lastrowid, "created", new_value=request.form["item_name"])
        db.commit()
        flash("Material added.")
        return redirect(url_for("inventory_materials_list"))
    db = get_db()
    sites = db.execute("SELECT DISTINCT site FROM inventory_materials ORDER BY site").fetchall()
    return render_template("inventory/new_material.html", sites=[s["site"] for s in sites])


@app.route("/inventory/materials/<int:material_id>/delete", methods=["POST"])
@login_required
def inventory_delete_material(material_id):
    if not _authorized("action:sitepulse:manage_inventory"):
        flash("Not authorized.", "error")
        return redirect(url_for("inventory_materials_list"))
    db = get_db()
    m = db.execute("SELECT * FROM inventory_materials WHERE id = ?", (material_id,)).fetchone()
    db.execute("DELETE FROM inventory_materials WHERE id = ?", (material_id,))
    log_activity("inventory", "material", material_id, "deleted", old_value=m["item_name"] if m else None)
    db.commit()
    flash("Material deleted.")
    return redirect(url_for("inventory_materials_list"))


def send_due_concrete_reminders():
    """The day before a scheduled pour, send the full order details (the
    same notification the old immediate-on-schedule message used to be).
    No cron here -- same pattern as apply_due_scheduled_moves: check for
    anything due whenever the concrete list page loads, and send it once.
    """
    db = get_db()
    tomorrow = (date.today() + timedelta(days=1)).isoformat()
    due = db.execute(
        "SELECT * FROM inventory_concrete_requests WHERE status = 'Scheduled' AND pour_date = ? "
        "AND (full_reminder_sent_at IS NULL OR full_reminder_sent_at = '')",
        (tomorrow,)
    ).fetchall()
    for r in due:
        order_chat_id = whatsapp_chat_id_for_site(r["project"], r["job_site_address"])
        send_whatsapp_group_message(
            "📋 Tomorrow's concrete order:\n\n" + build_concrete_order_notification(r),
            chat_id=order_chat_id
        )
        db.execute(
            "UPDATE inventory_concrete_requests SET full_reminder_sent_at = ? WHERE id = ?",
            (datetime.utcnow().isoformat(), r["id"])
        )
    if due:
        db.commit()


@app.route("/inventory/concrete")
@login_required
def inventory_concrete_list():
    if not _authorized("module:sitepulse:view"):
        flash("You don't have access to SitePulse.", "error")
        return redirect(url_for("home"))
    send_due_concrete_reminders()
    db = get_db()

    project_filter = request.args.get("project", "")
    status_filter = request.args.get("status", "")
    pour_date_filter = request.args.get("pour_date", "")

    conditions, params = [], []
    if project_filter:
        conditions.append("project = ?")
        params.append(project_filter)
    if status_filter:
        conditions.append("status = ?")
        params.append(status_filter)
    if pour_date_filter:
        conditions.append("pour_date = ?")
        params.append(pour_date_filter)
    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""

    rows = db.execute(f"SELECT * FROM inventory_concrete_requests {where} ORDER BY pour_date DESC", params).fetchall()
    projects = [p["project"] for p in db.execute(
        "SELECT DISTINCT project FROM inventory_concrete_requests WHERE project IS NOT NULL AND project != '' ORDER BY project"
    ).fetchall()]

    return render_template(
        "inventory/concrete_requests.html", requests=rows, projects=projects,
        status_options=CONCRETE_STATUS_OPTIONS, project_filter=project_filter,
        status_filter=status_filter, pour_date_filter=pour_date_filter
    )


def create_concrete_request(fields, requested_by):
    """Shared insert+notify logic for a new concrete request -- used by
    both the web form and the voice assistant, so both paths behave
    identically (same notification, same defaults, same activity log).
    `fields` is a dict of form-field-name -> value; missing keys default
    the same way request.form.get(..., "") would.
    """
    db = get_db()
    now = datetime.utcnow().isoformat()

    def f(key, default=""):
        return fields.get(key) or default

    # Defense in depth: even if the client-side toggle didn't clear it
    # (JS disabled, browser autofill restoring a stale value after the
    # page loaded), never actually store a pump size/time for a request
    # that doesn't have a real pump selected.
    pump_size = f("pump_size") if f("pump_type") in ("Ground Pump", "Overhead Pump") else ""
    pump_arrival_time = f("pump_arrival_time") if f("pump_type") in ("Ground Pump", "Overhead Pump") else ""
    drilling_time = f("drilling_time") if f("drilling_required") == "Yes" else ""

    cur = db.execute(
        """INSERT INTO inventory_concrete_requests (project, project_id, job_site_address, area_description,
           pour_date, pour_time, mix_design_psi, mix_slump, concrete_amount, truck_spacing,
           pump_type, pump_size, pump_arrival_time, lab_required, lab_time, drilling_required, drilling_time,
           requested_by, requested_signature, requested_date, ordered_by, ordered_signature,
           ordered_date, status, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (f("project"), fields.get("project_id") or None, f("job_site_address"), f("area_description"), f("pour_date"), f("pour_time"),
         f("mix_design_psi"), f("mix_slump"), f("concrete_amount"), f("truck_spacing"),
         f("pump_type"), pump_size, pump_arrival_time, f("lab_required", "No"), f("lab_time"),
         f("drilling_required", "No"), drilling_time, requested_by, f("requested_signature"),
         f("requested_date"), f("ordered_by"), f("ordered_signature"), f("ordered_date"),
         "Submitted", now, now)
    )
    log_activity("inventory", "concrete_request", cur.lastrowid, "created", new_value=f("project"))
    db.commit()
    send_whatsapp_group_message(
        f"🧱 New concrete request submitted\n"
        f"Project: {f('project') or '—'}\n"
        f"Pour date: {friendly_date(f('pour_date'))}{' at ' + friendly_time(f('pour_time')) if f('pour_time') else ''}\n"
        f"Amount: {f('concrete_amount') or '—'}\n"
        f"Requested by: {requested_by}",
        chat_id=whatsapp_chat_id_for_site(f("project"), f("job_site_address"))
    )
    return cur.lastrowid


CONCRETE_REQUEST_REQUIRED_FIELDS = [
    "project", "job_site_address", "area_description", "pour_date", "pour_time",
    "mix_design_psi", "mix_slump", "concrete_amount", "truck_spacing",
    "pump_type", "pump_size", "pump_arrival_time", "lab_required", "lab_time",
    "drilling_required", "drilling_time", "requested_date",
]


@app.route("/inventory/concrete/new", methods=["GET", "POST"])
@login_required
def inventory_new_concrete():
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    db = get_db()
    tracker_projects = db.execute(
        "SELECT id, name, client FROM tracker_projects WHERE status NOT IN ('Archived','Cancelled') ORDER BY name"
    ).fetchall()
    if request.method == "POST":
        needs_pump = request.form.get("pump_type") in ("Ground Pump", "Overhead Pump")
        needs_lab = request.form.get("lab_required") == "Yes"
        needs_drilling = request.form.get("drilling_required") == "Yes"
        required_fields = [
            f for f in CONCRETE_REQUEST_REQUIRED_FIELDS
            if (needs_pump or f not in ("pump_size", "pump_arrival_time"))
            and (needs_lab or f != "lab_time")
            and (needs_drilling or f != "drilling_time")
        ]
        missing = [f for f in required_fields if not request.form.get(f, "").strip()]
        if missing:
            flash("Please fill in every field on the form before submitting.", "error")
            return render_template("inventory/new_concrete_request.html", today=date.today().isoformat(), form=request.form, tracker_projects=tracker_projects)
        create_concrete_request(request.form.to_dict(), current_user.name or current_user.email)
        flash("Concrete request submitted.")
        return redirect(url_for("inventory_concrete_list"))
    return render_template("inventory/new_concrete_request.html", today=date.today().isoformat(), tracker_projects=tracker_projects)


def _time_minus_hours(hhmm, hours):
    """'08:00' minus 1 hour -> '07:00'. Returns '' if hhmm is blank/unparseable."""
    if not hhmm:
        return ""
    try:
        t = datetime.strptime(hhmm, "%H:%M")
        t -= timedelta(hours=hours)
        return t.strftime("%H:%M")
    except ValueError:
        return ""


def build_concrete_order_notification(r):
    """Build the plain-text order notification in the format procurement
    texts/emails out once an order is placed -- e.g.:
    'Concrete scheduled tomorrow (2nd road pour) 07/21/2026 at 8:00 AM'
    """
    if not r["pour_date"]:
        return ""
    try:
        pour_dt = datetime.strptime(r["pour_date"], "%Y-%m-%d").date()
        date_display = pour_dt.strftime("%m/%d/%Y")
        delta = (pour_dt - date.today()).days
        when = "today" if delta == 0 else ("tomorrow" if delta == 1 else pour_dt.strftime("%A"))
    except ValueError:
        date_display = r["pour_date"]
        when = ""

    def fmt_time(t):
        if not t:
            return None
        try:
            return datetime.strptime(t, "%H:%M").strftime("%-I:%M %p")
        except ValueError:
            return t

    pour_time = fmt_time(r["pour_time"])
    # The header announces the confirmed/scheduled delivery, so it should
    # prefer the confirmed arrival time over the originally requested
    # pour_time when the two differ (e.g. requested 7:00 AM, confirmed
    # slot 8:00 AM) -- same precedence the concrete-company line below
    # already uses. `pour_time` itself is left untouched: it's still the
    # requested-time fallback used further down.
    scheduled_time_display = fmt_time(r["concrete_arrival_time"]) or pour_time
    lines = []
    header = f"Concrete scheduled {when}".strip()
    if r["area_description"]:
        header += f" ({r['area_description']})"
    header += f" {date_display}"
    if scheduled_time_display:
        header += f" at {scheduled_time_display}"
    lines.append(header)

    amount_line = " ".join(x for x in [r["concrete_amount"], f"plus {r['mix_design_psi']} PSI" if r["mix_design_psi"] else ""] if x)
    if amount_line:
        lines.append(amount_line)

    if r["pump_company"] or r["pump_size"]:
        pump_time = fmt_time(r["pump_arrival_time"])
        pump_line = r["pump_type"] if r["pump_type"] else "Pump"
        if r["pump_company"]:
            pump_line += f"-{r['pump_company']}"
        if r["pump_company_phone"]:
            pump_line += f" #{r['pump_company_phone']}"
        if pump_time:
            pump_line += f" @{pump_time}"
        lines.append(pump_line)

    if r["concrete_company"]:
        concrete_time = fmt_time(r["concrete_arrival_time"]) or pour_time
        concrete_line = r["concrete_company"]
        if concrete_time:
            concrete_line += f" @{concrete_time}"
        if r["concrete_company_phone"]:
            concrete_line += f" #{r['concrete_company_phone']}"
        lines.append(concrete_line)

    if r["lab_required"] == "Yes" and (r["lab_company"] or r["lab_time"]):
        lab_time = fmt_time(r["lab_time"])
        lab_line = r["lab_company"] or "Lab"
        if lab_time:
            lab_line += f" at {lab_time}"
        lines.append(lab_line)

    if r["drilling_required"] == "Yes" and (r["drilling_company"] or r["drilling_time"]):
        drill_time = fmt_time(r["drilling_time"])
        drill_line = r["drilling_company"] or "Drilling company"
        if r["drilling_company_phone"]:
            drill_line += f" #{r['drilling_company_phone']}"
        if drill_time:
            drill_line += f" at {drill_time}"
        lines.append(drill_line)

    return "\n\n".join(lines)


@app.route("/inventory/concrete/<int:request_id>")
@login_required
def inventory_view_concrete(request_id):
    if not _authorized("module:sitepulse:view"):
        flash("You don't have access to SitePulse.", "error")
        return redirect(url_for("home"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_concrete_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_concrete_list"))
    notification = build_concrete_order_notification(r) if r["ordered_by"] else None
    return render_template("inventory/concrete_request_detail.html", r=r, notification=notification)


@app.route("/inventory/concrete/<int:request_id>/edit", methods=["GET", "POST"])
@login_required
def inventory_edit_concrete(request_id):
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_concrete_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_concrete_list"))
    if request.method == "POST":
        # Same defense-in-depth as create: never store a pump size/time
        # unless a real pump is actually selected.
        pump_type_val = request.form.get("pump_type", "")
        needs_pump = pump_type_val in ("Ground Pump", "Overhead Pump")
        pump_size_val = request.form.get("pump_size", "") if needs_pump else ""
        pump_arrival_val = request.form.get("pump_arrival_time", "") if needs_pump else ""
        drilling_required_val = request.form.get("drilling_required", "No")
        drilling_time_val = request.form.get("drilling_time", "") if drilling_required_val == "Yes" else ""
        db.execute(
            """UPDATE inventory_concrete_requests SET project=?, project_id=?, job_site_address=?, area_description=?,
               pour_date=?, pour_time=?, mix_design_psi=?, mix_slump=?, concrete_amount=?, truck_spacing=?,
               pump_type=?, pump_size=?, pump_arrival_time=?, lab_required=?, lab_time=?, drilling_required=?,
               drilling_time=?, updated_at=?
               WHERE id=?""",
            (request.form.get("project", ""), _clean_project_id(db, request.form.get("project_id")),
             request.form.get("job_site_address", ""),
             request.form.get("area_description", ""), request.form["pour_date"], request.form.get("pour_time", ""),
             request.form.get("mix_design_psi", ""), request.form.get("mix_slump", ""),
             request.form.get("concrete_amount", ""), request.form.get("truck_spacing", ""),
             pump_type_val, pump_size_val, pump_arrival_val,
             request.form.get("lab_required", "No"), request.form.get("lab_time", ""),
             drilling_required_val, drilling_time_val,
             datetime.utcnow().isoformat(), request_id)
        )
        log_activity("inventory", "concrete_request", request_id, "updated", new_value=request.form.get("project", ""))
        db.commit()
        flash("Concrete request updated.")
        return redirect(url_for("inventory_view_concrete", request_id=request_id))
    tracker_projects = db.execute(
        "SELECT id, name, client FROM tracker_projects WHERE status NOT IN ('Archived','Cancelled') ORDER BY name"
    ).fetchall()
    return render_template("inventory/edit_concrete_request.html", r=r, today=date.today().isoformat(), tracker_projects=tracker_projects)


@app.route("/inventory/concrete/<int:request_id>/order", methods=["GET", "POST"])
@login_required
def inventory_place_concrete_order(request_id):
    if not is_procurement():
        flash("Only procurement (Ayoub, Rebecca, or Marilu) can place a concrete order.", "error")
        return redirect(url_for("inventory_view_concrete", request_id=request_id))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_concrete_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_concrete_list"))
    if request.method == "POST":
        now = datetime.utcnow().isoformat()
        db.execute(
            """UPDATE inventory_concrete_requests SET
               concrete_company=?, concrete_company_phone=?, concrete_arrival_time=?, pump_company=?, pump_company_phone=?,
               pump_arrival_time=?, lab_company=?, lab_time=?, drilling_company=?, drilling_company_phone=?,
               drilling_time=?, ordered_by=?, ordered_date=?, status='Scheduled', updated_at=?
               WHERE id=?""",
            (request.form.get("concrete_company", ""), request.form.get("concrete_company_phone", ""),
             request.form.get("concrete_arrival_time", ""), request.form.get("pump_company", ""), request.form.get("pump_company_phone", ""),
             request.form.get("pump_arrival_time", ""), request.form.get("lab_company", ""),
             request.form.get("lab_time", ""), request.form.get("drilling_company", ""),
             request.form.get("drilling_company_phone", ""), request.form.get("drilling_time", ""),
             current_user.name or current_user.email, date.today().isoformat(), now, request_id)
        )
        log_activity("inventory", "concrete_request", request_id, "updated", field="status",
                     old_value=r["status"], new_value="Scheduled")
        db.commit()
        updated_r = db.execute("SELECT * FROM inventory_concrete_requests WHERE id = ?", (request_id,)).fetchone()
        order_chat_id = whatsapp_chat_id_for_site(updated_r["project"], updated_r["job_site_address"])
        pour_date_display = friendly_date(updated_r["pour_date"])
        # Use the confirmed/scheduled arrival time (just captured on this
        # very form) for the notification announcing the schedule, not the
        # original requested pour_time -- they can legitimately differ
        # (e.g. requested 7:00 AM, confirmed slot 8:00 AM). pour_time
        # remains the requested-time record for audit/history and is
        # intentionally left untouched by this route. Falls back to
        # pour_time only if no confirmed arrival time was set.
        scheduled_time = updated_r["concrete_arrival_time"] or updated_r["pour_time"]
        pour_time_display = " at " + friendly_time(scheduled_time) if scheduled_time else ""
        send_whatsapp_group_message(
            f"Concrete Scheduled for {pour_date_display}{pour_time_display}",
            chat_id=order_chat_id
        )
        flash("Order placed and marked Scheduled.")
        return redirect(url_for("inventory_view_concrete", request_id=request_id))
    return render_template(
        "inventory/place_concrete_order.html", r=r, today=date.today().isoformat(),
        default_pump_arrival=_time_minus_hours(r["pour_time"], 1) if not r["pump_arrival_time"] else "",
        default_drilling_time=_time_minus_hours(r["pour_time"], 1) if not r["drilling_time"] else "",
        default_lab_time=r["pour_time"] if not r["lab_time"] else "",
    )


@app.route("/inventory/concrete/<int:request_id>/status", methods=["POST"])
@login_required
def inventory_update_concrete_status(request_id):
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_concrete_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_concrete_list"))
    new_status = request.form["status"]
    if new_status == "Scheduled" and not is_procurement():
        flash("Only procurement can mark a concrete request Scheduled -- use Place Order.", "error")
        return redirect(url_for("inventory_view_concrete", request_id=request_id))
    db.execute("UPDATE inventory_concrete_requests SET status = ?, updated_at = ? WHERE id = ?",
               (new_status, datetime.utcnow().isoformat(), request_id))
    log_activity("inventory", "concrete_request", request_id, "updated", field="status",
                 old_value=r["status"], new_value=new_status)
    db.commit()
    flash(f"Marked as {new_status}.")
    return redirect(url_for("inventory_view_concrete", request_id=request_id))


@app.route("/inventory/concrete/<int:request_id>/delete", methods=["POST"])
@login_required
def inventory_delete_concrete(request_id):
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_concrete_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_concrete_list"))
    db.execute("DELETE FROM inventory_concrete_requests WHERE id = ?", (request_id,))
    log_activity("inventory", "concrete_request", request_id, "deleted", old_value=r["project"])
    db.commit()
    flash("Concrete request deleted.")
    return redirect(url_for("inventory_concrete_list"))


@app.route("/inventory/purchase")
@login_required
def inventory_purchase_list():
    if not _authorized("module:sitepulse:view"):
        flash("You don't have access to SitePulse.", "error")
        return redirect(url_for("home"))
    db = get_db()
    rows = db.execute("SELECT * FROM inventory_purchase_requests ORDER BY request_date DESC").fetchall()
    return render_template("inventory/purchase_requests.html", requests=rows)


def _generate_pr_number(db):
    """Auto-generate a purchase request number in the same MMDDYYYY
    format already used historically, with a -2/-3 suffix if more than
    one request happens to land on the same day."""
    base = date.today().strftime("%m%d%Y")
    existing = {
        row["pr_number"] for row in db.execute(
            "SELECT pr_number FROM inventory_purchase_requests WHERE pr_number LIKE ?", (f"{base}%",)
        ).fetchall()
    }
    if base not in existing:
        return base
    n = 2
    while f"{base}-{n}" in existing:
        n += 1
    return f"{base}-{n}"


@app.route("/inventory/purchase/new", methods=["GET", "POST"])
@login_required
def inventory_new_purchase():
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    db = get_db()
    tracker_projects = db.execute(
        "SELECT id, name, client FROM tracker_projects WHERE status NOT IN ('Archived','Cancelled') ORDER BY name"
    ).fetchall()
    if request.method == "POST":
        db = get_db()
        now = datetime.utcnow().isoformat()
        requestor_display = current_user.name or current_user.email

        cur = db.execute(
            """INSERT INTO inventory_purchase_requests (pr_number, request_date, job_name, project_id,
               location_description, requested_by, needed_on, source_of_supply,
               requestor_signature, requestor_date, status, created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (_generate_pr_number(db), request.form["request_date"], request.form.get("job_name", ""),
             request.form.get("project_id") or None,
             request.form.get("location_description", ""), requestor_display,
             request.form.get("needed_on", ""), request.form.get("source_of_supply", ""),
             request.form.get("requestor_signature", ""), request.form.get("requestor_date", ""),
             "Submitted", now, now)
        )
        pr_id = cur.lastrowid
        items = request.form.getlist("item[]")
        descriptions = request.form.getlist("description[]")
        suppliers = request.form.getlist("supplier[]")
        qtys = request.form.getlist("qty[]")
        units = request.form.getlist("unit[]")
        for item, desc, sup, qty, unit in zip(items, descriptions, suppliers, qtys, units):
            if item.strip() or desc.strip():
                db.execute(
                    "INSERT INTO inventory_purchase_request_items (purchase_request_id, item, description, supplier, qty, unit) VALUES (?,?,?,?,?,?)",
                    (pr_id, item, desc, sup, qty, unit)
                )
        log_activity("inventory", "purchase_request", pr_id, "created", new_value=request.form.get("job_name", ""))
        db.commit()

        # If a procurement person is the one actually logged in and
        # submitting, skip the submission notification -- the group only
        # needs to hear about it once the order is actually placed, which
        # already sends its own notification further down the flow.
        if current_user.email not in PROCUREMENT_EMAILS:
            item_summary = "; ".join(i.strip() for i in items if i.strip())[:200]
            send_whatsapp_group_message(
                f"🛒 New purchase request submitted\n"
                f"Job: {request.form.get('job_name', '') or '—'}\n"
                f"Needed by: {friendly_date(request.form.get('needed_on', '')) or '—'}\n"
                f"Items: {item_summary or '—'}\n"
                f"Requested by: {requestor_display}",
                chat_id=whatsapp_chat_id_for_site(request.form.get("job_name", ""), request.form.get("location_description", ""))
            )
        flash("Purchase request submitted.")
        return redirect(url_for("inventory_purchase_list"))

    return render_template("inventory/new_purchase_request.html", today=date.today().isoformat(), tracker_projects=tracker_projects)


@app.route("/inventory/purchase/<int:request_id>")
@login_required
def inventory_view_purchase(request_id):
    if not _authorized("module:sitepulse:view"):
        flash("You don't have access to SitePulse.", "error")
        return redirect(url_for("home"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_purchase_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_purchase_list"))
    items = db.execute("SELECT * FROM inventory_purchase_request_items WHERE purchase_request_id = ?", (request_id,)).fetchall()
    return render_template("inventory/purchase_request_detail.html", r=r, items=items)


@app.route("/inventory/activity")
@login_required
def inventory_activity_log():
    if not _authorized("action:activity_log:view"):
        flash("Not authorized.", "error")
        return redirect(url_for("inventory_materials_list"))
    db = get_db()
    entries = db.execute(
        "SELECT * FROM activity_log WHERE section = 'inventory' ORDER BY created_at DESC LIMIT 300"
    ).fetchall()
    return render_template("inventory/activity_log.html", entries=entries)


@app.route("/inventory/concrete/<int:request_id>/activity")
@login_required
def inventory_concrete_activity_log(request_id):
    if not _authorized("action:activity_log:view"):
        flash("Not authorized.", "error")
        return redirect(url_for("inventory_concrete_list"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_concrete_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_concrete_list"))
    entries = db.execute(
        "SELECT * FROM activity_log WHERE section='inventory' AND entity_type='concrete_request' AND entity_id = ? ORDER BY created_at DESC",
        (request_id,)
    ).fetchall()
    return render_template("inventory/activity_log.html", entries=entries, record_name=r["project"])


@app.route("/inventory/purchase/<int:request_id>/activity")
@login_required
def inventory_purchase_activity_log(request_id):
    if not _authorized("action:activity_log:view"):
        flash("Not authorized.", "error")
        return redirect(url_for("inventory_purchase_list"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_purchase_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_purchase_list"))
    entries = db.execute(
        "SELECT * FROM activity_log WHERE section='inventory' AND entity_type='purchase_request' AND entity_id = ? ORDER BY created_at DESC",
        (request_id,)
    ).fetchall()
    return render_template("inventory/activity_log.html", entries=entries, record_name=r["job_name"])


@app.route("/inventory/purchase/<int:request_id>/edit", methods=["GET", "POST"])
@login_required
def inventory_edit_purchase(request_id):
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_purchase_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_purchase_list"))
    if request.method == "POST":
        db.execute(
            """UPDATE inventory_purchase_requests SET pr_number=?, request_date=?, job_name=?, project_id=?,
               location_description=?, needed_on=?, source_of_supply=?, updated_at=? WHERE id=?""",
            (request.form.get("pr_number", ""), request.form["request_date"], request.form.get("job_name", ""),
             _clean_project_id(db, request.form.get("project_id")),
             request.form.get("location_description", ""), request.form.get("needed_on", ""),
             request.form.get("source_of_supply", ""), datetime.utcnow().isoformat(), request_id)
        )
        # Items: simplest reliable approach is replace-all -- delete the old
        # lines and insert whatever's on the form now, same as the create route.
        db.execute("DELETE FROM inventory_purchase_request_items WHERE purchase_request_id = ?", (request_id,))
        items = request.form.getlist("item[]")
        descriptions = request.form.getlist("description[]")
        suppliers = request.form.getlist("supplier[]")
        qtys = request.form.getlist("qty[]")
        units = request.form.getlist("unit[]")
        for item, desc, sup, qty, unit in zip(items, descriptions, suppliers, qtys, units):
            if item.strip() or desc.strip():
                db.execute(
                    "INSERT INTO inventory_purchase_request_items (purchase_request_id, item, description, supplier, qty, unit) VALUES (?,?,?,?,?,?)",
                    (request_id, item, desc, sup, qty, unit)
                )
        log_activity("inventory", "purchase_request", request_id, "updated", new_value=request.form.get("job_name", ""))
        db.commit()
        flash("Purchase request updated.")
        return redirect(url_for("inventory_view_purchase", request_id=request_id))
    existing_items = db.execute("SELECT * FROM inventory_purchase_request_items WHERE purchase_request_id = ?", (request_id,)).fetchall()
    tracker_projects = db.execute(
        "SELECT id, name, client FROM tracker_projects WHERE status NOT IN ('Archived','Cancelled') ORDER BY name"
    ).fetchall()
    return render_template("inventory/edit_purchase_request.html", r=r, items=existing_items, tracker_projects=tracker_projects)


@app.route("/inventory/purchase/<int:request_id>/status", methods=["POST"])
@login_required
def inventory_update_purchase_status(request_id):
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_purchase_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_purchase_list"))
    new_status = request.form["status"]
    if new_status == "Scheduled" and not is_procurement():
        flash("Only procurement can mark a purchase request Scheduled -- use Place Order.", "error")
        return redirect(url_for("inventory_view_purchase", request_id=request_id))
    db.execute("UPDATE inventory_purchase_requests SET status = ?, updated_at = ? WHERE id = ?",
               (new_status, datetime.utcnow().isoformat(), request_id))
    log_activity("inventory", "purchase_request", request_id, "updated", field="status",
                 old_value=r["status"], new_value=new_status)
    db.commit()
    flash(f"Marked as {new_status}.")
    return redirect(url_for("inventory_view_purchase", request_id=request_id))


@app.route("/inventory/purchase/<int:request_id>/order", methods=["GET", "POST"])
@login_required
def inventory_place_purchase_order(request_id):
    if not is_procurement():
        flash("Only procurement (Ayoub, Rebecca, or Marilu) can place this order.", "error")
        return redirect(url_for("inventory_view_purchase", request_id=request_id))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_purchase_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_purchase_list"))
    if request.method == "POST":
        db.execute(
            """UPDATE inventory_purchase_requests SET vendor_company=?, vendor_company_phone=?,
               ordered_by=?, ordered_date=?, expected_delivery_date=?, status='Scheduled', updated_at=? WHERE id=?""",
            (request.form.get("vendor_company", ""), request.form.get("vendor_company_phone", ""),
             current_user.name or current_user.email, date.today().isoformat(),
             request.form.get("expected_delivery_date", ""),
             datetime.utcnow().isoformat(), request_id)
        )
        log_activity("inventory", "purchase_request", request_id, "updated", field="status",
                     old_value=r["status"], new_value="Scheduled")
        db.commit()
        updated_r = db.execute("SELECT * FROM inventory_purchase_requests WHERE id = ?", (request_id,)).fetchone()
        order_chat_id = whatsapp_chat_id_for_site(updated_r["job_name"], updated_r["location_description"])
        send_whatsapp_group_message(
            f"✅ Purchase order placed by {current_user.name or current_user.email}\n"
            f"Job: {updated_r['job_name'] or '—'}\n"
            f"Vendor: {updated_r['vendor_company'] or '—'}"
            + (f" #{updated_r['vendor_company_phone']}" if updated_r['vendor_company_phone'] else ""),
            chat_id=order_chat_id
        )
        try:
            items = db.execute(
                "SELECT * FROM inventory_purchase_request_items WHERE purchase_request_id = ?", (request_id,)
            ).fetchall()
            pdf_bytes = build_purchase_order_pdf(updated_r, items)
            send_whatsapp_document(
                pdf_bytes, f"Purchase_Order_{request_id}.pdf",
                chat_id=order_chat_id, caption="Purchase order details"
            )
        except Exception as e:
            print(f"[whatsapp] failed to build/send purchase order PDF: {e}")
        flash("Order placed and marked Scheduled.")
        return redirect(url_for("inventory_view_purchase", request_id=request_id))
    return render_template("inventory/place_purchase_order.html", r=r, today=date.today().isoformat())


@app.route("/inventory/purchase/<int:request_id>/delete", methods=["POST"])
@login_required
def inventory_delete_purchase(request_id):
    if not _authorized("action:sitepulse:manage"):
        flash("You don't have permission to make changes in SitePulse.", "error")
        return redirect(url_for("inventory_home"))
    db = get_db()
    r = db.execute("SELECT * FROM inventory_purchase_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("inventory_purchase_list"))
    db.execute("DELETE FROM inventory_purchase_request_items WHERE purchase_request_id = ?", (request_id,))
    db.execute("DELETE FROM inventory_purchase_requests WHERE id = ?", (request_id,))
    log_activity("inventory", "purchase_request", request_id, "deleted", old_value=r["job_name"])
    db.commit()
    flash("Purchase request deleted.")
    return redirect(url_for("inventory_purchase_list"))


# ---------------------------------------------------------------------------
# Bid Tracker -- Projects + Quotes (core of Command Center)
# ---------------------------------------------------------------------------

TR_STATUS_OPTIONS = ["In Progress", "Submitted", "Awarded", "Unmeant", "On Hold", "Cancelled", "Pending", "Archived"]
TR_STATUS_BADGE_CLASS = {
    "In Progress": "status-inprogress", "Submitted": "status-submitted",
    "Awarded": "status-awarded", "Unmeant": "status-lost",
    "On Hold": "status-onhold", "Cancelled": "status-cancelled", "Pending": "status-pending",
    "Archived": "status-cancelled",
}


MARKDOWN_ALLOWED_TAGS = [
    "p", "br", "strong", "em", "b", "i", "u", "s", "del", "sup", "sub",
    "ul", "ol", "li", "a", "code", "pre", "blockquote",
    "h1", "h2", "h3", "h4", "h5", "h6", "hr",
    "table", "thead", "tbody", "tr", "th", "td",
]
MARKDOWN_ALLOWED_ATTRS = {
    "a": ["href", "title", "rel"],
    "th": ["align"],
    "td": ["align"],
}
MARKDOWN_ALLOWED_PROTOCOLS = ["http", "https", "mailto"]


@app.template_filter("markdown")
def tr_markdown_filter(text):
    """Renders Markdown to HTML, then sanitizes the result through bleach
    (a maintained HTML-sanitization library, not a homemade regex filter)
    before it's ever marked |safe in a template. python-markdown passes
    raw HTML in the source straight through unescaped -- without this
    sanitization step, a value like '<script>...</script>' or
    '<img onerror=...>' stored in a field such as a quote's RFQ/follow-up
    email text would execute as live HTML for anyone who views that page
    (stored XSS). Only a fixed allowlist of real Markdown-output tags/
    attributes survives; everything else (script, style, iframe, event
    handler attributes, javascript: URLs, etc.) is stripped, not merely
    escaped for display."""
    if not text:
        return ""
    rendered = md_lib.markdown(text, extensions=["extra"])
    return bleach.clean(
        rendered,
        tags=MARKDOWN_ALLOWED_TAGS,
        attributes=MARKDOWN_ALLOWED_ATTRS,
        protocols=MARKDOWN_ALLOWED_PROTOCOLS,
        strip=True,
    )


@app.template_filter("statusclass")
def tr_statusclass_filter(status):
    return TR_STATUS_BADGE_CLASS.get(status, "status-pending")


@app.template_filter("daysleft")
def tr_daysleft_filter(due_date_str):
    if not due_date_str:
        return None
    try:
        due = datetime.strptime(due_date_str, "%Y-%m-%d").date()
        return (due - date.today()).days
    except ValueError:
        return None


def tr_format_currency(value):
    if not value:
        return value
    value = value.strip()
    if not value:
        return value
    cleaned = value.replace("$", "").replace(",", "").strip()
    try:
        num = float(cleaned)
        if "." in cleaned:
            return "${:,.2f}".format(num)
        return "${:,.0f}".format(num)
    except ValueError:
        return value


def tr_format_phone(value):
    if not value:
        return value
    digits = "".join(c for c in value if c.isdigit())
    if len(digits) == 10:
        return f"({digits[0:3]}) {digits[3:6]}-{digits[6:10]}"
    if len(digits) == 11 and digits[0] == "1":
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:11]}"
    return value


def gather_business_snapshot():
    """Pull a compact, current snapshot of the business for the voice
    assistant to answer questions against -- equipment status, open
    requests, and active bids. Kept short on purpose: this gets stuffed
    into every assistant prompt, so it's a summary, not a full export.
    """
    db = get_db()
    lines = []

    assets = db.execute("SELECT name, status, location FROM sitepulse_assets ORDER BY name").fetchall()
    status_counts = {}
    for a in assets:
        status_counts[a["status"]] = status_counts.get(a["status"], 0) + 1
    lines.append("EQUIPMENT (" + ", ".join(f"{v} {k}" for k, v in status_counts.items()) + f", {len(assets)} total):")
    for a in assets:
        if a["status"] != "Available":
            lines.append(f"  - {a['name']}: {a['status']}" + (f" @ {a['location']}" if a["location"] else ""))

    open_concrete = db.execute(
        """SELECT c.project, c.pour_date, c.status, tp.client AS linked_client
           FROM inventory_concrete_requests c
           LEFT JOIN tracker_projects tp ON tp.id = c.project_id
           WHERE c.status != 'Completed' ORDER BY c.pour_date"""
    ).fetchall()
    lines.append(f"\nCONCRETE REQUESTS (open, {len(open_concrete)}):")
    for r in open_concrete[:15]:
        client_note = f" for {r['linked_client']}" if r["linked_client"] else ""
        lines.append(f"  - {r['project'] or 'Untitled'}{client_note}: {r['status']}, pour {r['pour_date'] or 'TBD'}")

    open_purchase = db.execute(
        """SELECT p.job_name, p.needed_on, p.status, tp.client AS linked_client
           FROM inventory_purchase_requests p
           LEFT JOIN tracker_projects tp ON tp.id = p.project_id
           WHERE p.status != 'Completed' ORDER BY p.needed_on"""
    ).fetchall()
    lines.append(f"\nPURCHASE REQUESTS (open, {len(open_purchase)}):")
    for r in open_purchase[:15]:
        client_note = f" for {r['linked_client']}" if r["linked_client"] else ""
        lines.append(f"  - {r['job_name'] or 'Untitled'}{client_note}: {r['status']}, needed {r['needed_on'] or 'TBD'}")

    try:
        projects = db.execute(
            "SELECT name, client, status, bid_due_date FROM tracker_projects WHERE status != 'Lost' AND status != 'Archived' ORDER BY bid_due_date"
        ).fetchall()
        lines.append(f"\nBID TRACKER (active, {len(projects)}):")
        for p in projects[:15]:
            lines.append(f"  - {p['name']}" + (f" ({p['client']})" if p["client"] else "") + f": {p['status']}, due {p['bid_due_date'] or 'TBD'}")
    except sqlite3.OperationalError:
        pass

    return "\n".join(lines)


CONCRETE_REQUEST_FIELDS = """
- project (text, REQUIRED) -- the job/project name
- job_site_address (text, REQUIRED) -- delivery address
- area_description (text, REQUIRED) -- what area/scope is being poured
- pour_date (date, REQUIRED) -- format YYYY-MM-DD
- pour_time (time, REQUIRED) -- format HH:MM 24-hour
- mix_design_psi (text, REQUIRED) -- e.g. "4000"
- mix_slump (text, REQUIRED) -- e.g. "4 inch"
- concrete_amount (text, REQUIRED) -- e.g. "130 yds"
- truck_spacing (text, REQUIRED) -- spacing between trucks, e.g. "15 min"
- pump_type (REQUIRED) -- one of: None, Ground Pump, Overhead Pump
- pump_size (text, REQUIRED only if pump_type is Ground Pump or Overhead Pump -- skip asking if pump_type is None)
- pump_arrival_time (time, REQUIRED only if pump_type is Ground Pump or Overhead Pump -- skip asking if pump_type is None) -- format HH:MM 24-hour
- lab_required (Yes/No, REQUIRED)
- lab_time (time, REQUIRED only if lab_required is Yes -- skip asking if lab_required is No) -- format HH:MM 24-hour
- drilling_required (Yes/No, REQUIRED)
- drilling_time (time, REQUIRED only if drilling_required is Yes -- skip asking if drilling_required is No) -- format HH:MM 24-hour
""".strip()

# The one field on the web form NOT asked about by voice -- "requested_date"
# is the date the request itself was made, which is filled in automatically
# with today's date at submission time. Asking someone "what's today's
# date" out loud would be a strange thing for Atlas to ask.
VOICE_REQUIRED_FIELDS = [f for f in CONCRETE_REQUEST_REQUIRED_FIELDS if f != "requested_date"]


def _parse_assistant_reply(raw_text):
    """Split Claude's raw response into the spoken part and the trailing
    <state>...</state> JSON block. Falls back gracefully if the model
    didn't include a state block (treats it as a plain chat answer)."""
    import re
    match = re.search(r"<state>(.*?)</state>", raw_text, re.DOTALL)
    if not match:
        return raw_text.strip(), {"mode": "chat", "fields": {}, "action": "none"}
    spoken = raw_text[:match.start()].strip()
    try:
        state = json.loads(match.group(1))
    except (json.JSONDecodeError, ValueError):
        state = {"mode": "chat", "fields": {}, "action": "none"}
    return spoken, state


ATLAS_VOICE_ID = os.environ.get("ELEVENLABS_VOICE_ID")  # No hardcoded fallback -- unset means "use the free browser voice," not "use some baked-in voice."


def _elevenlabs_tts_call(text):
    """The actual HTTP call to ElevenLabs' (non-streaming) text-to-speech
    endpoint for one piece of text -- a full reply, or one sentence when
    called from _synthesize_sentence_chunks below. Returns
    (base64_audio_or_None, error_or_None). Split out from
    generate_atlas_speech so both the old single-shot path and the new
    sentence-buffered streaming path share exactly one place that knows
    how to talk to ElevenLabs -- no duplicated request-building logic to
    drift out of sync.
    """
    if not ATLAS_VOICE_ID:
        return None, None
    api_key = os.environ.get("ELEVENLABS_API_KEY")
    if not api_key:
        return None, "ELEVENLABS_VOICE_ID is set but ELEVENLABS_API_KEY is not."
    if not text or not text.strip():
        return None, None
    import urllib.request
    import urllib.error
    import base64
    body = json.dumps({
        "text": text,
        "model_id": "eleven_multilingual_v2",
        "voice_settings": {"stability": 0.5, "similarity_boost": 0.75},
    }).encode("utf-8")
    req = urllib.request.Request(
        f"https://api.elevenlabs.io/v1/text-to-speech/{ATLAS_VOICE_ID}",
        data=body,
        headers={"Content-Type": "application/json", "xi-api-key": api_key, "Accept": "audio/mpeg"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            audio_bytes = resp.read()
            return base64.b64encode(audio_bytes).decode("ascii"), None
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")[:300]
        return None, f"ElevenLabs error {e.code}: {detail}"
    except (urllib.error.URLError, TimeoutError) as e:
        return None, f"ElevenLabs connection error: {str(e)}"


def generate_atlas_speech(text):
    """Generate speech audio for a complete piece of text in one call.
    Kept as the single-shot entry point for backward compatibility (used
    by the final-leftover-text case in stream_atlas_turn) -- the new
    sentence-by-sentence streaming path is _synthesize_sentence_chunks
    below, which calls the same underlying _elevenlabs_tts_call per
    sentence instead of once for the whole reply.
    """
    return _elevenlabs_tts_call(text)


_SENTENCE_BOUNDARY_RE = None  # compiled lazily, see _split_ready_sentences


def _split_ready_sentences(buffered_text):
    """Splits buffered_text into (ready_sentences, remainder) at real
    sentence boundaries (., !, ?, or a newline, followed by whitespace or
    end of string) -- NOT at arbitrary token/character counts. This is
    the "sensible sentence/phrase buffering" the spec asks for instead of
    either (a) one giant TTS call for the whole reply (current/old
    behavior -- all the latency lands up front) or (b) a TTS call per
    token/every-few-characters (naturalness suffers, and ElevenLabs
    credit usage would balloon -- short fragments still cost close to a
    full request's overhead). A sentence is a natural, cheap-enough, and
    prosody-safe unit to synthesize independently.

    Deliberately conservative: a boundary is only "ready" if there's
    already something after it in the buffer (or it's clearly terminal
    punctuation followed by whitespace) -- so we don't cut mid-sentence
    on a period that's actually a decimal point or abbreviation followed
    by more of the same sentence still streaming in. The very last
    (possibly incomplete) fragment is always returned as `remainder` and
    is only flushed by the caller once the stream is known to be done.
    """
    global _SENTENCE_BOUNDARY_RE
    if _SENTENCE_BOUNDARY_RE is None:
        import re
        _SENTENCE_BOUNDARY_RE = re.compile(r'([.!?]+["\')]?|\n)(\s+)')

    ready = []
    last_end = 0
    for m in _SENTENCE_BOUNDARY_RE.finditer(buffered_text):
        # Only treat this as a real boundary if there's more text after
        # it already buffered -- otherwise we can't yet tell whether
        # it's a genuine sentence end or Claude just hasn't continued
        # the sentence past the period yet.
        if m.end() < len(buffered_text):
            ready.append(buffered_text[last_end:m.end()].strip())
            last_end = m.end()
    remainder = buffered_text[last_end:]
    ready = [s for s in ready if s]
    return ready, remainder


def _build_atlas_system_prompt(snapshot, fields, project_context=None):
    """The fixed instructions + live context, sent as the system prompt on
    every turn. The conversation itself travels separately as a real
    messages array now, not flattened into this text.

    PROJECT CONTEXT (item 6): project_context (session-scoped, see
    ATLAS_SESSIONS[token]["project_context"]) is surfaced here as plain
    fact, not as an instruction to use any tool -- the live conversation
    loop does not dynamically dispatch to the Tool Registry today (only
    the snapshot above and the concrete-request field-collection state
    machine below actually run inside a turn); this prompt line is
    exactly the "make it available to the prompt so the model can reason
    consistently" connection point, nothing more. The one thing the
    model IS asked to do here is state, in <state>, the project NAME
    when the person establishes or changes which project they mean --
    never an id (the model never invents an id). That name is then
    resolved for real, server-side, via the same _find_project()-backed
    set_project_context tool call every other project lookup uses --
    ambiguous or unmatched names never get silently set.
    """
    if project_context and project_context.get("project_id"):
        context_line = f"CURRENT PROJECT CONTEXT: {project_context.get('name')} (project_id={project_context.get('project_id')}) -- assume this is the project unless the person clearly means a different one.\n\n"
    else:
        context_line = "CURRENT PROJECT CONTEXT: none established yet.\n\n"
    return (
        "You are Atlas, the assistant inside BuildIQ. If asked your name, "
        "say Atlas. People talk to you like they'd talk to Claude or "
        "ChatGPT -- hold a real conversation, remember what's already been "
        "said, and don't repeat a question that's already been answered. "
        "Replies may be read aloud by text-to-speech, so keep them "
        "conversational and avoid markdown or lists unless they're truly "
        "needed.\n\n"
        "You can do two things:\n"
        "1. Answer questions about the current state of the business using "
        "the snapshot below.\n"
        "2. Help someone submit a new concrete request by asking for "
        "whatever's still missing, one or two things at a time -- never "
        "more than that in one turn. A concrete request has these fields:\n"
        f"{CONCRETE_REQUEST_FIELDS}\n\n"
        "Rules for filling out a request:\n"
        "- Only ask about fields that are still blank in the current draft.\n"
        "- Never invent or assume a value the person didn't actually say.\n"
        "- If a project context is already established below and the "
        "request is for that project, use it for the 'project' field "
        "instead of asking the person to repeat the project name.\n"
        "- Once every REQUIRED field is filled, read back a short spoken "
        "summary of the whole request and ask them to confirm before "
        "submitting anything.\n"
        "- Only submit (action=submit) on the turn where they clearly "
        "confirm (yes / go ahead / submit it / that's right) AND every "
        "required field is already filled in the draft.\n"
        "- If they ask an unrelated question mid-request, just answer it "
        "normally -- don't force them back to the form.\n"
        "- If they say cancel/never mind/start over, clear the fields and "
        "set mode back to chat.\n\n"
        "PROJECT CONTEXT rules:\n"
        "- If the person establishes or changes which project they're "
        "talking about (e.g. \"we're working on Patel Farm\", \"switch to "
        "the Overlook Tower job\"), put that project's name -- exactly as "
        "they said it -- in the state block's project_name field.\n"
        "- If you are not confident which specific project they mean (the "
        "name could plausibly match more than one project), ask them to "
        "clarify in your reply instead, and leave project_name null this "
        "turn. Never guess.\n"
        "- If the project hasn't changed this turn, leave project_name "
        "null -- don't repeat it every turn.\n\n"
        + context_line +
        "CURRENT BUSINESS SNAPSHOT:\n" + snapshot + "\n\n"
        "CURRENT DRAFT (fields collected so far, empty if none in progress):\n"
        + json.dumps(fields) + "\n\n"
        "Respond in exactly two parts:\n"
        "1. Your natural reply.\n"
        "2. On its own line, a state block in EXACTLY this format:\n"
        '<state>{"mode": "concrete_request" or "chat", "fields": {<all fields known so far>}, "action": "none" or "submit", "project_name": "<name, or null if unchanged>"}</state>'
    )


class ToolResult:
    """What every tool handler's outcome gets wrapped into before it's
    ever seen by execute_tool's caller. Atlas never sees a raw exception,
    a raw DB row, or a raw traceback -- only this."""
    def __init__(self, success, data=None, error=None):
        self.success = success
        self.data = data
        self.error = error

    def to_dict(self):
        return {"success": self.success, "data": self.data, "error": self.error}


class ToolWriteRejected(Exception):
    """Raised by a tool handler (never caught anywhere except
    execute_tool) to fail a write CLOSED with a specific, structured
    reason -- as opposed to execute_tool's generic except-Exception
    catch-all ("something went wrong running that"). Used specifically
    for canonical-project-identity integrity failures (project_not_found)
    so an Atlas write that attempted to use a project_id never silently
    degrades into an unlinked write just because that id turned out to
    be invalid or stale -- see _tool_create_concrete_request."""
    pass


class AtlasTool:
    """One registered Atlas capability. `parameters` is a light schema:
    {name: {"type": "string"|"integer"|"number", "required": bool, "enum": [...]}}.
    `permission` is the *manual* permission a human doing this through the
    UI would need; `atlas_permission` is the separate Atlas-specific one --
    execute_tool requires BOTH, which is the mechanism that makes "a user
    can do X manually but not hand it to Atlas, or vice versa" actually
    enforced rather than just a design intention.
    """
    def __init__(self, name, description, parameters, permission, atlas_permission,
                 kind, handler, confirm=None):
        self.name = name
        self.description = description
        self.parameters = parameters or {}
        self.permission = permission
        self.atlas_permission = atlas_permission
        self.kind = kind  # "read" | "write"
        self.confirm = confirm if confirm is not None else (kind == "write")
        self.handler = handler


ATLAS_TOOLS = {}


def register_tool(name, description, parameters, permission, atlas_permission, kind, handler, confirm=None):
    """The only way a capability becomes callable by Atlas. Nothing else
    -- no raw SQL, no arbitrary Python, no route dispatch -- is ever
    reachable from the model's output. If it's not in ATLAS_TOOLS, it
    does not run.

    Rejects a malformed registration immediately (ValueError, at import
    time, loud and unmissable) rather than letting it into ATLAS_TOOLS --
    a tool with a missing/empty permission or atlas_permission would
    otherwise call user_has_permission(user, None), and after the
    fail-closed fix that correctly denies everyone... but "correctly
    denies everyone" for a tool that was supposed to work is still a
    bug worth catching at registration time rather than discovering at
    first use. This is a stricter check than the resolver's fail-closed
    behavior needs to provide on its own -- defense at both layers."""
    if not name or not isinstance(name, str):
        raise ValueError("register_tool: name must be a non-empty string")
    if not permission or not isinstance(permission, str):
        raise ValueError(f"register_tool({name!r}): permission must be a non-empty string, got {permission!r}")
    if not atlas_permission or not isinstance(atlas_permission, str):
        raise ValueError(f"register_tool({name!r}): atlas_permission must be a non-empty string, got {atlas_permission!r}")
    if kind not in ("read", "write"):
        raise ValueError(f"register_tool({name!r}): kind must be 'read' or 'write', got {kind!r}")
    if not callable(handler):
        raise ValueError(f"register_tool({name!r}): handler must be callable")
    ATLAS_TOOLS[name] = AtlasTool(name, description, parameters, permission, atlas_permission, kind, handler, confirm)


def _validate_tool_params(tool, raw_params):
    """Minimal but real schema enforcement: every required parameter must
    be present and non-empty; every provided parameter must be declared
    on the tool (unknown keys are dropped, not passed through); enum
    fields must match one of the declared values. Returns (clean_params,
    error_message_or_None).
    """
    if not isinstance(raw_params, dict):
        return None, "parameters must be an object"
    clean = {}
    for pname, spec in tool.parameters.items():
        value = raw_params.get(pname)
        required = spec.get("required", False)
        if (value is None or value == "") and required:
            return None, f"missing required parameter: {pname}"
        if value is None or value == "":
            continue
        enum = spec.get("enum")
        if enum and value not in enum:
            return None, f"invalid value for {pname}: must be one of {enum}"
        ptype = spec.get("type", "string")
        if ptype == "integer":
            try:
                value = int(value)
            except (TypeError, ValueError):
                return None, f"invalid value for {pname}: must be an integer"
        clean[pname] = value
    return clean, None


def execute_tool(tool_name, raw_params, user, confirmed=False, session_context=None):
    """The single centralized gateway every Atlas action must pass
    through -- this is what Phase 2's authorization requirement actually
    means in code. Every call is validated, permission-checked against
    BOTH the manual and Atlas-specific permission, stale-project-checked,
    confirmation-gated if it's a write, executed, and logged -- in
    EXACTLY that order, every time, with no bypass path:
        validate params -> permission check -> stale-context rejection
        -> confirmation check -> handler
    The stale-context check deliberately sits AFTER permission (so an
    unauthorized caller learns only "not permitted", never anything
    about whether a project exists) but BEFORE confirmation (so a stale
    canonical-project write is rejected on its very first call --
    confirmed or not -- and never enters the confirm/retry exchange at
    all; see the write_context_stale block below for exactly why this
    ordering, not just "check it somewhere", is the actual fix). The
    parser in stream_atlas_turn never calls a handler directly; it only
    ever calls this.

    PROJECT CONTEXT (item 6): `session_context` is the current Atlas
    session's project-context dict (see ATLAS_SESSIONS[token] --
    "project_context": {"project_id": int, "name": str} or {}), owned by
    the caller and passed in explicitly -- execute_tool never reaches
    into session state on its own. Two things happen here, both
    generically, so no per-tool special-casing is needed as more tools
    gain project_id support later:
      1. If the tool declares a `project_id` parameter and the caller
         didn't supply one, but session_context has a resolved
         project_id, it's filled in automatically -- this is what lets
         "create a purchase request for 40 sheets of plywood" reuse a
         project established earlier in the same conversation without
         the model having to re-ask or re-guess it. set_project_context
         itself is deliberately EXCLUDED from this injection (it's the
         one tool whose whole purpose is to CHANGE session_context --
         auto-filling its project_id from the very context it's meant
         to update would make switching projects by name alone
         impossible, since the old project_id would keep winning).
      2. If this IS the dedicated set_project_context tool and it
         resolves successfully, the result is written back into
         session_context so it persists for the rest of the session.
         This is the ONLY tool allowed to mutate session_context --
         every other tool only ever reads project_id like any other
         parameter.
    Nothing here infers an ambiguous project silently: resolution
    (including the "ambiguous -- ask the user" case) is entirely
    _find_project()'s existing logic, reused as-is.
    """
    tool = ATLAS_TOOLS.get(tool_name)
    if not tool:
        log_activity("atlas", "tool_call", 0, "atlas_unknown_tool", new_value=tool_name)
        get_db().commit()
        return ToolResult(False, error=f"unknown tool: {tool_name}")

    raw_params = dict(raw_params or {})
    write_context_stale = False
    if (session_context and tool_name != "set_project_context"
            and "project_id" in tool.parameters and not raw_params.get("project_id")):
        # STICKY STALE MARKER: if a PRIOR call already found this
        # session's project context stale and popped project_id/name
        # (below), that popping alone is not enough to protect a
        # SUBSEQUENT call on the same session_context object -- once
        # project_id is gone, a second call has literally no evidence
        # left that a project was ever intended, and would look
        # identical to a genuinely-unlinked request. The
        # "_project_context_stale" marker is what actually carries that
        # evidence forward: it keeps failing write attempts closed until
        # set_project_context succeeds again (which clears the marker as
        # part of establishing fresh, valid context -- see below) or an
        # explicit project_id is supplied directly (which bypasses
        # session context entirely and is unaffected by this marker).
        if session_context.get("_project_context_stale") and tool.kind == "write":
            write_context_stale = True
        else:
            ctx_project_id = session_context.get("project_id")
            if ctx_project_id:
                # Re-verify the stored project still exists every time
                # it's used, rather than trusting whatever was resolved
                # whenever set_project_context last ran -- a project can
                # be deleted from Project Hunt at any point after
                # context was established.
                still_exists = get_db().execute(
                    "SELECT 1 FROM tracker_projects WHERE id = ?", (ctx_project_id,)
                ).fetchone()
                if still_exists:
                    raw_params["project_id"] = ctx_project_id
                else:
                    # Fail-safe, but NOT the same fail-safe for every
                    # tool kind. Clearing the stale id/name so the
                    # person is asked to re-establish it is always
                    # correct. But for a WRITE tool, simply proceeding
                    # without project_id would silently convert "the
                    # person meant this to attach to a specific project"
                    # into an unlinked write -- a real canonical-identity
                    # loss, not a graceful degradation. Reads have no
                    # such risk (there's nothing to attach), so they're
                    # allowed to continue gracefully, same as before.
                    # The actual short-circuit for writes happens below,
                    # AFTER the normal permission/confirmation gates --
                    # this flag only records that it must happen; it
                    # never skips or reorders those checks. The sticky
                    # marker (see top of this block) is what makes this
                    # protection survive a second call on the same
                    # session_context, not just this one.
                    session_context.pop("project_id", None)
                    session_context.pop("name", None)
                    if tool.kind == "write":
                        session_context["_project_context_stale"] = True
                        write_context_stale = True

    clean_params, err = _validate_tool_params(tool, raw_params or {})
    if err:
        log_activity("atlas", "tool_call", 0, "atlas_validation_failed", field=tool_name, new_value=err)
        get_db().commit()
        return ToolResult(False, error=err)

    manual_ok = user_has_permission(user, tool.permission)
    atlas_ok = user_has_permission(user, tool.atlas_permission)
    if not (manual_ok and atlas_ok):
        log_activity("atlas", "tool_call", 0, "atlas_denied", field=tool_name,
                      new_value=f"manual_ok={manual_ok} atlas_ok={atlas_ok}")
        get_db().commit()
        return ToolResult(False, error="not permitted")

    if write_context_stale:
        # MUST be checked here -- after permission (so an unauthorized
        # caller still gets "not permitted", not a way to probe project
        # existence) but BEFORE the confirmation check below. Root cause
        # of the bug this ordering fixes: confirmation is a two-step
        # exchange (unconfirmed call -> "confirmation required" -> caller
        # retries with confirmed=True). Context was being cleared, once,
        # unconditionally, the moment staleness was DETECTED -- which
        # happens on every call, confirmed or not. If the confirmation
        # check ran first, the FIRST (unconfirmed) call would see
        # "confirmation required" while the stale project_id was already
        # wiped from session_context; the caller, told only to confirm,
        # would retry with confirmed=True, and by then there would be
        # nothing left to reject -- the write would go through unlinked,
        # exactly the silent identity loss this whole mechanism exists
        # to prevent. Rejecting BEFORE the confirmation check closes
        # that: the very first call (confirmed or not) that would have
        # depended on a stale project gets "project_context_stale"
        # immediately, writes nothing, and never enters the confirm
        # exchange at all -- so there is no confirmed retry to exploit.
        log_activity("atlas", "tool_call", 0, "atlas_stale_context", field=tool_name)
        get_db().commit()
        return ToolResult(False, error="project_context_stale")

    if tool.kind == "write" and tool.confirm and not confirmed:
        log_activity("atlas", "tool_call", 0, "atlas_unconfirmed", field=tool_name)
        get_db().commit()
        return ToolResult(False, error="confirmation required")

    try:
        data = tool.handler(user=user, **clean_params)
    except ToolWriteRejected as e:
        # Structured rejection from the handler itself (e.g. an
        # explicitly-supplied project_id that doesn't resolve to any
        # current tracker_projects row) -- distinct from an actual
        # unexpected error, so the caller gets the real, specific reason
        # instead of a generic "something went wrong".
        log_activity("atlas", "tool_call", 0, "atlas_rejected", field=tool_name, new_value=str(e))
        get_db().commit()
        return ToolResult(False, error=str(e))
    except Exception as e:
        log_activity("atlas", "tool_call", 0, "atlas_error", field=tool_name, new_value=str(e))
        get_db().commit()
        return ToolResult(False, error="something went wrong running that")

    if tool_name == "set_project_context" and session_context is not None and isinstance(data, dict) and data.get("found"):
        # The only tool allowed to write back into session_context, and
        # only ever with a project genuinely resolved by _find_project
        # (never a guess -- an ambiguous result is NOT written here,
        # leaving session_context unchanged until the user disambiguates).
        session_context["project_id"] = data.get("project_id")
        session_context["name"] = data.get("name")
        # A fresh, valid resolution clears the sticky stale marker (if
        # any) -- this is the ONLY way _project_context_stale ever gets
        # removed, which is exactly the "require Atlas to re-establish
        # valid project context" behavior: the marker persists across
        # any number of write attempts until this specific, successful
        # re-resolution happens.
        session_context.pop("_project_context_stale", None)

    action_label = f"atlas_{tool.kind}_{tool_name}"
    entity_id = data.get("id") if isinstance(data, dict) else 0
    log_activity("atlas", tool_name, entity_id or 0, action_label, new_value=str(clean_params))
    get_db().commit()
    return ToolResult(True, data=data)


def _tool_create_concrete_request(user, **fields):
    """Handler for the 'create_concrete_request' tool -- a thin wrapper
    around the exact same create_concrete_request() the web form has
    always used. No new insert logic, no new validation logic; only the
    call site changed (see stream_atlas_turn's submit branch).

    CANONICAL IDENTITY CORRECTION (item 6 follow-up): this is
    deliberately done HERE, in the Atlas-only wrapper, and NOT inside
    create_concrete_request() itself. The web form's "project" text
    field and its optional "Link to a Project Hunt project" project_id
    dropdown are independently editable by design (someone can type a
    shorthand/external job name while still linking a real project_id
    for cross-referencing) -- changing create_concrete_request() to
    force-overwrite project text from project_id would silently change
    that existing, intentional web-form behavior. Atlas is different: a
    project_id reaching this handler came either directly from the
    model or from execute_tool()'s session-context injection, and in
    both cases the free-text `project` the model also produced is only
    ever a best-effort guess at a display name, never an independent,
    deliberate choice the way a human filling out the web form makes.
    So here, and only here: if a project_id is present, it -- not the
    model's free-text guess -- decides the stored project identity.
    Resolved fresh against tracker_projects every call (never trusts a
    stale id). SECURITY INVARIANT: if it doesn't resolve, this fails
    CLOSED -- raises ToolWriteRejected("project_not_found") and creates
    NOTHING, rather than dropping the id and quietly writing an unlinked
    request. A project_id reaching this handler means canonical identity
    was actually attempted (by the model directly, or injected from
    Atlas session context by execute_tool); if that attempt can't be
    honored, that is an integrity failure to report back, not something
    to paper over by degrading to an unlinked write. Genuinely
    unlinked/external requests -- where no project_id was ever attempted
    at all -- are a completely different path and are unaffected: they
    still work exactly as they always have.
    """
    project_id = fields.get("project_id")
    if project_id:
        db = get_db()
        canonical = db.execute("SELECT name FROM tracker_projects WHERE id = ?", (project_id,)).fetchone()
        if canonical:
            fields = dict(fields)
            fields["project"] = canonical["name"]
        else:
            # FAIL CLOSED, not a silent downgrade to an unlinked request.
            # A project_id reaching this point means canonical identity
            # was explicitly attempted (either the model supplied one
            # directly, or execute_tool injected one from session
            # context) -- if it doesn't resolve, that's an integrity
            # failure to report, not something to quietly paper over by
            # dropping the id and writing an unlinked record anyway.
            # Genuinely-unlinked/external requests (no project_id at
            # all) are unaffected and still work exactly as before --
            # this branch only runs when a project_id was present.
            raise ToolWriteRejected("project_not_found")
    submitted_id = create_concrete_request(fields, user.name or user.email)
    return {"id": submitted_id, "submitted_id": submitted_id}


register_tool(
    name="create_concrete_request",
    description="Submit a new concrete pour request once every required field has been collected and the person has confirmed.",
    parameters={
        "project": {"type": "string", "required": True},
        "project_id": {"type": "integer", "required": False},
        "job_site_address": {"type": "string", "required": False},
        "area_description": {"type": "string", "required": False},
        "pour_date": {"type": "string", "required": True},
        "pour_time": {"type": "string", "required": False},
        "mix_design_psi": {"type": "string", "required": False},
        "mix_slump": {"type": "string", "required": False},
        "concrete_amount": {"type": "string", "required": False},
        "truck_spacing": {"type": "string", "required": False},
        "pump_type": {"type": "string", "required": False},
        "pump_size": {"type": "string", "required": False},
        "pump_arrival_time": {"type": "string", "required": False},
        "lab_required": {"type": "string", "required": False},
        "lab_time": {"type": "string", "required": False},
        "drilling_required": {"type": "string", "required": False},
        "drilling_time": {"type": "string", "required": False},
    },
    permission="action:sitepulse:manage",
    atlas_permission="atlas:create_requests",
    kind="write",
    confirm=True,
    handler=_tool_create_concrete_request,
)


def stream_atlas_turn(user_text, draft):
    """Streams one turn of the assistant as an SSE generator. Yields
    'data: {...}\\n\\n' lines; the caller (the Flask route) is responsible
    for actually returning a streaming Response built from this generator.

    Prior turns travel as a real messages array (draft["history"] is a
    list of {"role","content"} dicts fed straight to the API), not
    flattened into a text blob in the system prompt -- this is what gives
    it real multi-turn memory instead of a blurry paraphrase of the
    conversation so far.

    The trailing <state>...</state> block the model emits is never shown
    to the person -- it's buffered out of the visible stream and parsed
    once the response finishes.

    VOICE UPGRADE (progressive sentence-buffered TTS): as visible text
    arrives, completed sentences are dispatched to ElevenLabs on a
    background thread and sent to the client as `audio_chunk` events
    *while the rest of Claude's reply is still streaming in* -- audio
    for the first sentence can start playing well before the model has
    finished the whole response. This is progressive/sentence-buffered,
    NOT true low-level ElevenLabs streaming TTS (which would relay
    ElevenLabs' own chunked response incrementally) -- see the release
    review's TTS analysis for exact request-count/latency implications.
    Deliberately does NOT call ElevenLabs per token/every-few-characters
    (unnatural prosody, and would multiply API request count); see
    _split_ready_sentences for the real sentence-boundary logic. If
    ATLAS_VOICE_ID isn't configured, _elevenlabs_tts_call short-circuits
    to (None, None) per sentence at effectively zero cost. Each ready
    sentence's synthesis call is submitted to a small per-turn
    ThreadPoolExecutor immediately (non-blocking) so that Claude-stream
    consumption is NOT paused for the duration of each ElevenLabs HTTP
    call the way a fully inline/synchronous call would pause it --
    audio_chunk events are still emitted in strict sentence order
    (never a later sentence's audio ahead of an earlier one still in
    flight), just not necessarily blocking the read loop while waiting.

    WRITE-CONFIRMATION SECURITY -- IMPORTANT, READ BEFORE CHANGING:
    BuildIQ's own code requires the SAME complete field set to be
    proposed via action="submit" on TWO CONSECUTIVE model turns (via a
    fields-hash held in the per-session draft) before a write is even
    eligible to run. This hardens the boundary against a single stray
    model output triggering a write -- but the model is still the one
    interpreting whether the user's utterance means "yes, submit," and
    this mechanism does not give BuildIQ independent, out-of-band proof
    that the user explicitly confirmed. It only proves the model
    proposed the same complete action twice in a row.

    Separately, and just as important: this generator's Python code
    executes synchronously start-to-finish regardless of whether the
    client is still connected. Flask/WSGI only notice a client
    disconnect at the next attempted write to the socket -- a frontend
    AbortController.abort() does NOT stop this function from continuing
    to run past that point, including past an execute_tool() call, if
    one were made inline here. That is exactly why the actual write is
    NOT performed in this function even after the two-turn confirmation
    passes: see the `pending_write` handling below, which defers the
    real execute_tool() call to a separate /assistant/confirm_write
    request that the client only ever sends after it has verifiably
    finished receiving this entire SSE response. If the client aborts,
    the tab closes, or a barge-in interrupts before that happens, that
    follow-up request is simply never sent and the write never occurs
    -- deterministic by construction, not by hoping a disconnect gets
    detected mid-generator.
    """
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        msg = "This assistant isn't set up yet -- ask Ayoub to add an Anthropic API key."
        yield f"data: {json.dumps({'type': 'delta', 'text': msg})}\n\n"
        yield f"data: {json.dumps({'type': 'audio_chunk', 'seq': 0, 'final': True, 'text': msg, 'audio': None, 'audio_error': None})}\n\n"
        yield f"data: {json.dumps({'type': 'done', 'mode': draft.get('mode', 'chat'), 'submitted_id': None, 'audio': None, 'audio_error': None, 'pending_write_token': None})}\n\n"
        return

    snapshot = gather_business_snapshot()
    system = _build_atlas_system_prompt(snapshot, draft.get("fields", {}), draft.get("project_context"))

    history = draft.get("history", [])[-20:]  # last 10 exchanges, real turns
    messages = [{"role": h["role"], "content": h["content"]} for h in history]
    messages.append({"role": "user", "content": user_text})

    raw_text = ""
    visible_sent = ""
    STATE_TAG = "<state>"

    # Sentence-buffered TTS state for this turn. Each ready sentence's
    # ElevenLabs call is dispatched to this small per-turn executor as
    # soon as the sentence is ready, so Claude-stream consumption below
    # isn't paused for the duration of each HTTP call the way a fully
    # inline synchronous call would pause it. Shut down (without waiting
    # on stragglers -- they're joined explicitly at final flush instead)
    # at the very end of the turn.
    import concurrent.futures
    tts_buffer = ""
    chunk_seq = 0
    pending_futures = []  # ordered list of (seq, text, future) -- strict sentence order preserved on drain
    tts_executor = concurrent.futures.ThreadPoolExecutor(max_workers=2)

    def _submit_ready_sentences():
        """Splits whatever's newly available in tts_buffer and dispatches
        each ready sentence to the executor immediately (non-blocking) --
        does not yield anything itself."""
        nonlocal tts_buffer, chunk_seq
        ready, remainder = _split_ready_sentences(tts_buffer)
        tts_buffer = remainder
        for sentence in ready:
            future = tts_executor.submit(_elevenlabs_tts_call, sentence)
            pending_futures.append((chunk_seq, sentence, future))
            chunk_seq += 1

    def _drain_completed_chunks(block=False):
        """Yields audio_chunk SSE lines for whatever's ready at the FRONT
        of pending_futures, in strict order -- never pops a later
        sentence's future ahead of an earlier one still in flight, so
        playback order on the client is always correct even though
        synthesis itself may finish out of order in the background.
        block=True (used only at final flush) waits for every remaining
        future to complete rather than skipping ones still in flight."""
        while pending_futures:
            seq, sentence, future = pending_futures[0]
            if not block and not future.done():
                break
            audio_b64, audio_err = future.result()
            pending_futures.pop(0)
            yield f"data: {json.dumps({'type': 'audio_chunk', 'seq': seq, 'final': False, 'text': sentence, 'audio': audio_b64, 'audio_error': audio_err})}\n\n"

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type": "application/json", "x-api-key": api_key, "anthropic-version": "2023-06-01"},
            json={"model": "claude-sonnet-4-6", "max_tokens": 600, "system": system, "messages": messages, "stream": True},
            stream=True, timeout=60,
        )
        resp.raise_for_status()
        # One continuous pass over the stream, all the way to the end --
        # breaking out early to "drain the rest" in a second loop was the
        # bug that caused this: requests' iter_lines() buffers whatever
        # it's already pulled off the socket inside that generator's own
        # frame, and abandoning the generator via `break` throws away
        # anything sitting in that buffer, up to and including the
        # closing </state> tag. That silently truncated the state block,
        # so the model would say "confirming submission" while the
        # backend never actually saw action=submit and nothing was ever
        # created -- exactly the "I don't actually have the ability to
        # submit" contradiction on the next turn.
        for line in resp.iter_lines(decode_unicode=True):
            if not line or not line.startswith("data:"):
                continue
            payload = line[len("data:"):].strip()
            if payload in ("", "[DONE]"):
                continue
            try:
                event = json.loads(payload)
            except json.JSONDecodeError:
                continue
            if event.get("type") != "content_block_delta":
                continue
            delta_text = event.get("delta", {}).get("text", "")
            if not delta_text:
                continue
            raw_text += delta_text

            # Only stream out text we're sure isn't (part of) the state
            # tag -- hold back a small trailing window until we know.
            # Once the tag shows up, stop emitting new deltas but keep
            # looping so the rest of the stream (the full state JSON and
            # its closing tag) still gets read and accumulated.
            tag_idx = raw_text.find(STATE_TAG)
            if tag_idx != -1:
                safe_upto = tag_idx
            else:
                safe_upto = max(0, len(raw_text) - len(STATE_TAG))
            if safe_upto > len(visible_sent):
                new_chunk = raw_text[len(visible_sent):safe_upto]
                if new_chunk:
                    yield f"data: {json.dumps({'type': 'delta', 'text': new_chunk})}\n\n"
                    visible_sent = raw_text[:safe_upto]
                    tts_buffer += new_chunk
                    _submit_ready_sentences()
                    yield from _drain_completed_chunks(block=False)
    except requests.exceptions.RequestException as e:
        # requests' default str(e) for an HTTPError is just the generic
        # status line ("400 Client Error: Bad Request for url: ...") --
        # it throws away the actual JSON error body Anthropic sends back,
        # which is where the real, actionable reason lives (e.g. "model:
        # not_found_error", an account/key access issue, or a malformed
        # request). Surface that body when we have it instead of just
        # the generic status line, so a 400 is actually diagnosable from
        # the chat transcript rather than needing server log access.
        detail = str(e)
        resp_obj = getattr(e, "response", None)
        if resp_obj is not None:
            try:
                detail = resp_obj.text[:500]
            except Exception:
                pass
        err_msg = f"I hit an error talking to Claude: {detail}"
        yield f"data: {json.dumps({'type': 'delta', 'text': err_msg})}\n\n"
        yield f"data: {json.dumps({'type': 'audio_chunk', 'seq': 0, 'final': True, 'text': err_msg, 'audio': None, 'audio_error': None})}\n\n"
        yield f"data: {json.dumps({'type': 'done', 'mode': draft.get('mode', 'chat'), 'submitted_id': None, 'audio': None, 'audio_error': None, 'pending_write_token': None})}\n\n"
        tts_executor.shutdown(wait=False, cancel_futures=True)
        return

    spoken, state = _parse_assistant_reply(raw_text)
    if not visible_sent and spoken:
        # Fallback: state tag was never found mid-stream (model skipped
        # it, or it arrived in one big chunk) -- send the whole spoken
        # reply now rather than showing nothing.
        yield f"data: {json.dumps({'type': 'delta', 'text': spoken})}\n\n"
        tts_buffer += spoken

    mode = state.get("mode", "chat")
    fields = state.get("fields", {}) if mode == "concrete_request" else {}
    action = state.get("action", "none")

    # PROJECT CONTEXT (item 6): carry the session's established project
    # context forward turn-to-turn (it's session-scoped, not turn-scoped
    # -- new_draft otherwise replaces the whole draft each turn). If the
    # model named a project this turn, resolve it for real through the
    # exact same execute_tool("set_project_context", ...) gateway every
    # other tool call goes through -- permission-checked, logged, and
    # backed by _find_project()'s existing exact/unique-substring/
    # ambiguous logic. The model's stated name is only ever a hint to
    # resolve; an ambiguous or unmatched name leaves project_context
    # unchanged rather than guessing, and this is enforced here in code,
    # not just by the prompt asking nicely.
    project_context = dict(draft.get("project_context") or {})
    model_project_name = state.get("project_name")
    if model_project_name and isinstance(model_project_name, str) and model_project_name.strip():
        execute_tool("set_project_context", {"project_name": model_project_name.strip()},
                     current_user, session_context=project_context)

    new_history = history + [
        {"role": "user", "content": user_text},
        {"role": "assistant", "content": spoken},
    ]
    new_draft = {"mode": mode, "fields": fields, "history": new_history[-20:], "pending_submit": None, "project_context": project_context}

    submitted_id = None
    pending_write_token = None
    if action == "submit":
        needs_pump = fields.get("pump_type") in ("Ground Pump", "Overhead Pump")
        needs_lab = fields.get("lab_required") == "Yes"
        needs_drilling = fields.get("drilling_required") == "Yes"
        required_now = [
            f for f in VOICE_REQUIRED_FIELDS
            if (needs_pump or f not in ("pump_size", "pump_arrival_time"))
            and (needs_lab or f != "lab_time")
            and (needs_drilling or f != "drilling_time")
        ]
        missing = [f for f in required_now if not str(fields.get(f, "")).strip()]
        if not missing:
            fields_hash = hashlib.sha256(json.dumps(fields, sort_keys=True).encode("utf-8")).hexdigest()
            prior_pending = draft.get("pending_submit")
            if prior_pending and prior_pending.get("fields_hash") == fields_hash:
                # Same complete field set proposed on two consecutive
                # model turns -- eligible to write. The actual write is
                # deliberately NOT performed here (see the big docstring
                # at the top of this function for exactly why): it's
                # deferred to a separate /assistant/confirm_write request
                # the client only sends after fully receiving this SSE
                # response. draft["pending_write"] is what that route
                # looks up.
                fields["requested_date"] = date.today().isoformat()
                pending_write_token = secrets.token_hex(16)
                new_draft["pending_write"] = {"token": pending_write_token, "fields": dict(fields), "issued_at": time.time()}
            else:
                # First time this exact, complete field set has been
                # proposed for submission -- hold it. Do not write
                # anything yet. A second matching confirmation on the
                # very next turn is required.
                extra = " Just to be safe, say that one more time (like 'yes, submit it') and I'll send it."
                yield f"data: {json.dumps({'type': 'delta', 'text': extra})}\n\n"
                spoken += extra
                tts_buffer += extra
                new_draft["pending_submit"] = {"fields_hash": fields_hash}
        else:
            extra = " Actually, I'm still missing something required -- let's finish that first."
            yield f"data: {json.dumps({'type': 'delta', 'text': extra})}\n\n"
            spoken += extra
            tts_buffer += extra
            new_draft["pending_submit"] = None

    _submit_ready_sentences()
    if tts_buffer.strip():
        # Whatever's left over (even a fragment with no terminal
        # punctuation) is the tail of the reply -- there's no more text
        # coming to complete it, so submit it for synthesis as-is.
        future = tts_executor.submit(_elevenlabs_tts_call, tts_buffer.strip())
        pending_futures.append((chunk_seq, tts_buffer.strip(), future))
        chunk_seq += 1
        tts_buffer = ""
    yield from _drain_completed_chunks(block=True)
    tts_executor.shutdown(wait=False)

    draft.clear()
    draft.update(new_draft)

    # `audio`/`audio_error` on `done` are kept (always null on this path)
    # purely for older-client backward compatibility -- all real audio
    # for this turn was already delivered via audio_chunk events above.
    yield f"data: {json.dumps({'type': 'done', 'mode': new_draft.get('mode'), 'submitted_id': submitted_id, 'audio': None, 'audio_error': None, 'pending_write_token': pending_write_token})}\n\n"


@app.route("/assistant")
@login_required
def assistant_page():
    if not is_atlas_allowed():
        flash("Ask Ayoub for access to the office assistant.", "error")
        return redirect(url_for("home"))
    return render_template("assistant.html")


def transcribe_via_whisper(audio_bytes, mime_type):
    """Transcribe recorded audio via OpenAI's Whisper API. Preferred over
    ElevenLabs for listening -- Whisper costs roughly $0.006/minute versus
    ElevenLabs' ~330 credits/minute, which burns through the free 10,000
    credit/month allowance fast on its own. Keeping listening off
    ElevenLabs leaves the full credit allowance for the voice (talking
    back), which is the part actually worth paying for.
    Returns (text_or_None, error_or_None).
    """
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        return None, None  # Not configured -- caller falls back to ElevenLabs, not an error.
    if not audio_bytes:
        return None, "No audio received."
    import urllib.request
    import urllib.error
    import uuid

    boundary = uuid.uuid4().hex
    ext = "webm" if "webm" in (mime_type or "") else "mp4" if "mp4" in (mime_type or "") else "wav"
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="model"\r\n\r\nwhisper-1\r\n'
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="atlas.{ext}"\r\n'
        f"Content-Type: {mime_type or 'audio/webm'}\r\n\r\n"
    ).encode("utf-8") + audio_bytes + f"\r\n--{boundary}--\r\n".encode("utf-8")

    req = urllib.request.Request(
        "https://api.openai.com/v1/audio/transcriptions",
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}", "Authorization": f"Bearer {api_key}"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return data.get("text", "").strip(), None
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")[:300]
        return None, f"Whisper transcription error {e.code}: {detail}"
    except (urllib.error.URLError, TimeoutError) as e:
        return None, f"Whisper connection error: {str(e)}"


def transcribe_via_elevenlabs(audio_bytes, mime_type):
    """Transcribe recorded audio via ElevenLabs Speech-to-Text. Fallback
    used only when OPENAI_API_KEY isn't set -- see transcribe_via_whisper
    for why Whisper is preferred when available. Returns
    (text_or_None, error_or_None).
    """
    api_key = os.environ.get("ELEVENLABS_API_KEY")
    if not api_key:
        return None, "Neither OPENAI_API_KEY nor ELEVENLABS_API_KEY is set."
    if not audio_bytes:
        return None, "No audio received."
    import urllib.request
    import urllib.error
    import uuid

    boundary = uuid.uuid4().hex
    ext = "webm" if "webm" in (mime_type or "") else "mp4" if "mp4" in (mime_type or "") else "wav"
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="model_id"\r\n\r\nscribe_v1\r\n'
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="atlas.{ext}"\r\n'
        f"Content-Type: {mime_type or 'audio/webm'}\r\n\r\n"
    ).encode("utf-8") + audio_bytes + f"\r\n--{boundary}--\r\n".encode("utf-8")

    req = urllib.request.Request(
        "https://api.elevenlabs.io/v1/speech-to-text",
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}", "xi-api-key": api_key},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return data.get("text", "").strip(), None
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")[:300]
        return None, f"ElevenLabs transcription error {e.code}: {detail}"
    except (urllib.error.URLError, TimeoutError) as e:
        return None, f"Transcription connection error: {str(e)}"


@app.route("/assistant/ask", methods=["POST"])
@login_required
def assistant_ask():
    if not is_atlas_allowed():
        return {"error": "not authorized"}, 403

    transcribe_error = None
    if request.content_type and "multipart/form-data" in request.content_type:
        audio_file = request.files.get("audio")
        if not audio_file:
            question = ""
        else:
            audio_bytes, mime_type = audio_file.read(), audio_file.mimetype
            question, transcribe_error = transcribe_via_whisper(audio_bytes, mime_type)
            if question is None and transcribe_error is None:
                # Whisper not configured -- fall back to ElevenLabs.
                question, transcribe_error = transcribe_via_elevenlabs(audio_bytes, mime_type)
            question = (question or "").strip()
    else:
        question = (request.get_json(silent=True) or {}).get("question", "").strip()

    token = session.get("atlas_token")
    if not token:
        token = secrets.token_hex(16)
        session["atlas_token"] = token
    draft = ATLAS_SESSIONS.setdefault(token, {"mode": "chat", "fields": {}, "history": [], "pending_submit": None, "pending_write": None, "project_context": {}})

    def generate():
        yield f"data: {json.dumps({'type': 'question', 'text': question, 'transcribe_error': transcribe_error})}\n\n"
        if transcribe_error:
            yield f"data: {json.dumps({'type': 'done', 'mode': draft.get('mode', 'chat'), 'submitted_id': None, 'audio': None, 'audio_error': None, 'pending_write_token': None})}\n\n"
            return
        if not question:
            no_question_msg = "I didn't catch a question."
            yield f"data: {json.dumps({'type': 'delta', 'text': no_question_msg})}\n\n"
            audio_b64, audio_err = _elevenlabs_tts_call(no_question_msg)
            yield f"data: {json.dumps({'type': 'audio_chunk', 'seq': 0, 'final': True, 'text': no_question_msg, 'audio': audio_b64, 'audio_error': audio_err})}\n\n"
            yield f"data: {json.dumps({'type': 'done', 'mode': draft.get('mode', 'chat'), 'submitted_id': None, 'audio': None, 'audio_error': None, 'pending_write_token': None})}\n\n"
            return
        for chunk in stream_atlas_turn(question, draft):
            yield chunk

    return Response(stream_with_context(generate()), mimetype="text/event-stream", headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/assistant/reset", methods=["POST"])
@login_required
def assistant_reset():
    token = session.get("atlas_token")
    if token:
        ATLAS_SESSIONS.pop(token, None)
    return {"ok": True}


@app.route("/assistant/confirm_write", methods=["POST"])
@login_required
def assistant_confirm_write():
    """The ONLY place create_concrete_request's execute_tool call
    actually happens for the voice/Atlas path. Deliberately separate
    from stream_atlas_turn's SSE response -- see that function's
    docstring for the full reasoning. In short: stream_atlas_turn's
    Python code runs to completion synchronously regardless of whether
    the client is still connected, so performing the write inline there
    would mean a client-side abort (barge-in, tab close, network drop)
    could NOT reliably prevent an already-in-flight write. By requiring
    this separate request -- which the client only ever sends after it
    has verifiably finished receiving the whole prior SSE response --
    the write is deterministically contingent on that later request
    actually arriving. No request here, no write. Ever.

    ATOMIC CLAIM (fixes a real concurrent-request race found in release
    review): checking that a token is still pending and then clearing it
    were originally two separate steps, with no protection between them
    -- two simultaneous or retried requests carrying the same valid
    token could both pass the check before either cleared it, and both
    go on to call execute_tool, double-writing the concrete request.
    Fixed by making "verify the token is still pending, check it hasn't
    expired, and clear it" one atomic operation under
    ATLAS_WRITE_CONFIRM_LOCK -- whichever request acquires the lock
    first is the only one that can ever see pending_write still present
    for that token; every other concurrent or replayed request
    (including one that arrives after this one already cleared it) sees
    it gone and is rejected. The lock is released before execute_tool()
    runs, so a slow write doesn't hold up unrelated confirm_write calls
    for other sessions/tokens.

    IF execute_tool FAILS AFTER THE TOKEN WAS CLAIMED: the token is NOT
    restored or re-armed. Once claimed, it's spent -- whether or not the
    underlying write actually succeeded. A failed submission requires a
    fresh two-turn confirmation from Atlas (a new token), not a retry of
    the same one. This is deliberate: reopening the same token for retry
    would reintroduce exactly the double-execution window this fix
    closes, for the sake of a failure path that's already surfaced to
    the user as an explicit error they can just ask Atlas to try again.
    """
    if not is_atlas_allowed():
        return {"success": False, "error": "not authorized"}, 403

    token = session.get("atlas_token")
    submitted_token = (request.get_json(silent=True) or {}).get("token", "")
    if not token or not submitted_token:
        return {"success": False, "error": "no matching pending confirmation"}, 400

    with ATLAS_WRITE_CONFIRM_LOCK:
        draft = ATLAS_SESSIONS.get(token)
        pending = (draft or {}).get("pending_write")
        if not pending or pending.get("token") != submitted_token:
            # Either nothing pending, already claimed by a concurrent/
            # earlier request, or the token just doesn't match -- do
            # not write anything.
            return {"success": False, "error": "no matching pending confirmation"}, 400

        issued_at = pending.get("issued_at")
        if issued_at is None or (time.time() - issued_at) > PENDING_WRITE_TTL_SECONDS:
            draft["pending_write"] = None  # clear the stale token too, not just reject this call
            return {"success": False, "error": "that confirmation has expired -- please ask again"}, 400

        # Claimed. Cleared HERE, still inside the lock, before
        # execute_tool ever runs -- this is what actually makes it
        # atomic. Any other request for this same token, concurrent or
        # not, will now find pending_write already gone.
        fields = dict(pending.get("fields", {}))
        draft["pending_write"] = None
        draft["pending_submit"] = None

    result = execute_tool("create_concrete_request", fields, current_user, confirmed=True,
                          session_context=draft.get("project_context", {}))
    if result.success:
        draft["mode"] = "chat"
        draft["fields"] = {}
        draft["history"] = []
        return {"success": True, "submitted_id": result.data.get("submitted_id"), "error": None}
    return {"success": False, "submitted_id": None, "error": result.error}



REQUEST_STATUSES = ["Submitted", "Reviewing", "Approved", "Building", "Testing", "Released", "On Hold", "Not Planned"]
# Fix 7 (optional approval notes): a reasonable server-side cap, checked
# in code -- not just the HTML maxlength attribute, which is only a UX
# hint and never a security/data-integrity boundary on its own. Applies
# to both the (required) Return reason and the (optional) Approve note,
# since both are stored in the same field.
APPROVAL_NOTE_MAX_LENGTH = 500
CONCRETE_STATUS_OPTIONS = ["Submitted", "Scheduled", "Completed"]
# Not previously a named constant -- inventory_update_purchase_status /
# inventory_place_purchase_order use these two literal strings directly.
# Naming it here doesn't change behavior; it just gives the Phase 3A
# Atlas tool (and anything else that needs it) one real source instead
# of a second guess at the vocabulary.
PURCHASE_STATUS_OPTIONS = ["Submitted", "Scheduled", "Completed"]

# Phase 3A: the 7 read tools + get_attention_items, and the shared
# attention engine they (and eventually Product Intelligence) both use.
# See intelligence.py for what each one actually does. Placed here,
# after SP_STATUS_OPTIONS/PURCHASE_STATUS_OPTIONS/register_tool/get_db/
# user_has_permission all already exist in this module.
import intelligence
intelligence.register_atlas_tools(register_tool, SP_STATUS_OPTIONS, PURCHASE_STATUS_OPTIONS)


def _log_request_status(db, request_id, status, changed_by, release_note=None):
    """Write one row to the employee-visible status timeline. This is the
    single source of truth both the employee's "My Requests" timeline and
    the admin's status history read from -- there's no separate "current
    status" tracking logic to keep in sync, the latest row here always
    reflects the truth (and feature_requests.status is kept in step
    alongside it for fast filtering/display).
    """
    now = datetime.utcnow().isoformat()
    db.execute(
        "INSERT INTO feature_request_status_history (feature_request_id, status, release_note, changed_by, changed_at) VALUES (?,?,?,?,?)",
        (request_id, status, release_note, changed_by, now)
    )
    db.execute("UPDATE feature_requests SET status = ?, updated_at = ? WHERE id = ?", (status, now, request_id))


@app.route("/requests", methods=["GET", "POST"])
@login_required
def request_center():
    db = get_db()
    department_options = [d["name"] for d in db.execute("SELECT name FROM departments ORDER BY name").fetchall()]
    if request.method == "POST":
        text = request.form.get("original_request", "").strip()
        if not text:
            flash("Please describe what you need before submitting.", "error")
        else:
            submitted_department = request.form.get("department", "").strip()
            if not submitted_department:
                user_row = db.execute("SELECT department FROM users WHERE email = ?", (current_user.email,)).fetchone()
                submitted_department = user_row["department"] if user_row else None
            department = submitted_department
            now = datetime.utcnow().isoformat()
            cur = db.execute(
                "INSERT INTO feature_requests (requester_email, requester_name, department, original_request, status, approval_status, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?)",
                (current_user.email, current_user.name or current_user.email, department, text, "Submitted", "Pending", now, now)
            )
            request_id = cur.lastrowid
            db.commit()
            _log_request_status(db, request_id, "Submitted", current_user.email)
            for screenshot in request.files.getlist("screenshots"):
                saved_name = save_photo(screenshot)
                if saved_name:
                    db.execute(
                        "INSERT INTO feature_request_attachments (feature_request_id, filename, uploaded_by, created_at) VALUES (?,?,?,?)",
                        (request_id, saved_name, current_user.email, now)
                    )
            db.commit()
            flash("Request submitted.")
        return redirect(url_for("request_center"))

    my_requests = db.execute(
        "SELECT * FROM feature_requests WHERE requester_email = ? ORDER BY created_at DESC",
        (current_user.email,)
    ).fetchall()
    requests_with_history = []
    for r in my_requests:
        history = db.execute(
            "SELECT * FROM feature_request_status_history WHERE feature_request_id = ? ORDER BY changed_at",
            (r["id"],)
        ).fetchall()
        attachments = db.execute(
            "SELECT * FROM feature_request_attachments WHERE feature_request_id = ? ORDER BY id",
            (r["id"],)
        ).fetchall()
        approval_timeline = _get_approval_timeline(db, r["id"])
        requests_with_history.append({"r": r, "history": history, "attachments": attachments, "approval_timeline": approval_timeline})
    return render_template("requests/my_requests.html", requests_with_history=requests_with_history, department_options=department_options)


@app.route("/requests/<int:request_id>/resubmit", methods=["GET", "POST"])
@login_required
def request_resubmit(request_id):
    """Update & Resubmit for a Returned request -- the missing employee
    feedback-loop workflow. SAME request_id throughout (no duplicate row
    ever created): this only ever UPDATEs the existing feature_requests
    row and appends one feature_request_resubmissions event; it never
    INSERTs a new feature_requests row.

    AUTHORIZATION (both GET and POST, server-side, never trusting a
    hidden field): the authoritative DB record is re-read fresh on every
    request. Only the ORIGINAL requester (current_user.email ==
    r['requester_email'], read from the DB row itself, never from any
    client-supplied value) may use this route, and only while
    approval_status is exactly 'Returned'. An Administrator does not get
    special access here merely by being an Administrator -- ownership,
    not role, is the boundary the brief specifically calls for. Pending,
    Approved, and Released requests all fail this check the same way a
    stranger's Returned request would.
    """
    db = get_db()
    r = db.execute("SELECT * FROM feature_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("request_center"))
    if r["requester_email"] != current_user.email:
        flash("You can only edit and resubmit your own requests.", "error")
        return redirect(url_for("request_center"))
    if r["approval_status"] != "Returned":
        flash("This request isn't in a state that can be resubmitted.", "error")
        return redirect(url_for("request_center"))

    if request.method == "POST":
        text = request.form.get("original_request", "").strip()
        if not text:
            flash("Please describe what you need before resubmitting.", "error")
            return redirect(url_for("request_resubmit", request_id=request_id))
        # BLOCKER 2 FIX: department must come from the same canonical,
        # controlled vocabulary Request Center already uses (the real
        # departments table) -- never arbitrary free text. A submitted
        # value that isn't in that authoritative list is rejected and
        # ignored safely: the request's EXISTING department is kept
        # rather than silently accepting a typo'd/invented value that
        # would quietly pollute Product Intelligence's department
        # filters and reporting.
        submitted_department = request.form.get("department", "").strip()
        valid_departments = {d["name"] for d in db.execute("SELECT name FROM departments").fetchall()}
        if submitted_department and submitted_department in valid_departments:
            department = submitted_department
        else:
            department = r["department"]
        now = datetime.utcnow().isoformat()

        # ATOMIC ONE-WINNER TRANSITION -- same principle as the
        # Approve/Return concurrency fix: the WHERE clause, not a
        # Python-side read-then-check, is what actually decides whether
        # this resubmission is allowed to apply. Two resubmit attempts
        # against the same Returned row (e.g. a double-click, or two
        # tabs) can only ever have exactly one winner; the loser's
        # UPDATE matches zero rows.
        try:
            cur = db.execute(
                """UPDATE feature_requests SET original_request = ?, department = ?,
                   approval_status = 'Pending', approval_decided_by = NULL,
                   approval_decided_at = NULL, approval_reason = NULL, updated_at = ?
                   WHERE id = ? AND approval_status = 'Returned' AND requester_email = ?""",
                (text, department, now, request_id, current_user.email)
            )
            if cur.rowcount == 1:
                db.execute(
                    "INSERT INTO feature_request_resubmissions (feature_request_id, resubmitted_by, resubmitted_at) VALUES (?,?,?)",
                    (request_id, current_user.email, now)
                )
                log_activity("product_intelligence", "feature_request", request_id, "updated",
                             field="approval_status", old_value="Returned", new_value="Pending")
                db.commit()
                flash("Request updated and resubmitted for procurement approval.")
                return redirect(url_for("request_center"))
            else:
                db.rollback()
                flash("This request has already changed -- please review its current state.", "error")
                return redirect(url_for("request_center"))
        except sqlite3.OperationalError as e:
            db.rollback()
            log_activity("product_intelligence", "feature_request", request_id, "resubmit_contention_error", new_value=str(e))
            db.commit()
            flash("That update couldn't be processed right now because of a conflicting update -- please try again.", "error")
            return redirect(url_for("request_center"))

    approval_timeline = _get_approval_timeline(db, request_id)
    department_options = [d["name"] for d in db.execute("SELECT name FROM departments ORDER BY name").fetchall()]
    return render_template("requests/resubmit_request.html", r=r, approval_timeline=approval_timeline, department_options=department_options)


def _get_approval_timeline(db, request_id):
    """The merged, chronological Approve/Return/Resubmit event log for
    ONE request -- Approved/Returned decisions from
    feature_request_approvals plus Resubmitted events from
    feature_request_resubmissions, sorted together. Used by both the
    employee card (Returned reason + resubmit affordance) and the
    approver's detail page (so a returned-then-resubmitted request's
    full story -- original Return reason, who resubmitted it, when --
    is visible, not just the latest state)."""
    approvals = db.execute(
        "SELECT 'approval' AS kind, decision, reason, decided_by AS actor, decided_at AS at "
        "FROM feature_request_approvals WHERE feature_request_id = ?", (request_id,)
    ).fetchall()
    resubmissions = db.execute(
        "SELECT 'resubmission' AS kind, 'Resubmitted' AS decision, NULL AS reason, resubmitted_by AS actor, resubmitted_at AS at "
        "FROM feature_request_resubmissions WHERE feature_request_id = ?", (request_id,)
    ).fetchall()
    combined = [dict(row) for row in approvals] + [dict(row) for row in resubmissions]
    combined.sort(key=lambda e: e["at"])
    return combined


def _render_my_requests_for(db, email):
    """Shared by the real employee view and the admin's PREVIEW mode --
    same query, same filtering, so preview is a genuine test of the real
    access control rather than a separately-maintained mockup.
    """
    my_requests = db.execute(
        "SELECT * FROM feature_requests WHERE requester_email = ? ORDER BY created_at DESC", (email,)
    ).fetchall()
    requests_with_history = []
    for r in my_requests:
        history = db.execute(
            "SELECT * FROM feature_request_status_history WHERE feature_request_id = ? ORDER BY changed_at",
            (r["id"],)
        ).fetchall()
        attachments = db.execute(
            "SELECT * FROM feature_request_attachments WHERE feature_request_id = ? ORDER BY id",
            (r["id"],)
        ).fetchall()
        approval_timeline = _get_approval_timeline(db, r["id"])
        requests_with_history.append({"r": r, "history": history, "attachments": attachments, "approval_timeline": approval_timeline})
    return requests_with_history


def _spark_points(counts, width=220, height=30, pad=3):
    """Turn a list of daily counts into an SVG polyline 'points' string,
    scaled to fit the given viewbox. Used for the small module-tile
    sparklines on the Command Center."""
    if not counts:
        return ""
    mx = max(counts) or 1
    n = len(counts)
    step = width / (n - 1) if n > 1 else 0
    pts = []
    for i, c in enumerate(counts):
        x = round(i * step, 1)
        y = round(height - pad - (c / mx) * (height - 2 * pad), 1) if mx else height / 2
        pts.append(f"{x},{y}")
    return " ".join(pts)


def _trend_paths(counts, width=900, height=140, pad_top=8, pad_bottom=8):
    """Turn a list of daily counts into an SVG line path plus a matching
    filled-area path (closed down to the baseline), scaled to the given
    viewbox. Used for the big Requests Trend chart on Command Center."""
    if not counts:
        return "", ""
    mx = max(counts) if max(counts) > 0 else 1
    n = len(counts)
    step = width / (n - 1) if n > 1 else 0
    pts = []
    for i, c in enumerate(counts):
        x = round(i * step, 1)
        y = round(height - pad_bottom - (c / mx) * (height - pad_top - pad_bottom), 1)
        pts.append((x, y))
    line = " ".join(f"L{x},{y}" if i > 0 else f"M{x},{y}" for i, (x, y) in enumerate(pts))
    fill = line + f" L{pts[-1][0]},{height} L{pts[0][0]},{height} Z"
    return line, fill


@app.route("/admin/product-intelligence")
@login_required
def product_intelligence():
    if not _authorized("module:product_intelligence:view"):
        flash("Product Intelligence is restricted to admins.", "error")
        return redirect(url_for("home"))
    db = get_db()
    status_filter = request.args.get("status", "")
    department_filter = request.args.get("department", "")
    approval_filter = request.args.get("approval", "")
    conditions, params = [], []
    if status_filter:
        # Supports a single status (the existing dropdown) or a
        # comma-separated list (the new clickable KPI cards, e.g.
        # "Submitted,Reviewing" for the combined New/Reviewing card).
        status_list = [s.strip() for s in status_filter.split(",") if s.strip()]
        placeholders = ",".join("?" * len(status_list))
        conditions.append(f"status IN ({placeholders})")
        params.extend(status_list)
    if department_filter:
        conditions.append("department = ?")
        params.append(department_filter)
    if approval_filter:
        # Item 3: independent filter dimension alongside status/department
        # above -- approval_status is a separate column, not a status
        # value, so this is its own condition rather than folded into the
        # status IN (...) clause.
        conditions.append("approval_status = ?")
        params.append(approval_filter)
    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    rows = db.execute(f"SELECT * FROM feature_requests {where} ORDER BY created_at DESC", params).fetchall()
    departments = [d["department"] for d in db.execute(
        "SELECT DISTINCT department FROM feature_requests WHERE department IS NOT NULL AND department != ''"
    ).fetchall()]

    # PI flow refinement: a safe "come back here" URL for links FROM this
    # page INTO a request's detail page -- captures whatever filters are
    # currently applied (request.full_path already includes the query
    # string) plus which section the link came from, via the anchor
    # appended at each call site below. product_intelligence_detail()
    # validates this is actually a path on this same route before ever
    # using it (never an open redirect) -- see that route.
    back_base_url = request.full_path.rstrip("?")

    # Dashboard data -- all real aggregate queries against feature_requests
    # and its related tables, computed fresh every load. Nothing here is
    # fabricated or estimated; every number traces back to an actual row.
    all_requests = db.execute("SELECT * FROM feature_requests").fetchall()
    # Gate correction: a request whose approval_status isn't 'Approved'
    # cannot have actually advanced through the development lifecycle
    # (change_status/release now enforce that server-side -- see the
    # POST handler below), so it must not be counted as though it's
    # already "awaiting product review" or otherwise progressing through
    # that pipeline. `approved_requests` is what every development-
    # lifecycle-facing count below (status_counts, the KPI strip except
    # Total/Pending Approval, the Request Lifecycle chart, the pipeline
    # breakdown, the backlog trend, and the "awaiting review" attention
    # item) is computed from. `all_requests`/`total_requests` stays the
    # TRUE, complete count -- Total Requests and the All Requests table
    # below intentionally still include Pending/Returned requests, since
    # that table is the full historical list, not a development-pipeline
    # view.
    approved_requests = [r for r in all_requests if r["approval_status"] == "Approved"]
    status_counts = {}
    for r in approved_requests:
        status_counts[r["status"]] = status_counts.get(r["status"], 0) + 1
    total_requests = len(all_requests)
    # Item 3: pending-approval count is a real COUNT against the real
    # approval_status column -- not derived from status, and not an
    # estimate. Historical rows backfilled to 'Approved' (see the
    # migration) never appear here.
    pending_approval_count = db.execute(
        "SELECT COUNT(*) FROM feature_requests WHERE approval_status = 'Pending'"
    ).fetchone()[0]
    # Item 2 (PI flow refinement): the actual inbox rows for the new
    # Pending Approval section -- real rows, oldest-first (first-in-
    # first-out, matching how an actual approval queue should read),
    # not a re-derivation of the count above.
    pending_approval_requests = db.execute(
        "SELECT * FROM feature_requests WHERE approval_status = 'Pending' ORDER BY created_at ASC"
    ).fetchall()
    kpi = {
        "total": total_requests,
        "new_reviewing": status_counts.get("Submitted", 0) + status_counts.get("Reviewing", 0),
        "building": status_counts.get("Building", 0),
        "testing": status_counts.get("Testing", 0),
        "released": status_counts.get("Released", 0),
        "stalled": status_counts.get("On Hold", 0) + status_counts.get("Not Planned", 0),
        "pending_approval": pending_approval_count,
    }

    # NOTE: a `pipeline` (status/pct breakdown) list used to be computed
    # here, but was verified NOT to be consumed anywhere in
    # templates/requests/product_intelligence.html -- the visually
    # similar "Request Lifecycle" section on that page actually renders
    # `lifecycle` (below), a different variable. Its denominator
    # (total_requests) was also stale relative to status_counts now
    # being Approved-only (see the approval gate correction above), but
    # since the whole list was confirmed dead rather than just
    # mis-denominated, it -- and its now-unused pipeline_order/
    # pipeline_colors helpers -- have been removed rather than "fixed",
    # since carrying forward a corrected-but-unused calculation isn't
    # meaningfully safer than removing it.

    dept_counts = {}
    for r in all_requests:
        d = r["department"] or "Unassigned"
        dept_counts[d] = dept_counts.get(d, 0) + 1
    dept_breakdown = sorted(
        [{"label": k, "count": v, "pct": round(100 * v / total_requests) if total_requests else 0} for k, v in dept_counts.items()],
        key=lambda x: -x["count"]
    )

    module_rows = db.execute(
        "SELECT buildiq_module, COUNT(*) as c FROM feature_request_intelligence "
        "WHERE buildiq_module IS NOT NULL AND buildiq_module != '' GROUP BY buildiq_module"
    ).fetchall()
    module_total = sum(m["c"] for m in module_rows) or 1
    module_breakdown = sorted(
        [{"label": m["buildiq_module"], "count": m["c"], "pct": round(100 * m["c"] / module_total)} for m in module_rows],
        key=lambda x: -x["count"]
    )

    # v4 (Latest Movement truthfulness fix): the previous query only
    # ever looked at feature_request_status_history, so a request whose
    # most recent REAL event was a procurement Approve/Return/Resubmit
    # (which don't touch status_history at all -- they're a separate,
    # independent dimension by design) still showed its stale original
    # dev-status entry (e.g. "Submitted") as if that were the latest
    # thing that happened. This does not invent a new unified status
    # column -- it's a read-only UNION of the three REAL, existing event
    # tables, ordered by actual timestamp, so whichever genuinely
    # happened most recently for a request is what's shown for it.
    recent_activity = db.execute(
        """SELECT * FROM (
             SELECT 'status' AS kind, h.status AS label, h.changed_by AS actor, h.changed_at AS at, h.feature_request_id AS request_id
             FROM feature_request_status_history h
             UNION ALL
             SELECT 'approval' AS kind, a.decision AS label, a.decided_by AS actor, a.decided_at AS at, a.feature_request_id AS request_id
             FROM feature_request_approvals a
             UNION ALL
             SELECT 'resubmission' AS kind, 'Resubmitted' AS label, r.resubmitted_by AS actor, r.resubmitted_at AS at, r.feature_request_id AS request_id
             FROM feature_request_resubmissions r
           ) combined
           JOIN feature_requests f ON f.id = combined.request_id
           ORDER BY combined.at DESC LIMIT 8"""
    ).fetchall()

    recently_released = db.execute(
        """SELECT f.id, f.original_request, f.requester_name, f.requester_email, f.department,
           h.release_note, h.changed_at, i.buildiq_module
           FROM feature_requests f
           JOIN feature_request_status_history h ON h.feature_request_id = f.id
           LEFT JOIN feature_request_intelligence i ON i.feature_request_id = f.id
           WHERE f.status = 'Released' AND h.status = 'Released'
           ORDER BY h.changed_at DESC LIMIT 5"""
    ).fetchall()

    # ---- Command Center: control-room layer ----
    # Everything below is a real query against existing tables -- nothing
    # here is fabricated. Priority queue, system gauges, the requests
    # trend chart, and per-module activity sparklines.
    today = date.today()

    attention_items = []
    upcoming_bids = db.execute(
        "SELECT id, name, bid_due_date FROM tracker_projects "
        "WHERE status = 'In Progress' AND bid_due_date IS NOT NULL AND bid_due_date != '' "
        "ORDER BY bid_due_date ASC"
    ).fetchall()
    for p in upcoming_bids:
        try:
            due = datetime.strptime(p["bid_due_date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        days_left = (due - today).days
        if 0 <= days_left <= 3:
            attention_items.append({
                "priority": "high", "title": f"Bid due in {days_left} day{'s' if days_left != 1 else ''} \u2014 {p['name']}",
                "meta": "PROJECT HUNT", "action_label": "Open",
                "url": url_for("tracker_view_project", project_id=p["id"])
            })

    overdue_rentals = db.execute(
        "SELECT id, equipment_description, due_date FROM sitepulse_rentals "
        "WHERE returned_date IS NULL AND due_date IS NOT NULL AND due_date != '' AND due_date < ?",
        (today.isoformat(),)
    ).fetchall()
    for r in overdue_rentals:
        try:
            due = datetime.strptime(r["due_date"], "%Y-%m-%d").date()
            days_late = (today - due).days
        except ValueError:
            days_late = None
        attention_items.append({
            "priority": "high", "title": f"Rental overdue \u2014 {r['equipment_description']}",
            "meta": f"EQUIPMENT CENTER \u00b7 {days_late} day{'s' if days_late != 1 else ''} late" if days_late is not None else "EQUIPMENT CENTER",
            "action_label": "Open", "url": url_for("sitepulse_rentals_list")
        })

    pending_requests_count = status_counts.get("Submitted", 0) + status_counts.get("Reviewing", 0)
    if pending_requests_count > 0:
        # Fix 3 (Attention Required -> Review navigation): when there's
        # exactly one matching request, open it directly instead of
        # dropping the user into a filtered All Requests list they then
        # have to search through themselves. With more than one match,
        # a single click still can't reasonably pick which one, so the
        # filtered list (unchanged, existing behavior) is the honest
        # destination. back= carries them straight back to this section
        # -- reusing the exact same safe, already-audited back-context
        # mechanism the rest of Product Intelligence uses, not a new one.
        awaiting_review_matches = [r for r in approved_requests if r["status"] in ("Submitted", "Reviewing")]
        if len(awaiting_review_matches) == 1:
            review_url = url_for("product_intelligence_detail", request_id=awaiting_review_matches[0]["id"],
                                  back=back_base_url + "#pi2-attention")
        else:
            review_url = url_for("product_intelligence", status="Submitted,Reviewing")
        attention_items.append({
            "priority": "med", "title": f"{pending_requests_count} new request{'s' if pending_requests_count != 1 else ''} awaiting review",
            "meta": "REQUEST CENTER", "action_label": "Review",
            "url": review_url
        })

    tomorrow = (today + timedelta(days=1)).isoformat()
    unordered_pours = db.execute(
        """SELECT c.id, c.project, c.project_id, tp.name AS linked_project_name
           FROM inventory_concrete_requests c
           LEFT JOIN tracker_projects tp ON tp.id = c.project_id
           WHERE c.pour_date = ? AND c.status = 'Submitted'""",
        (tomorrow,)
    ).fetchall()
    for c in unordered_pours:
        display_name = c["linked_project_name"] or c["project"]
        attention_items.append({
            "priority": "med", "title": "Concrete pour tomorrow \u2014 no order placed",
            "meta": f"SITEPULSE \u00b7 {display_name}", "action_label": "Open",
            "url": url_for("inventory_concrete_list")
        })

    priority_rank = {"high": 0, "med": 1, "low": 2}
    attention_items.sort(key=lambda x: priority_rank.get(x["priority"], 3))
    attention_items = attention_items[:6]

    # NOTE: fleet_uptime_pct/resolution_rate_pct (a blended fleet-uptime +
    # request-resolution percentage) used to power a "System Health" gauge
    # that could show misleadingly low numbers (e.g. 0%) and imply BuildIQ
    # itself was broken. Removed as part of Product Intelligence 2.0's
    # System Health correction -- see platform_state below for its
    # truthful replacement.

    day_labels = [(today - timedelta(days=i)) for i in range(13, -1, -1)]
    submitted_counts, resolved_counts = [], []
    for d in day_labels:
        d_str = d.isoformat()
        submitted_counts.append(db.execute(
            "SELECT COUNT(*) FROM feature_requests WHERE date(created_at) = ?", (d_str,)
        ).fetchone()[0])
        resolved_counts.append(db.execute(
            "SELECT COUNT(*) FROM feature_request_status_history WHERE status = 'Released' AND date(changed_at) = ?", (d_str,)
        ).fetchone()[0])
    submitted_line, _ = _trend_paths(submitted_counts)
    resolved_line, resolved_fill = _trend_paths(resolved_counts)

    week_days = [(today - timedelta(days=i)) for i in range(6, -1, -1)]

    def _section_spark(section):
        counts = []
        for d in week_days:
            counts.append(db.execute(
                "SELECT COUNT(*) FROM activity_log WHERE section = ? AND date(created_at) = ?",
                (section, d.isoformat())
            ).fetchone()[0])
        return _spark_points(counts)

    active_bids_count = db.execute("SELECT COUNT(*) FROM tracker_projects WHERE status = 'In Progress'").fetchone()[0]
    in_maintenance_count = db.execute("SELECT COUNT(*) FROM sitepulse_assets WHERE status = 'In Maintenance'").fetchone()[0]
    open_po_count = db.execute("SELECT COUNT(*) FROM inventory_purchase_requests WHERE status != 'Completed'").fetchone()[0]

    # Per-module status text: replaces the old healthy=True/False ->
    # "Healthy"/"Attention" pill, which asserted a certification about
    # the WHOLE module that BuildIQ doesn't actually have data to back
    # (SitePulse's was hardcoded True with no real check at all -- CTO
    # audit finding). Each label below is scoped to exactly what's
    # tracked, using the same real attention_items/overdue_rentals data
    # already computed above -- never a claim about overall module health.
    ph_attention_count = sum(1 for a in attention_items if a["meta"] == "PROJECT HUNT")
    ec_attention_count = len(overdue_rentals)
    sp_attention_count = sum(1 for a in attention_items if a["meta"].startswith("SITEPULSE"))

    def _module_status_text(count, singular_noun="Attention Item"):
        if count == 0:
            return "No Attention Items"
        noun = singular_noun if count == 1 else singular_noun + "s"
        return f"{count} {noun}"

    module_tiles = [
        {"name": "Project Hunt", "value": active_bids_count, "label": "Active bids", "color": "var(--teal)",
         "attention": ph_attention_count > 0, "status_text": _module_status_text(ph_attention_count),
         "spark": _section_spark("tracker"), "spark_color": "#5EEAD4", "url": url_for("tracker_dashboard")},
        {"name": "Equipment Center", "value": in_maintenance_count, "label": "In maintenance", "color": "var(--brass)",
         "attention": ec_attention_count > 0, "status_text": _module_status_text(ec_attention_count, "Overdue Rental"),
         "spark": _section_spark("sitepulse"), "spark_color": "#C9A24B", "url": url_for("sitepulse_dashboard")},
        {"name": "SitePulse", "value": open_po_count, "label": "Open POs", "color": "var(--cyan)",
         "attention": sp_attention_count > 0, "status_text": _module_status_text(sp_attention_count),
         "spark": _section_spark("inventory"), "spark_color": "#5AC8E0", "url": url_for("inventory_home")},
    ]

    # ---- Request Health: metrics specific to the requests pipeline,
    # separate from System Health (which is about the platform overall). ----
    released_map = {}
    for row in db.execute(
        "SELECT feature_request_id, MIN(changed_at) as released_at FROM feature_request_status_history "
        "WHERE status = 'Released' GROUP BY feature_request_id"
    ).fetchall():
        released_map[row["feature_request_id"]] = row["released_at"]

    created_map = {r["id"]: r["created_at"] for r in approved_requests}

    resolve_days = []
    for req_id, released_at in released_map.items():
        created_at = created_map.get(req_id)
        if not created_at:
            continue
        try:
            d1 = datetime.fromisoformat(created_at)
            d2 = datetime.fromisoformat(released_at)
            resolve_days.append(((d2 - d1).total_seconds() / 86400, released_at))
        except ValueError:
            continue
    resolve_days.sort(key=lambda x: x[1], reverse=True)
    recent_resolve_days = [d for d, _ in resolve_days[:20]]
    avg_resolve_days = round(sum(recent_resolve_days) / len(recent_resolve_days), 1) if recent_resolve_days else None

    backlog_counts = []
    for d in day_labels:
        d_end = d.isoformat() + "T23:59:59"
        count = 0
        for req_id, c_at in created_map.items():
            if c_at > d_end:
                continue
            r_at = released_map.get(req_id)
            if r_at and r_at <= d_end:
                continue
            count += 1
        backlog_counts.append(count)
    backlog_line, _ = _trend_paths(backlog_counts, width=260, height=56, pad_top=6, pad_bottom=6)
    open_backlog_now = backlog_counts[-1] if backlog_counts else 0
    backlog_trend_up = len(backlog_counts) >= 2 and backlog_counts[-1] > backlog_counts[0]

    roadmap_rows = db.execute("SELECT * FROM roadmap_items ORDER BY sort_order ASC").fetchall()
    roadmap_lanes = {"now": [], "next": [], "evolving": [], "later": []}
    for item in roadmap_rows:
        roadmap_lanes.setdefault(item["lane"], []).append(item)
    module_health = [r for r in roadmap_rows if r["lane"] in ("now", "next")][:4]

    # ---- Product Intelligence 2.0: Command Center data ----
    # Everything below is either a direct re-derivation of numbers already
    # computed above, or a small new real query -- nothing here is
    # fabricated, estimated, or hardcoded. Where BuildIQ genuinely doesn't
    # have reliable data for a concept, the value is left None/empty and
    # the template shows a truthful state instead of a number.

    # Request Lifecycle: reuses status_counts already computed for `kpi`
    # above. REQUESTED combines Submitted+Reviewing (BuildIQ's own two
    # earliest real statuses) since the reference story's 5-stage
    # lifecycle doesn't distinguish them -- everything else is a 1:1
    # mapping onto REQUEST_STATUSES, the same statuses every other part
    # of Product Intelligence already uses.
    lifecycle = [
        {"key": "requested", "label": "Requested", "count": status_counts.get("Submitted", 0) + status_counts.get("Reviewing", 0)},
        {"key": "approved", "label": "Approved", "count": status_counts.get("Approved", 0)},
        {"key": "building", "label": "Building", "count": status_counts.get("Building", 0)},
        {"key": "testing", "label": "Testing", "count": status_counts.get("Testing", 0)},
        {"key": "resolved", "label": "Resolved", "count": status_counts.get("Released", 0)},
    ]

    # Priority Builds hero: pulls specific named roadmap_items rows
    # (the same table Build Direction reads) rather than a separate
    # data source, so editing one place (the existing roadmap edit UI)
    # keeps both sections truthful and in sync.
    def _priority_build(name, status_label):
        row = next((r for r in roadmap_rows if r["name"] == name), None)
        if not row:
            return None
        return {"name": name, "status_label": status_label, "lane": row["lane"],
                "note": row["note"], "progress_pct": row["progress_pct"]}
    priority_builds = [b for b in [
        _priority_build("Product Core", "FOUNDATION"),
        _priority_build("Product Intelligence", "POLISHING \u00b7 ACTIVE"),
        _priority_build("Atlas", "EVOLVING"),
    ] if b]

    # Situation panel: active builds/in-testing come straight from `kpi`
    # above; blockers are the high-priority items already computed for
    # Attention Required; changes-this-week is a real count from the
    # same activity_log table the module sparklines already query.
    week_ago = (datetime.utcnow() - timedelta(days=7)).isoformat()
    changes_this_week = db.execute(
        "SELECT COUNT(*) FROM activity_log WHERE created_at >= ?", (week_ago,)
    ).fetchone()[0]
    situation = {
        "active_builds": kpi["building"],
        "blockers": len([a for a in attention_items if a["priority"] == "high"]),
        "in_testing": kpi["testing"],
        "changes_this_week": changes_this_week,
    }

    # BuildIQ Pulse: same 14-day window already built for the trend
    # lines above (day_labels/submitted_counts/resolved_counts) -- just
    # summarized as real totals instead of (or alongside) the line chart.
    requests_received_14d = sum(submitted_counts)
    resolved_14d = sum(resolved_counts)
    modules_active_14d = db.execute(
        "SELECT COUNT(DISTINCT section) FROM activity_log WHERE created_at >= ?",
        (day_labels[0].isoformat(),)
    ).fetchone()[0]
    platform_events_14d = db.execute(
        "SELECT COUNT(*) FROM activity_log WHERE created_at >= ?",
        (day_labels[0].isoformat(),)
    ).fetchone()[0]
    pulse = {
        "requests_received": requests_received_14d,
        "resolved": resolved_14d,
        "modules_active": modules_active_14d,
        "platform_events": platform_events_14d,
        "has_trend_data": any(submitted_counts) or any(resolved_counts),
    }

    # Platform State: replaces the old "System Health" percentage, which
    # mixed fleet uptime and request-resolution rate into one misleading
    # number that could show 0% and look like BuildIQ was broken. Only
    # facts BuildIQ genuinely knows are shown -- whether the Atlas LLM
    # key and WhatsApp integration are configured. "Application
    # Operational" was removed after CTO audit: the fact that this route
    # executed is not meaningful uptime/health monitoring, and asserting
    # it as a status fact would be its own small fabrication.
    platform_state = {
        "atlas_configured": bool(os.environ.get("ANTHROPIC_API_KEY")),
        "whatsapp_configured": bool(ULTRAMSG_TOKEN),
    }

    # BuildIQ Ecosystem: reuses module_tiles (already real, already
    # computed above) and adds Atlas (state only -- no fabricated count)
    # and Product Core (the identity/connectivity foundation -- shown as
    # a non-clickable conceptual node, not a metric-bearing module, since
    # there's no dedicated Project Core page to click into).
    ecosystem_atlas = {"name": "Atlas", "configured": platform_state["atlas_configured"],
                        "url": url_for("assistant_page") if is_atlas_allowed() else None}

    # Build Direction defaults to its clean read-only presentation for
    # everyone; the Edit Roadmap toggle (and the note/lane editing
    # controls it reveals) only ever render for someone who already has
    # the real permission the roadmap_item_update route itself enforces
    # -- this is a display decision, not a new authorization boundary.
    can_manage_roadmap = _authorized("action:product_intelligence:manage")

    return render_template(
        "requests/product_intelligence.html", requests=rows, statuses=REQUEST_STATUSES,
        departments=departments, status_filter=status_filter, department_filter=department_filter,
        kpi=kpi, dept_breakdown=dept_breakdown, module_breakdown=module_breakdown,
        recent_activity=recent_activity, recently_released=recently_released,
        attention_items=attention_items,
        submitted_line=submitted_line, resolved_line=resolved_line, resolved_fill=resolved_fill,
        module_tiles=module_tiles, avg_resolve_days=avg_resolve_days, backlog_line=backlog_line,
        open_backlog_now=open_backlog_now, backlog_trend_up=backlog_trend_up,
        roadmap_lanes=roadmap_lanes, module_health=module_health, can_manage_roadmap=can_manage_roadmap,
        lifecycle=lifecycle, priority_builds=priority_builds, situation=situation,
        pulse=pulse, platform_state=platform_state, ecosystem_atlas=ecosystem_atlas,
        approval_filter=approval_filter, pending_approval_requests=pending_approval_requests,
        APPROVAL_NOTE_MAX_LENGTH=APPROVAL_NOTE_MAX_LENGTH,
        back_base_url=back_base_url,
    )


@app.route("/admin/roadmap/<int:item_id>/update", methods=["POST"])
@login_required
def roadmap_item_update(item_id):
    if not _authorized("action:product_intelligence:manage"):
        flash("Product Intelligence is restricted to admins.", "error")
        return redirect(url_for("home"))
    db = get_db()
    lane = request.form.get("lane", "later")
    if lane not in ("now", "next", "evolving", "later"):
        lane = "later"
    # Progress % is no longer exposed in the Build Direction UI (NOW/NEXT/
    # EVOLVING/LATER is the visible roadmap state model) -- but the field
    # and column remain for backward compatibility, per instruction, with
    # no migration. Since the form no longer submits progress_pct at all,
    # defaulting a missing value to 0 would silently zero out whatever was
    # already stored on every single note/lane edit -- preserve the
    # existing value instead when the field isn't present in the request.
    existing_row = db.execute("SELECT progress_pct FROM roadmap_items WHERE id = ?", (item_id,)).fetchone()
    existing_pct = existing_row[0] if existing_row else 0
    try:
        progress_pct = max(0, min(100, int(request.form.get("progress_pct", existing_pct))))
    except (ValueError, TypeError):
        progress_pct = existing_pct
    note = request.form.get("note", "").strip()
    db.execute(
        "UPDATE roadmap_items SET lane = ?, progress_pct = ?, note = ?, updated_at = ? WHERE id = ?",
        (lane, progress_pct, note, datetime.utcnow().isoformat(), item_id)
    )
    db.commit()
    flash("Roadmap updated.")
    return redirect(url_for("product_intelligence") + "#cc-roadmap")


@app.route("/admin/product-intelligence/<int:request_id>", methods=["GET", "POST"])
@login_required
def product_intelligence_detail(request_id):
    # Item 3: a procurement approver (action:product_intelligence:approve_requests)
    # needs to reach this page to act on a Pending request even if they
    # don't hold the broader action:product_intelligence:manage permission
    # (that stays scoped to whoever runs the dev pipeline). Every
    # individual POST action below is separately, specifically gated --
    # this outer check only decides who can load the page at all.
    if not (_authorized("action:product_intelligence:manage") or is_product_request_approver()):
        flash("Product Intelligence is restricted to admins.", "error")
        return redirect(url_for("home"))
    db = get_db()

    # PI flow refinement: a safe "back" destination, carried through GET
    # (from a link on the main PI page) and, if the form includes it as
    # a hidden field, through POST redirects too -- so acting on a
    # request and returning lands the user back where they actually
    # came from (their filtered All Requests view, the Pending Approval
    # section, etc.) instead of always the bare PI page.
    #
    # SECURITY: this value is attacker-controlled (a GET query param or
    # POST form field), so it is validated STRICTLY, not just with a
    # prefix check -- a plain `.startswith("/admin/product-intelligence")`
    # would still accept e.g. "/admin/product-intelligence@evil.com" or
    # a value containing ".." (still same-origin, but not a real
    # verification that it's actually the product_intelligence route --
    # not itself an open redirect, but not the "same-route" guarantee
    # this is supposed to provide either). Using urlsplit() and checking
    # each component explicitly instead:
    #   - scheme must be empty (rejects "https://...", "javascript:...")
    #   - netloc must be empty (rejects "//evil.com" and "http://evil.com")
    #   - path must be EXACTLY the product_intelligence route, not just
    #     prefixed by it (rejects any lookalike or traversal attempt --
    #     urlsplit does not collapse "..", so a path containing it can
    #     never equal the exact expected path string)
    # Query string and fragment are passed through as-is (that's the
    # actual filter/anchor state being preserved) since they can't
    # change which host/path the browser navigates to.
    from urllib.parse import urlsplit
    pi_base_path = url_for("product_intelligence")

    def _safe_pi_back_url(raw):
        if not raw:
            return None
        parsed = urlsplit(raw)
        if parsed.scheme or parsed.netloc or parsed.path != pi_base_path:
            return None
        return raw

    raw_back = request.args.get("back") or request.form.get("back") or ""
    back_url = _safe_pi_back_url(raw_back) or pi_base_path
    r = db.execute("SELECT * FROM feature_requests WHERE id = ?", (request_id,)).fetchone()
    if not r:
        flash("Request not found.", "error")
        return redirect(url_for("product_intelligence"))

    if request.method == "POST":
        action = request.form.get("action")
        now = datetime.utcnow().isoformat()

        # PI flow refinement: lets the new Pending Approval inbox cards
        # (on the main Product Intelligence page) submit approve/return
        # directly, then land back on that same page/section instead of
        # the request detail page -- "I remain oriented on the page."
        # Deliberately a closed whitelist value, never a raw URL/path
        # from the client, so this can never become an open redirect.
        return_to = request.form.get("return_to", "")

        def redirect_after_action():
            if return_to == "pending_approval":
                return redirect(url_for("product_intelligence") + "#pi2-pending-approval")
            if back_url != pi_base_path:
                return redirect(url_for("product_intelligence_detail", request_id=request_id, back=back_url))
            return redirect(url_for("product_intelligence_detail", request_id=request_id))

        if action == "change_department":
            if not _authorized("action:product_intelligence:manage"):
                flash("You don't have permission to make that change.", "error")
                return redirect(url_for("product_intelligence_detail", request_id=request_id))
            new_department = request.form.get("department", "").strip()
            db.execute("UPDATE feature_requests SET department = ?, updated_at = ? WHERE id = ?",
                       (new_department, now, request_id))
            db.commit()
            flash("Department corrected.")

        elif action == "save_details":
            if not _authorized("action:product_intelligence:manage"):
                flash("You don't have permission to make that change.", "error")
                return redirect(url_for("product_intelligence_detail", request_id=request_id))
            # Saves the admin-only intelligence fields WITHOUT touching
            # status at all -- exactly the separation asked for.
            db.execute(
                """INSERT INTO feature_request_intelligence
                   (feature_request_id, buildiq_module, internal_notes, solution_built, testing_notes, user_feedback, updated_at)
                   VALUES (?,?,?,?,?,?,?)
                   ON CONFLICT(feature_request_id) DO UPDATE SET
                   buildiq_module=excluded.buildiq_module, internal_notes=excluded.internal_notes,
                   solution_built=excluded.solution_built, testing_notes=excluded.testing_notes,
                   user_feedback=excluded.user_feedback, updated_at=excluded.updated_at""",
                (request_id, request.form.get("buildiq_module", ""), request.form.get("internal_notes", ""),
                 request.form.get("solution_built", ""), request.form.get("testing_notes", ""),
                 request.form.get("user_feedback", ""), now)
            )
            db.commit()
            flash("Details saved.")

        elif action == "change_status":
            if not _authorized("action:product_intelligence:manage"):
                flash("You don't have permission to make that change.", "error")
                return redirect(url_for("product_intelligence_detail", request_id=request_id))
            # Gate correction: approval and development status stay
            # independent FIELDS (no merging), but the WORKFLOW between
            # them is now enforced here, server-side -- a request that
            # hasn't cleared procurement approval cannot be advanced
            # through the development lifecycle. Template hiding alone
            # was previously the only thing stopping this; a direct POST
            # from someone holding action:product_intelligence:manage
            # but not approval authority could still move a Pending
            # request forward. This check closes that.
            if r["approval_status"] != "Approved":
                flash(f"This request cannot enter the development pipeline yet -- approval_status is {r['approval_status']}, not Approved.", "error")
                return redirect(url_for("product_intelligence_detail", request_id=request_id))
            new_status = request.form.get("status", "")
            if new_status in REQUEST_STATUSES and new_status != "Released":
                _log_request_status(db, request_id, new_status, current_user.email)
                db.commit()
                flash(f"Status moved to {new_status}.")
            else:
                flash("Invalid status change.", "error")

        elif action == "release":
            if not _authorized("action:product_intelligence:manage"):
                flash("You don't have permission to make that change.", "error")
                return redirect(url_for("product_intelligence_detail", request_id=request_id))
            # Same gate as change_status above -- a request that isn't
            # Approved cannot be released either.
            if r["approval_status"] != "Approved":
                flash(f"This request cannot be released -- approval_status is {r['approval_status']}, not Approved.", "error")
                return redirect(url_for("product_intelligence_detail", request_id=request_id))
            # Deliberate, separate action -- requires a note, confirmed
            # via the checkbox on the form. This is the only path that
            # can ever set status to Released.
            release_note = request.form.get("release_note", "").strip()
            confirmed = request.form.get("confirm_release") == "yes"
            if not release_note or not confirmed:
                flash("A release note and confirmation are both required to release a request.", "error")
            else:
                _log_request_status(db, request_id, "Released", current_user.email, release_note=release_note)
                db.execute(
                    """INSERT INTO feature_request_intelligence (feature_request_id, release_date, updated_at)
                       VALUES (?,?,?)
                       ON CONFLICT(feature_request_id) DO UPDATE SET release_date=excluded.release_date, updated_at=excluded.updated_at""",
                    (request_id, now, now)
                )
                db.commit()
                flash("Request released.")

        elif action in ("approve_request", "return_request"):
            # Item 3: procurement approval gate. Server-side authorization
            # in TWO independent parts, both required:
            #   1. the actor must hold the dedicated approve permission
            #      (never action:product_intelligence:manage alone --
            #      someone with only the broader manage permission but
            #      not explicitly granted approval authority cannot
            #      approve/return through this action).
            #   2. the actor cannot be the request's own requester, even
            #      if they otherwise hold approval authority.
            # Both checks happen here, on the POST handler itself -- not
            # only in the UI -- so a direct POST from someone lacking
            # either can never succeed.
            if not is_product_request_approver():
                flash("You don't have permission to approve or return requests.", "error")
                return redirect_after_action()
            if current_user.email == r["requester_email"]:
                flash("You cannot approve or return your own request.", "error")
                return redirect_after_action()
            if r["approval_status"] != "Pending":
                # Fast, friendly path only -- NOT the actual security
                # guarantee. `r` was read at the top of this request,
                # before any of the checks above ran; a concurrent
                # decision could still land in the gap between that read
                # and the UPDATE below. This early check just avoids
                # bothering an already-doomed request with reason
                # validation etc. The UPDATE...WHERE clause further down
                # is the ONLY thing that actually decides who wins.
                flash(f"This request has already been handled (now {r['approval_status']}) -- no action was taken.", "error")
                return redirect_after_action()

            reason = request.form.get("reason", "").strip()
            if action == "return_request" and not reason:
                flash("A reason is required to return a request.", "error")
                return redirect_after_action()
            if len(reason) > APPROVAL_NOTE_MAX_LENGTH:
                flash(f"That note is too long (max {APPROVAL_NOTE_MAX_LENGTH} characters) -- please shorten it and try again.", "error")
                return redirect_after_action()

            decision = "Approved" if action == "approve_request" else "Returned"

            # ATOMIC ONE-WINNER TRANSITION (concurrency fix). The
            # earlier design read approval_status, decided in Python
            # that it was Pending, then unconditionally wrote the new
            # state -- correct against a SEQUENTIAL stale request (the
            # second POST re-reads `r` at the top and sees the already-
            # changed value), but not against two truly concurrent
            # writers who could each pass that same Python-side check
            # before either commits. SQLite guarantees that a single
            # UPDATE statement is atomic with respect to the specific
            # rows it matches: the WHERE clause is evaluated against
            # the database's actual current state at the moment the
            # statement runs (under the connection's write lock), not
            # against a value read earlier in Python. So folding the
            # "still Pending" check directly into the UPDATE's WHERE
            # clause, and trusting ONLY `cur.rowcount` afterward, is
            # what makes this genuinely race-proof: whichever of two
            # concurrent UPDATEs against the same row actually acquires
            # SQLite's write lock first is guaranteed to see
            # approval_status = 'Pending' (still true) and win,
            # updating exactly one row; the second writer's UPDATE then
            # runs against a row that is no longer 'Pending' (the
            # winner's change is already committed, or the DB's locking
            # serializes the second UPDATE to run strictly after the
            # first), so its WHERE clause matches zero rows.
            #
            # ORDERING: the UPDATE runs FIRST, before the audit-history
            # INSERT, specifically so a loser (rowcount == 0) can never
            # produce a history row -- if the INSERT happened first and
            # the UPDATE lost the race, we would have "corrected" that
            # by deleting the row we just wrote (extra complexity, and
            # a real window where a losing decision is briefly visible
            # in the audit trail). Checking rowcount immediately after
            # the UPDATE and only inserting history when it's exactly 1
            # means a losing writer's transaction contains no writes at
            # all -- there's nothing to roll back beyond ending the
            # (otherwise-empty) transaction.
            try:
                cur = db.execute(
                    """UPDATE feature_requests SET approval_status = ?, approval_decided_by = ?,
                       approval_decided_at = ?, approval_reason = ?, updated_at = ?
                       WHERE id = ? AND approval_status = 'Pending'""",
                    (decision, current_user.email, now, reason or None, now, request_id)
                )
                if cur.rowcount == 1:
                    # This request won the race (or there was no race
                    # at all) -- exactly one approval-history row for
                    # exactly one winning decision.
                    db.execute(
                        "INSERT INTO feature_request_approvals (feature_request_id, decision, reason, decided_by, decided_at) VALUES (?,?,?,?,?)",
                        (request_id, decision, reason or None, current_user.email, now)
                    )
                    log_activity("product_intelligence", "feature_request", request_id, "updated",
                                 field="approval_status", old_value="Pending", new_value=decision)
                    db.commit()
                    if decision == "Approved":
                        flash("Request approved and moved to the development queue.")
                    else:
                        flash("Request returned to the requester.")
                else:
                    # Lost the race (or the sequential-stale case) --
                    # another decision already committed. No history
                    # row, no overwrite of the winner's data. Roll back
                    # explicitly so this connection doesn't hold an
                    # open (empty) transaction/lock any longer than
                    # necessary.
                    db.rollback()
                    fresh = db.execute("SELECT approval_status FROM feature_requests WHERE id = ?", (request_id,)).fetchone()
                    current_state = fresh["approval_status"] if fresh else "unknown"
                    flash(f"This request has already been handled (now {current_state}) -- no action was taken.", "error")
            except sqlite3.OperationalError as e:
                # Fail safe rather than let a raw SQLite locking error
                # (e.g. a busy-timeout still being exceeded under heavy
                # contention) reach the user as an unhandled exception.
                db.rollback()
                log_activity("product_intelligence", "feature_request", request_id, "approval_contention_error", new_value=str(e))
                db.commit()
                flash("That decision couldn't be processed right now because of a conflicting update -- please check the request's current status and try again if needed.", "error")

        return redirect_after_action()

    history = db.execute(
        "SELECT * FROM feature_request_status_history WHERE feature_request_id = ? ORDER BY changed_at",
        (request_id,)
    ).fetchall()
    # v4: merged Approve/Return/Resubmit timeline (was: approvals only)
    # -- so an approver reviewing a resubmitted request can see the full
    # story (original Return + reason, who resubmitted it, when) instead
    # of only the latest decision.
    approval_history = _get_approval_timeline(db, request_id)
    intel = db.execute("SELECT * FROM feature_request_intelligence WHERE feature_request_id = ?", (request_id,)).fetchone()
    attachments = db.execute(
        "SELECT * FROM feature_request_attachments WHERE feature_request_id = ? ORDER BY id", (request_id,)
    ).fetchall()
    department_options = [d["name"] for d in db.execute("SELECT name FROM departments ORDER BY name").fetchall()]
    can_manage = _authorized("action:product_intelligence:manage")
    can_approve = is_product_request_approver() and r["approval_status"] == "Pending" and current_user.email != r["requester_email"]
    return render_template("requests/product_intelligence_detail.html", r=r, history=history, intel=intel,
                            statuses=REQUEST_STATUSES, attachments=attachments, department_options=department_options,
                            approval_history=approval_history, can_manage_product_intelligence=can_manage,
                            can_approve_request=can_approve, back_url=back_url, APPROVAL_NOTE_MAX_LENGTH=APPROVAL_NOTE_MAX_LENGTH)


@app.route("/admin/product-intelligence/preview")
@login_required
def product_intelligence_preview():
    if not _authorized("module:product_intelligence:view"):
        flash("Product Intelligence is restricted to admins.", "error")
        return redirect(url_for("home"))
    db = get_db()
    preview_email = request.args.get("email", "")
    all_users = db.execute("SELECT email, name FROM users ORDER BY name").fetchall()
    requests_with_history = _render_my_requests_for(db, preview_email) if preview_email else []
    return render_template(
        "requests/preview_employee_view.html", requests_with_history=requests_with_history,
        all_users=all_users, preview_email=preview_email
    )


@app.route("/admin/users", methods=["GET", "POST"])
@login_required
def admin_users():
    if not _authorized("module:team_admin:view"):
        flash("This page is restricted to admins.", "error")
        return redirect(url_for("home"))
    db = get_db()
    if request.method == "POST":
        # Viewing this page only requires module:team_admin:view; actually
        # changing a department or adding one is a write action and gets
        # its own, stricter check per the "write actions must be
        # separately protected" requirement.
        if not _authorized("action:team_admin:manage_users"):
            flash("You don't have permission to make changes here.", "error")
            return redirect(url_for("admin_users"))

        action = request.form.get("action", "assign_department")

        if action == "add_department":
            new_dept = request.form.get("new_department", "").strip()
            if not new_dept:
                flash("Enter a department name.", "error")
            else:
                try:
                    db.execute(
                        "INSERT INTO departments (name, created_at) VALUES (?, ?)",
                        (new_dept, datetime.utcnow().isoformat())
                    )
                    db.commit()
                    flash(f"'{new_dept}' added to Departments.")
                except sqlite3.IntegrityError:
                    flash(f"'{new_dept}' already exists.", "error")
            return redirect(url_for("admin_users"))

        user_id = request.form.get("user_id")
        department = request.form.get("department", "").strip()
        db.execute("UPDATE users SET department = ? WHERE id = ?", (department, user_id))
        db.commit()
        flash("Department updated.")
        return redirect(url_for("admin_users"))

    users = db.execute("SELECT id, name, email, department FROM users ORDER BY name").fetchall()
    department_options = [d["name"] for d in db.execute("SELECT name FROM departments ORDER BY name").fetchall()]
    return render_template("requests/admin_users.html", users=users, department_options=department_options)


def _actor_is_administrator(db, user_id):
    """Does this user currently hold the Administrator role? Used only
    by the privilege-escalation guard below -- deliberately a role
    membership check, not a new permission key, per the "don't invent
    a permission for this" decision (see admin_user_permissions'
    docstring for the reasoning)."""
    row = db.execute(
        "SELECT 1 FROM user_roles ur JOIN roles r ON r.id = ur.role_id "
        "WHERE ur.user_id = ? AND r.name = 'Administrator' LIMIT 1",
        (user_id,)
    ).fetchone()
    return row is not None


# Permissions whose grant IS an escalation path (manage_users can be used
# to repeat this whole exercise on someone else) or unlocks destructive
# system-wide action (system_data:manage covers backup/restore/import/
# export). Narrower than "everything Administrator has" -- this is what
# actually needs a full-Administrator actor to hand out, not every
# permission in the catalog.
HIGH_PRIVILEGE_PERMISSION_KEYS = {
    "action:team_admin:manage_users",
    "action:system_data:manage",
}


@app.route("/admin/users/<int:user_id>/permissions", methods=["GET", "POST"])
@login_required
def admin_user_permissions(user_id):
    """Manage one user's roles and permission overrides.

    PRIVILEGE-ESCALATION GUARD (added after CTO audit): action:team_admin:
    manage_users is meant to let ordinary admin staff handle routine user
    management -- assigning a role, tweaking one permission. Left
    unguarded, that same permission could be used to hand out Administrator
    access (to anyone, including the actor themselves), grant other
    high-privilege permissions, or strip the last Administrator from the
    system entirely -- turning "manage users" into unrestricted privilege
    escalation. Deliberately NOT solved by adding a new permission key
    (there's nothing a new key would let this code check that role
    membership doesn't already tell us): the guard below requires the
    ACTOR to already hold the Administrator role before they can (a)
    grant the Administrator role to anyone, (b) grant any key in
    HIGH_PRIVILEGE_PERMISSION_KEYS to anyone, (c) remove the Administrator
    role from anyone, or (d) remove an explicit DENY on a high-privilege
    key (which would silently restore role-inherited access to it). It
    also blocks a user from changing their OWN roles/overrides through
    this page at all, and blocks removing the last remaining
    Administrator. None of this changes what action:team_admin:
    manage_users itself means or who has it -- only what an ordinary
    (non-Administrator) holder of it can do with it.
    """
    if not _authorized("action:team_admin:manage_users"):
        flash("This page is restricted to admins.", "error")
        return redirect(url_for("home"))

    db = get_db()
    target_user = db.execute("SELECT id, name, email FROM users WHERE id = ?", (user_id,)).fetchone()
    if not target_user:
        flash("User not found.", "error")
        return redirect(url_for("admin_users"))

    if request.method == "POST":
        action = request.form.get("action")
        now = datetime.utcnow().isoformat()
        admin_role_row = db.execute("SELECT id FROM roles WHERE name = 'Administrator'").fetchone()
        admin_role_id = admin_role_row["id"] if admin_role_row else None
        actor_is_admin = _actor_is_administrator(db, current_user.id)

        def _valid_role_id(raw):
            """Parses and validates a role_id from form input. Returns the
            int id if it's a real integer AND a real row in roles, else
            None. Guards against hostile/malformed input (non-numeric
            strings, negative numbers, ids for roles that don't exist)
            causing either a server 500 from int()/int-column comparisons,
            or a meaningless orphan user_roles row pointing at a
            nonexistent role."""
            try:
                rid = int(raw)
            except (TypeError, ValueError):
                return None
            row = db.execute("SELECT id FROM roles WHERE id = ?", (rid,)).fetchone()
            return rid if row else None

        def _valid_permission_id(raw):
            """Same validation as _valid_role_id, for permission_id."""
            try:
                pid = int(raw)
            except (TypeError, ValueError):
                return None
            row = db.execute("SELECT id FROM permissions WHERE id = ?", (pid,)).fetchone()
            return pid if row else None

        # Self-modification block: never let this page change the acting
        # user's own roles/overrides, in either direction -- no
        # self-escalation, and no accidental self-lockout either.
        if user_id == current_user.id:
            flash("You can't change your own roles or permissions from this page.", "error")
            return redirect(url_for("admin_user_permissions", user_id=user_id))

        if action == "assign_role":
            role_id = _valid_role_id(request.form.get("role_id"))
            if role_id is None:
                flash("Not a valid role.", "error")
                return redirect(url_for("admin_user_permissions", user_id=user_id))
            if role_id == admin_role_id and not actor_is_admin:
                flash("Only an Administrator can grant Administrator access.", "error")
                return redirect(url_for("admin_user_permissions", user_id=user_id))
            # Product Intelligence 2.0 role simplification: this is "the
            # normal Users & Permissions interface" the CTO's instruction
            # refers to -- only the 3 simplified roles can be newly
            # assigned through it. Existing legacy role assignments are
            # untouched (this only guards new assignment, never removal),
            # and nothing about permission resolution itself changes.
            role_name_row = db.execute("SELECT name FROM roles WHERE id = ?", (role_id,)).fetchone()
            if role_name_row and role_name_row["name"] not in ASSIGNABLE_ROLES:
                flash(f"\"{role_name_row['name']}\" is a legacy role and can no longer be newly assigned here.", "error")
                return redirect(url_for("admin_user_permissions", user_id=user_id))
            db.execute("INSERT OR IGNORE INTO user_roles (user_id, role_id) VALUES (?, ?)", (user_id, role_id))
            db.commit()
            flash("Role assigned.")

        elif action == "remove_role":
            role_id = _valid_role_id(request.form.get("role_id"))
            if role_id is None:
                flash("Not a valid role.", "error")
                return redirect(url_for("admin_user_permissions", user_id=user_id))
            if admin_role_id and role_id == admin_role_id:
                if not actor_is_admin:
                    flash("Only an Administrator can remove Administrator access.", "error")
                    return redirect(url_for("admin_user_permissions", user_id=user_id))
                remaining = db.execute(
                    "SELECT COUNT(*) FROM user_roles WHERE role_id = ? AND user_id != ?",
                    (admin_role_id, user_id)
                ).fetchone()[0]
                if remaining == 0:
                    flash("Can't remove the last Administrator.", "error")
                    return redirect(url_for("admin_user_permissions", user_id=user_id))
            db.execute("DELETE FROM user_roles WHERE user_id = ? AND role_id = ?", (user_id, role_id))
            db.commit()
            flash("Role removed.")

        elif action == "set_override":
            permission_id = _valid_permission_id(request.form.get("permission_id"))
            state = request.form.get("state")  # 'grant' or 'deny'
            if permission_id is None:
                flash("Not a valid permission.", "error")
                return redirect(url_for("admin_user_permissions", user_id=user_id))
            if state in ("grant", "deny"):
                if state == "grant":
                    perm_row = db.execute("SELECT key FROM permissions WHERE id = ?", (permission_id,)).fetchone()
                    if perm_row and perm_row["key"] in HIGH_PRIVILEGE_PERMISSION_KEYS and not actor_is_admin:
                        flash("Only an Administrator can grant this permission.", "error")
                        return redirect(url_for("admin_user_permissions", user_id=user_id))
                db.execute(
                    """INSERT INTO user_permission_overrides (user_id, permission_id, state, granted_by, updated_at)
                       VALUES (?, ?, ?, ?, ?)
                       ON CONFLICT(user_id, permission_id) DO UPDATE SET state=excluded.state, granted_by=excluded.granted_by, updated_at=excluded.updated_at""",
                    (user_id, permission_id, state, current_user.email, now)
                )
                db.commit()
                flash(f"Permission {state}ed.")

        elif action == "remove_override":
            permission_id = _valid_permission_id(request.form.get("permission_id"))
            if permission_id is None:
                flash("Not a valid permission.", "error")
                return redirect(url_for("admin_user_permissions", user_id=user_id))
            existing_override = db.execute(
                "SELECT state, permission_id FROM user_permission_overrides WHERE user_id = ? AND permission_id = ?",
                (user_id, permission_id)
            ).fetchone()
            if existing_override and existing_override["state"] == "deny":
                perm_row = db.execute("SELECT key FROM permissions WHERE id = ?", (permission_id,)).fetchone()
                if perm_row and perm_row["key"] in HIGH_PRIVILEGE_PERMISSION_KEYS and not actor_is_admin:
                    flash("Only an Administrator can remove this restriction.", "error")
                    return redirect(url_for("admin_user_permissions", user_id=user_id))
            db.execute("DELETE FROM user_permission_overrides WHERE user_id = ? AND permission_id = ?", (user_id, permission_id))
            db.commit()
            flash("Override removed -- back to role-inherited.")

        return redirect(url_for("admin_user_permissions", user_id=user_id))

    all_roles = db.execute("SELECT id, name, description FROM roles ORDER BY name").fetchall()
    assignable_role_ids = {r["id"] for r in all_roles if r["name"] in ASSIGNABLE_ROLES}
    user_role_ids = {r["role_id"] for r in db.execute("SELECT role_id FROM user_roles WHERE user_id = ?", (user_id,)).fetchall()}
    overrides = {r["permission_id"]: r["state"] for r in db.execute(
        "SELECT permission_id, state FROM user_permission_overrides WHERE user_id = ?", (user_id,)
    ).fetchall()}
    role_granted_perm_ids = set()
    for rid in user_role_ids:
        for r in db.execute("SELECT permission_id FROM role_permissions WHERE role_id = ?", (rid,)).fetchall():
            role_granted_perm_ids.add(r["permission_id"])

    all_permissions = db.execute("SELECT id, key, category, label, description FROM permissions ORDER BY category, label").fetchall()
    permissions_by_category = {}
    for p in all_permissions:
        state = overrides.get(p["id"])
        if state == "grant":
            effective, source = True, "granted"
        elif state == "deny":
            effective, source = False, "denied"
        elif p["id"] in role_granted_perm_ids:
            effective, source = True, "inherited"
        else:
            effective, source = False, "inherited"
        permissions_by_category.setdefault(p["category"], []).append({
            "id": p["id"], "key": p["key"], "label": p["label"], "description": p["description"],
            "effective": effective, "source": source,
        })

    return render_template(
        "requests/admin_user_permissions.html", target_user=target_user, all_roles=all_roles,
        assignable_role_ids=assignable_role_ids,
        user_role_ids=user_role_ids, permissions_by_category=permissions_by_category
    )


def tr_call_claude(prompt, max_tokens=800):
    api_key = os.environ.get("ANTHROPIC_API_KEY")


def tr_build_rfq_prompt(quote, project):
    today = datetime.now().strftime("%B %d, %Y")
    return f"""Write a short, professional Request for Quote (RFQ) email to a vendor,
asking them to provide pricing for a specific trade/scope on a construction project.
This is an initial outreach, not a follow-up — the vendor has not been contacted yet.

Today's date: {today}
Project: {project['name']}
Trade/Scope needed: {quote['trade']}
Vendor contact: {quote['vendor_contact'] or 'there'}
Vendor company: {quote['vendor_name'] or ''}

Keep it brief (under 120 words), professional, and clear about what's being requested
(a quote for the specified trade/scope). Mention we'd appreciate pricing at their
earliest convenience. No subject line, just the email body. Sign off generically as
"the estimating team," not a specific person's name."""


def tr_build_followup_prompt(quote, project):
    today = datetime.now().strftime("%B %d, %Y")
    return f"""Write a short, professional follow-up email to a vendor who has not yet
responded to a request for quote (RFQ).

Today's date: {today}
Project: {project['name']}
Trade: {quote['trade']}
Vendor contact: {quote['vendor_contact'] or 'there'}
Vendor company: {quote['vendor_name'] or ''}
RFQ sent date: {quote['rfq_sent_date'] or 'recently'}

Keep it brief (under 100 words), polite but direct about needing a response soon since
the bid deadline is approaching. No subject line, just the email body. Sign off generically
as "the estimating team," not a specific person's name."""


TR_UPLOAD_DIR = UPLOAD_DIR
TR_ALLOWED_UPLOAD_EXTENSIONS = {"pdf", "doc", "docx", "xls", "xlsx", "png", "jpg", "jpeg"}
TR_FILE_SIGNATURES = {
    "pdf": [b"%PDF"], "png": [b"\x89PNG"], "jpg": [b"\xff\xd8\xff"], "jpeg": [b"\xff\xd8\xff"],
    "docx": [b"PK\x03\x04"], "xlsx": [b"PK\x03\x04"],
    "doc": [b"\xd0\xcf\x11\xe0", b"PK\x03\x04"], "xls": [b"\xd0\xcf\x11\xe0", b"PK\x03\x04"],
}


def tr_allowed_upload_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in TR_ALLOWED_UPLOAD_EXTENSIONS


def tr_file_content_matches_extension(file_storage, extension):
    file_storage.seek(0)
    header = file_storage.read(8)
    file_storage.seek(0)
    signatures = TR_FILE_SIGNATURES.get(extension, [])
    if not signatures:
        return True
    return any(header.startswith(sig) for sig in signatures)


@app.route("/tracker/")
@login_required
def tracker_dashboard():
    db = get_db()
    filter_status = request.args.get("filter", "")
    if "filter" not in request.args:
        # No filter specified at all (fresh visit to Project Hunt) --
        # default to showing only active (In Progress) bids. "all" is
        # the explicit escape hatch to see everything else.
        filter_status = "active"

    all_projects = db.execute("SELECT * FROM tracker_projects ORDER BY bid_due_date ASC").fetchall()
    active_projects = [p for p in all_projects if p["status"] != "Archived"]

    def parse_val(v):
        if not v:
            return 0.0
        cleaned = v.replace("$", "").replace(",", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0

    kpis = {
        "active": len([p for p in active_projects if p["status"] == "In Progress"]),
        "submitted": len([p for p in active_projects if p["status"] == "Submitted"]),
        "awarded": len([p for p in active_projects if p["status"] == "Awarded"]),
        "unmeant": len([p for p in active_projects if p["status"] == "Unmeant"]),
    }

    total_blocking = 0
    upcoming = []
    call_today = []
    for p in active_projects:
        quotes = db.execute("SELECT * FROM tracker_quotes WHERE project_id = ?", (p["id"],)).fetchall()
        blocking = [q for q in quotes if q["is_submit_blocking"] and q["status"] != "Received"]
        total_blocking += len(blocking)
        dl = tr_daysleft_filter(p["bid_due_date"])
        if dl is not None and p["status"] in ("In Progress", "Pending") and dl <= 14:
            upcoming.append((p, dl))

        if p["status"] not in ("In Progress", "Pending"):
            continue
        for q in quotes:
            if q["status"] == "Sent" and q["rfq_sent_date"]:
                try:
                    sent = datetime.strptime(q["rfq_sent_date"], "%Y-%m-%d").date()
                    days_waiting = (date.today() - sent).days
                    if days_waiting >= 3:
                        call_today.append((p, q, days_waiting))
                except ValueError:
                    pass
    upcoming.sort(key=lambda x: x[1])
    call_today.sort(key=lambda x: -x[2])

    if filter_status == "active":
        display_projects = [p for p in active_projects if p["status"] == "In Progress"]
    elif filter_status == "submitted":
        display_projects = [p for p in active_projects if p["status"] == "Submitted"]
    elif filter_status == "awarded":
        display_projects = [p for p in active_projects if p["status"] == "Awarded"]
    elif filter_status == "unmeant":
        display_projects = [p for p in active_projects if p["status"] == "Unmeant"]
    else:
        # "all" (explicit) or any other/blank value -- show everything
        # that isn't archived.
        display_projects = active_projects

    client_filter = request.args.get("client", "").strip().lower()
    if client_filter:
        display_projects = [p for p in display_projects if p["client"] and client_filter in p["client"].lower()]

    status_filter = request.args.get("status", "")
    if status_filter:
        display_projects = [p for p in display_projects if p["status"] == status_filter]

    sort_by = request.args.get("sort", "due")
    sort_dir = request.args.get("dir", "asc")
    sort_key_map = {
        "name": lambda p: (p["name"] or "").lower(),
        "client": lambda p: (p["client"] or "").lower(),
        "status": lambda p: (p["status"] or ""),
        "value": lambda p: parse_val(p["estimated_value"]),
        "due": lambda p: p["bid_due_date"] or "9999-99-99",
    }
    if sort_by in sort_key_map:
        display_projects = sorted(display_projects, key=sort_key_map[sort_by], reverse=(sort_dir == "desc"))

    return render_template("tracker/dashboard.html", projects=display_projects, kpis=kpis,
                            total_blocking=total_blocking, upcoming=upcoming[:5],
                            call_today=call_today[:8], filter_status=filter_status,
                            client_filter=request.args.get("client", ""), status_filter=status_filter,
                            sort_by=sort_by, sort_dir=sort_dir, status_options=TR_STATUS_OPTIONS)


@app.route("/tracker/archive")
@login_required
def tracker_archive():
    db = get_db()
    projects = db.execute("SELECT * FROM tracker_projects WHERE status = 'Archived' ORDER BY updated_at DESC").fetchall()
    return render_template("tracker/archive.html", projects=projects)


@app.route("/tracker/project/<int:project_id>/delete", methods=["POST"])
@login_required
def tracker_delete_project(project_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (project_id,)).fetchone()
    log_activity("tracker", "project", project_id, "deleted", asset_id=project_id,
                 old_value=project["name"] if project else None)
    db.execute("DELETE FROM tracker_quotes WHERE project_id = ?", (project_id,))
    db.execute("DELETE FROM tracker_docs WHERE project_id = ?", (project_id,))
    # Phase 1: unlink (never cascade-delete) anything referencing this
    # project from other modules -- the concrete request, PO, rental, or
    # usage log entry is still a real record of work that happened; it
    # just stops being tied to a project that no longer exists. The
    # original free-text value is untouched either way.
    for table in ("inventory_concrete_requests", "inventory_purchase_requests", "sitepulse_usage_log", "sitepulse_rentals"):
        db.execute(f"UPDATE {table} SET project_id = NULL WHERE project_id = ?", (project_id,))
    db.execute("DELETE FROM tracker_projects WHERE id = ?", (project_id,))
    db.commit()
    flash("Project deleted.")
    return redirect(url_for("tracker_dashboard"))


@app.route("/tracker/quote/<int:quote_id>/upload", methods=["POST"])
@login_required
def tracker_upload_quote_file(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    from flask import send_file
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    uploaded_file = request.files.get("quote_file")
    if not uploaded_file or uploaded_file.filename == "":
        flash("No file selected.", "error")
        return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))
    if not tr_allowed_upload_file(uploaded_file.filename):
        flash("File type not allowed. Use PDF, Word, Excel, or image files.", "error")
        return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))
    original_name = uploaded_file.filename
    ext = original_name.rsplit(".", 1)[1].lower()
    if not tr_file_content_matches_extension(uploaded_file, ext):
        flash("That file's content doesn't match its extension — upload rejected for safety.", "error")
        return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))
    stored_name = f"quote_{quote_id}_{secrets.token_hex(6)}.{ext}"
    uploaded_file.save(os.path.join(TR_UPLOAD_DIR, stored_name))
    db.execute(
        "UPDATE tracker_quotes SET attachment_filename = ?, attachment_original_name = ?, updated_at = ? WHERE id = ?",
        (stored_name, original_name, datetime.utcnow().isoformat(), quote_id)
    )
    db.commit()
    flash("File uploaded.")
    return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))


@app.route("/tracker/quote/<int:quote_id>/download")
@login_required
def tracker_download_quote_file(quote_id):
    from flask import send_file
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote or not quote["attachment_filename"]:
        flash("No file attached to this quote.", "error")
        return redirect(url_for("tracker_dashboard"))
    file_path = os.path.join(TR_UPLOAD_DIR, quote["attachment_filename"])
    if not os.path.exists(file_path):
        flash("File not found on server.", "error")
        return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))
    return send_file(file_path, as_attachment=True, download_name=quote["attachment_original_name"])


@app.route("/tracker/quote/<int:quote_id>/delete_file", methods=["POST"])
@login_required
def tracker_delete_quote_file(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    if quote["attachment_filename"]:
        file_path = os.path.join(TR_UPLOAD_DIR, quote["attachment_filename"])
        if os.path.exists(file_path):
            os.remove(file_path)
    db.execute("UPDATE tracker_quotes SET attachment_filename = NULL, attachment_original_name = NULL WHERE id = ?", (quote_id,))
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))


@app.route("/tracker/quote/<int:quote_id>/edit", methods=["GET", "POST"])
@login_required
def tracker_edit_quote(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (quote["project_id"],)).fetchone()
    if request.method == "POST":
        new_fields = {
            "trade": request.form["trade"], "vendor_name": request.form.get("vendor_name", ""),
            "vendor_contact": request.form.get("vendor_contact", ""), "vendor_email": request.form.get("vendor_email", ""),
            "vendor_phone": tr_format_phone(request.form.get("vendor_phone", "")),
            "rfq_sent_date": request.form.get("rfq_sent_date", ""), "status": request.form.get("status", quote["status"]),
            "is_submit_blocking": 1 if request.form.get("is_submit_blocking") else 0,
            "amount": tr_format_currency(request.form.get("amount", "")), "notes": request.form.get("notes", ""),
        }
        db.execute(
            """UPDATE tracker_quotes SET trade = ?, vendor_name = ?, vendor_contact = ?, vendor_email = ?,
               vendor_phone = ?, rfq_sent_date = ?, status = ?, is_submit_blocking = ?,
               amount = ?, notes = ?, updated_at = ? WHERE id = ?""",
            (new_fields["trade"], new_fields["vendor_name"], new_fields["vendor_contact"],
             new_fields["vendor_email"], new_fields["vendor_phone"], new_fields["rfq_sent_date"],
             new_fields["status"], new_fields["is_submit_blocking"], new_fields["amount"],
             new_fields["notes"], datetime.utcnow().isoformat(), quote_id)
        )
        for field_name, new_val in new_fields.items():
            old_val = quote[field_name]
            if old_val != new_val:
                log_activity("tracker", "quote", quote_id, "updated", asset_id=quote["project_id"],
                             field=field_name, old_value=old_val, new_value=new_val)
        db.commit()
        return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))
    return render_template("tracker/edit_quote.html", q=quote, project=project)


@app.route("/tracker/quote/<int:quote_id>/delete", methods=["POST"])
@login_required
def tracker_delete_quote(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    project_id = quote["project_id"]
    log_activity("tracker", "quote", quote_id, "deleted", asset_id=project_id,
                 old_value=f"{quote['trade']} - {quote['vendor_name']}")
    db.execute("DELETE FROM tracker_quotes WHERE id = ?", (quote_id,))
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=project_id))


@app.route("/tracker/project/new", methods=["GET", "POST"])
@login_required
def tracker_new_project():
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    if request.method == "POST":
        db = get_db()
        now = datetime.utcnow().isoformat()
        cur = db.execute(
            """INSERT INTO tracker_projects (name, client, address, bid_due_date, estimated_value, status,
               assigned_to, notes, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (request.form["name"], request.form.get("client", ""), request.form.get("address", ""),
             request.form.get("bid_due_date", ""),
             tr_format_currency(request.form.get("estimated_value", "")), request.form.get("status", "In Progress"),
             request.form.get("assigned_to", ""), request.form.get("notes", ""), now, now)
        )
        new_project_id = cur.lastrowid
        log_activity("tracker", "project", new_project_id, "created", asset_id=new_project_id, new_value=request.form["name"])
        db.commit()
        return redirect(url_for("tracker_view_project", project_id=new_project_id))
    return render_template("tracker/new_project.html", status_options=TR_STATUS_OPTIONS)


@app.route("/tracker/project/<int:project_id>")
@login_required
def tracker_view_project(project_id):
    db = get_db()
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    quotes = db.execute(
        "SELECT * FROM tracker_quotes WHERE project_id = ? ORDER BY is_submit_blocking DESC, trade ASC, vendor_name ASC",
        (project_id,)
    ).fetchall()
    docs = db.execute("SELECT * FROM tracker_docs WHERE project_id = ? ORDER BY created_at ASC", (project_id,)).fetchall()

    trade_groups = []
    seen_trades = {}
    for q in quotes:
        key = q["trade"]
        if key not in seen_trades:
            seen_trades[key] = {"trade": key, "quotes": [], "any_blocking": False}
            trade_groups.append(seen_trades[key])
        seen_trades[key]["quotes"].append(q)
        if q["is_submit_blocking"] and q["status"] != "Received":
            seen_trades[key]["any_blocking"] = True

    # Deterministic Back-to-dashboard navigation: carries through whichever
    # dashboard filter the user came from (e.g. ?filter=active), rather than
    # relying only on the browser's back button, so "Back" always lands on
    # the same filtered view the user was looking at -- not just wherever
    # browser history happens to point.
    back_filter = request.args.get("filter", "")
    return render_template("tracker/project.html", p=project, trade_groups=trade_groups, docs=docs,
                            status_options=TR_STATUS_OPTIONS, back_filter=back_filter)


@app.route("/tracker/project/<int:project_id>/update", methods=["POST"])
@login_required
def tracker_update_project(project_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    old_project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (project_id,)).fetchone()
    new_status = request.form["status"]
    new_value = tr_format_currency(request.form.get("estimated_value", ""))
    db.execute(
        "UPDATE tracker_projects SET status = ?, estimated_value = ?, updated_at = ? WHERE id = ?",
        (new_status, new_value, datetime.utcnow().isoformat(), project_id)
    )
    if old_project and old_project["status"] != new_status:
        log_activity("tracker", "project", project_id, "updated", asset_id=project_id,
                     field="status", old_value=old_project["status"], new_value=new_status)
    if old_project and old_project["estimated_value"] != new_value:
        log_activity("tracker", "project", project_id, "updated", asset_id=project_id,
                     field="estimated_value", old_value=old_project["estimated_value"], new_value=new_value)
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=project_id))


@app.route("/tracker/project/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
def tracker_edit_project(project_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    if request.method == "POST":
        new_fields = {
            "name": request.form["name"], "client": request.form.get("client", ""),
            "address": request.form.get("address", ""), "bid_due_date": request.form.get("bid_due_date", ""),
            "estimated_value": tr_format_currency(request.form.get("estimated_value", "")),
            "status": request.form.get("status", project["status"]),
            "assigned_to": request.form.get("assigned_to", ""), "notes": request.form.get("notes", ""),
        }
        db.execute(
            """UPDATE tracker_projects SET name = ?, client = ?, address = ?, bid_due_date = ?, estimated_value = ?,
               status = ?, assigned_to = ?, notes = ?, updated_at = ? WHERE id = ?""",
            (new_fields["name"], new_fields["client"], new_fields["address"], new_fields["bid_due_date"],
             new_fields["estimated_value"], new_fields["status"], new_fields["assigned_to"], new_fields["notes"],
             datetime.utcnow().isoformat(), project_id)
        )
        for field_name, new_val in new_fields.items():
            old_val = project[field_name]
            if old_val != new_val:
                log_activity("tracker", "project", project_id, "updated", asset_id=project_id,
                             field=field_name, old_value=old_val, new_value=new_val)
        db.commit()
        return redirect(url_for("tracker_view_project", project_id=project_id))
    return render_template("tracker/edit_project.html", p=project, status_options=TR_STATUS_OPTIONS)


@app.route("/tracker/project/<int:project_id>/quote/new", methods=["GET", "POST"])
@login_required
def tracker_new_quote(project_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    if request.method == "POST":
        now = datetime.utcnow().isoformat()
        cur = db.execute(
            """INSERT INTO tracker_quotes (project_id, trade, vendor_name, vendor_contact, vendor_email,
               vendor_phone, rfq_sent_date, status, is_submit_blocking, notes, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (project_id, request.form["trade"], request.form.get("vendor_name", ""),
             request.form.get("vendor_contact", ""), request.form.get("vendor_email", ""),
             tr_format_phone(request.form.get("vendor_phone", "")), request.form.get("rfq_sent_date", ""),
             request.form.get("status", "Not Sent"), 1 if request.form.get("is_submit_blocking") else 0,
             request.form.get("notes", ""), now, now)
        )
        log_activity("tracker", "quote", cur.lastrowid, "created", asset_id=project_id,
                     new_value=f"{request.form['trade']} - {request.form.get('vendor_name', '')}")
        db.commit()
        return redirect(url_for("tracker_view_project", project_id=project_id))
    return render_template("tracker/new_quote.html", project=project)


@app.route("/tracker/quote/<int:quote_id>/update_status", methods=["POST"])
@login_required
def tracker_update_quote_status(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    new_status = request.form["status"]
    new_amount = tr_format_currency(request.form.get("amount", quote["amount"]))
    db.execute("UPDATE tracker_quotes SET status = ?, amount = ?, updated_at = ? WHERE id = ?",
               (new_status, new_amount, datetime.utcnow().isoformat(), quote_id))
    if quote["status"] != new_status:
        log_activity("tracker", "quote", quote_id, "updated", asset_id=quote["project_id"],
                     field="status", old_value=quote["status"], new_value=new_status)
    if quote["amount"] != new_amount:
        log_activity("tracker", "quote", quote_id, "updated", asset_id=quote["project_id"],
                     field="amount", old_value=quote["amount"], new_value=new_amount)
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))


@app.route("/tracker/quote/<int:quote_id>/generate_rfq", methods=["POST"])
@login_required
def tracker_generate_rfq(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (quote["project_id"],)).fetchone()
    result = tr_call_claude(tr_build_rfq_prompt(quote, project))
    db.execute("UPDATE tracker_quotes SET rfq_email = ?, updated_at = ? WHERE id = ?",
               (result, datetime.utcnow().isoformat(), quote_id))
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))


@app.route("/tracker/quote/<int:quote_id>/generate_followup", methods=["POST"])
@login_required
def tracker_generate_followup(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (quote["project_id"],)).fetchone()
    result = tr_call_claude(tr_build_followup_prompt(quote, project))
    db.execute("UPDATE tracker_quotes SET follow_up_email = ?, updated_at = ? WHERE id = ?",
               (result, datetime.utcnow().isoformat(), quote_id))
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))


@app.route("/tracker/quote/<int:quote_id>/clear_rfq", methods=["POST"])
@login_required
def tracker_clear_rfq(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    db.execute("UPDATE tracker_quotes SET rfq_email = NULL WHERE id = ?", (quote_id,))
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))


@app.route("/tracker/quote/<int:quote_id>/clear_followup", methods=["POST"])
@login_required
def tracker_clear_followup(quote_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    quote = db.execute("SELECT * FROM tracker_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        flash("Quote not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    db.execute("UPDATE tracker_quotes SET follow_up_email = NULL WHERE id = ?", (quote_id,))
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=quote["project_id"]))


@app.route("/tracker/project/<int:project_id>/doc/new", methods=["POST"])
@login_required
def tracker_new_doc(project_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    cur = db.execute(
        "INSERT INTO tracker_docs (project_id, doc_name, doc_type, status, notes, link, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (project_id, request.form["doc_name"], request.form.get("doc_type", ""),
         request.form.get("status", "Needed"), request.form.get("notes", ""),
         request.form.get("link", ""), datetime.utcnow().isoformat())
    )
    log_activity("tracker", "doc", cur.lastrowid, "created", asset_id=project_id, new_value=request.form["doc_name"])
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=project_id))


@app.route("/tracker/doc/<int:doc_id>/edit", methods=["GET", "POST"])
@login_required
def tracker_edit_doc(doc_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    doc = db.execute("SELECT * FROM tracker_docs WHERE id = ?", (doc_id,)).fetchone()
    if not doc:
        flash("Document not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (doc["project_id"],)).fetchone()
    if request.method == "POST":
        new_fields = {
            "doc_name": request.form["doc_name"], "doc_type": request.form.get("doc_type", ""),
            "status": request.form.get("status", doc["status"]), "link": request.form.get("link", ""),
            "notes": request.form.get("notes", ""),
        }
        db.execute(
            "UPDATE tracker_docs SET doc_name = ?, doc_type = ?, status = ?, link = ?, notes = ? WHERE id = ?",
            (new_fields["doc_name"], new_fields["doc_type"], new_fields["status"],
             new_fields["link"], new_fields["notes"], doc_id)
        )
        for field_name, new_val in new_fields.items():
            old_val = doc[field_name]
            if old_val != new_val:
                log_activity("tracker", "doc", doc_id, "updated", asset_id=doc["project_id"],
                             field=field_name, old_value=old_val, new_value=new_val)
        db.commit()
        return redirect(url_for("tracker_view_project", project_id=doc["project_id"]))
    return render_template("tracker/edit_doc.html", d=doc, project=project)


@app.route("/tracker/doc/<int:doc_id>/delete", methods=["POST"])
@login_required
def tracker_delete_doc(doc_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    doc = db.execute("SELECT * FROM tracker_docs WHERE id = ?", (doc_id,)).fetchone()
    if not doc:
        flash("Document not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    project_id = doc["project_id"]
    log_activity("tracker", "doc", doc_id, "deleted", asset_id=project_id, old_value=doc["doc_name"])
    db.execute("DELETE FROM tracker_docs WHERE id = ?", (doc_id,))
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=project_id))


@app.route("/tracker/doc/<int:doc_id>/update", methods=["POST"])
@login_required
def tracker_update_doc(doc_id):
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    doc = db.execute("SELECT * FROM tracker_docs WHERE id = ?", (doc_id,)).fetchone()
    new_status = request.form["status"]
    db.execute("UPDATE tracker_docs SET status = ? WHERE id = ?", (new_status, doc_id))
    if doc and doc["status"] != new_status:
        log_activity("tracker", "doc", doc_id, "updated", asset_id=doc["project_id"],
                     field="status", old_value=doc["status"], new_value=new_status)
    db.commit()
    return redirect(url_for("tracker_view_project", project_id=doc["project_id"]))


@app.route("/tracker/unit-prices")
@login_required
def tracker_unit_prices():
    db = get_db()
    prices = db.execute("SELECT * FROM tracker_unit_prices ORDER BY category ASC, item ASC").fetchall()
    return render_template("tracker/unit_prices.html", prices=prices)


@app.route("/tracker/unit-prices/new", methods=["POST"])
@login_required
def tracker_new_unit_price():
    if not _authorized("action:project_hunt:manage"):
        flash("You don't have permission to make changes in Project Hunt.", "error")
        return redirect(url_for("tracker_dashboard"))
    db = get_db()
    db.execute(
        "INSERT INTO tracker_unit_prices (category, item, unit, price, notes, updated_at) VALUES (?, ?, ?, ?, ?, ?)",
        (request.form.get("category", ""), request.form["item"], request.form.get("unit", ""),
         tr_format_currency(request.form.get("price", "")), request.form.get("notes", ""), datetime.utcnow().isoformat())
    )
    db.commit()
    return redirect(url_for("tracker_unit_prices"))


@app.route("/tracker/activity-log")
@login_required
def tracker_activity_log():
    db = get_db()
    entries = db.execute(
        """SELECT a.*, p.name AS project_name FROM activity_log a
           LEFT JOIN tracker_projects p ON a.asset_id = p.id
           WHERE a.section = 'tracker' ORDER BY a.created_at DESC LIMIT 300"""
    ).fetchall()
    return render_template("tracker/activity_log.html", entries=entries)


@app.route("/tracker/project/<int:project_id>/activity")
@login_required
def tracker_project_activity_log(project_id):
    db = get_db()
    project = db.execute("SELECT * FROM tracker_projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("tracker_dashboard"))
    entries = db.execute(
        "SELECT * FROM activity_log WHERE section = 'tracker' AND asset_id = ? ORDER BY created_at DESC",
        (project_id,)
    ).fetchall()
    return render_template("tracker/activity_log.html", entries=entries, project=project)


@app.route("/admin/backup")
@login_required
def admin_backup():
    if not _authorized("action:system_data:manage"):
        flash("Only admins can download the full database backup.", "error")
        return redirect(url_for("home"))
    if not os.path.exists(DB_PATH):
        flash("No database file found.", "error")
        return redirect(url_for("home"))
    backup_name = f"buildiq_backup_{datetime.now().strftime('%Y-%m-%d_%H%M')}.db"
    return send_file(DB_PATH, as_attachment=True, download_name=backup_name)


@app.route("/admin/restore", methods=["GET", "POST"])
@login_required
def admin_restore():
    if not _authorized("action:system_data:manage"):
        flash("Only admins can restore the database from a backup.", "error")
        return redirect(url_for("home"))
    if request.method == "POST":
        uploaded_file = request.files.get("backup_file")
        if not uploaded_file or uploaded_file.filename == "":
            flash("No file selected.", "error")
            return redirect(url_for("admin_restore"))
        temp_path = DB_PATH + ".upload_tmp"
        uploaded_file.save(temp_path)
        try:
            test_conn = sqlite3.connect(temp_path)
            test_conn.execute("SELECT COUNT(*) FROM sitepulse_assets").fetchone()
            test_conn.close()
        except Exception:
            os.remove(temp_path)
            flash("That file doesn't look like a valid BuildIQ backup. Nothing was changed.", "error")
            return redirect(url_for("admin_restore"))
        close_db(None)
        os.replace(temp_path, DB_PATH)
        # Restoring swaps in a DB file that may predate newer tables/columns
        # (departments, etc.) -- re-run migrations immediately so the
        # restored DB is brought up to the current schema before anyone
        # hits a route that assumes it.
        init_db()
        flash("Database restored successfully from backup.")
        return redirect(url_for("home"))
    return render_template("admin_restore.html")


@app.route("/admin/export/excel")
@login_required
def admin_export_excel():
    if not _authorized("action:system_data:manage"):
        flash("Only admins can export company data.", "error")
        return redirect(url_for("home"))
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    db = get_db()
    wb = Workbook()
    NAVY = "0A1420"
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=NAVY)
    title_font = Font(name="Arial", size=16, bold=True, color=NAVY)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    def style_title(ws, subtitle):
        ws["A1"] = f"BuildIQ — {subtitle} Export"
        ws["A1"].font = title_font
        ws["A2"] = f"Exported {datetime.now().strftime('%B %d, %Y')}"
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color="6B7280")

    def style_headers(ws, headers, row=4):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=i + 1, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

    ws = wb.active
    ws.title = "Equipment"
    style_title(ws, "Equipment")
    style_headers(ws, ["Name", "Description", "Status", "Location", "Value", "Daily Rate"])
    row = 5
    for a in db.execute("SELECT * FROM sitepulse_assets ORDER BY name"):
        for col, val in enumerate([a["name"], a["description"] or "", a["status"], a["location"] or "",
                                    a["value"] or "", a["daily_rate"] or ""], start=1):
            ws.cell(row=row, column=col, value=val).border = border
        row += 1

    ws2 = wb.create_sheet("Rentals")
    style_title(ws2, "Rentals")
    style_headers(ws2, ["Equipment", "Vendor", "Rented", "Due Back", "Returned"])
    row = 5
    for r in db.execute("SELECT * FROM sitepulse_rentals ORDER BY due_date"):
        for col, val in enumerate([r["equipment_description"], r["vendor"], r["rented_date"],
                                    r["due_date"] or "", r["returned_date"] or ""], start=1):
            ws2.cell(row=row, column=col, value=val).border = border
        row += 1

    ws3 = wb.create_sheet("SitePulse")
    style_title(ws3, "SitePulse")
    style_headers(ws3, ["Item", "Site", "Quantity", "Unit", "Shelf/Location"])
    row = 5
    for m in db.execute("SELECT * FROM inventory_materials ORDER BY site, item_name"):
        for col, val in enumerate([m["item_name"], m["site"], m["quantity"] or "", m["unit"] or "",
                                    m["shelf_location"] or ""], start=1):
            ws3.cell(row=row, column=col, value=val).border = border
        row += 1

    ws4 = wb.create_sheet("Project Hunt")
    style_title(ws4, "Project Hunt")
    style_headers(ws4, ["Project", "Client", "Status", "Due Date", "Value"])
    row = 5
    for p in db.execute("SELECT * FROM tracker_projects ORDER BY bid_due_date"):
        for col, val in enumerate([p["name"], p["client"] or "", p["status"], p["bid_due_date"] or "",
                                    p["estimated_value"] or ""], start=1):
            ws4.cell(row=row, column=col, value=val).border = border
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"buildiq_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/admin/import", methods=["GET", "POST"])
@login_required
def admin_import():
    if not _authorized("action:system_data:manage"):
        flash("Not authorized.", "error")
        return redirect(url_for("home"))

    if request.method == "POST":
        db = get_db()
        results = []

        sp_file = request.files.get("sitepulse_backup")
        if sp_file and sp_file.filename:
            tmp_path = "/tmp/import_sitepulse.db"
            sp_file.save(tmp_path)
            try:
                src = sqlite3.connect(tmp_path)
                src.row_factory = sqlite3.Row
                now = datetime.utcnow().isoformat()

                asset_id_map = {}
                for a in src.execute("SELECT * FROM assets"):
                    cur = db.execute(
                        """INSERT INTO sitepulse_assets (name, description, year, serial_number, value,
                           daily_rate, weekly_rate, monthly_rate, status, location, hours_mileage,
                           created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (a["name"], a["description"], a["year"], a["serial_number"], a["value"],
                         a["daily_rate"], a["weekly_rate"], a["monthly_rate"], a["status"],
                         a["location"], a["hours_mileage"], a["created_at"] or now, a["updated_at"] or now))
                    asset_id_map[a["id"]] = cur.lastrowid
                results.append(f"{len(asset_id_map)} equipment assets")

                usage_count = 0
                for u in src.execute("SELECT * FROM usage_log"):
                    new_asset_id = asset_id_map.get(u["asset_id"])
                    if new_asset_id:
                        db.execute(
                            """INSERT INTO sitepulse_usage_log (asset_id, usage_type, job_name, job_address,
                               client, out_date, return_date, notes, created_at) VALUES (?,?,?,?,?,?,?,?,?)""",
                            (new_asset_id, u["usage_type"], u["job_name"], u["job_address"], u["client"],
                             u["out_date"], u["return_date"], u["notes"], u["created_at"] or now))
                        usage_count += 1
                results.append(f"{usage_count} usage log entries")

                maint_count = 0
                for m in src.execute("SELECT * FROM maintenance_log"):
                    new_asset_id = asset_id_map.get(m["asset_id"])
                    if new_asset_id:
                        db.execute(
                            """INSERT INTO sitepulse_maintenance_log (asset_id, entry_date, work_done, parts,
                               hours_at_service, reported_by, resolved, created_at) VALUES (?,?,?,?,?,?,?,?)""",
                            (new_asset_id, m["entry_date"], m["work_done"], m["parts"], m["hours_at_service"],
                             m["reported_by"], m["resolved"], m["created_at"] or now))
                        maint_count += 1
                results.append(f"{maint_count} maintenance log entries")

                rental_count = 0
                for r in src.execute("SELECT * FROM rentals"):
                    db.execute(
                        """INSERT INTO sitepulse_rentals (vendor, equipment_description, job_name, rate_amount,
                           rate_period, rented_date, due_date, returned_date, notes, created_at, updated_at)
                           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                        (r["vendor"], r["equipment_description"], r["job_name"], r["rate_amount"],
                         r["rate_period"], r["rented_date"], r["due_date"], r["returned_date"], r["notes"],
                         r["created_at"] or now, r["updated_at"] or now))
                    rental_count += 1
                results.append(f"{rental_count} rentals")

                mat_count = 0
                try:
                    for mtl in src.execute("SELECT * FROM materials"):
                        db.execute(
                            """INSERT INTO inventory_materials (item_name, site, quantity, unit, shelf_location,
                               notes, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?)""",
                            (mtl["item_name"], mtl["site"], mtl["quantity"], mtl["unit"], mtl["shelf_location"],
                             mtl["notes"], mtl["created_at"] or now, mtl["updated_at"] or now))
                        mat_count += 1
                except sqlite3.OperationalError:
                    pass
                results.append(f"{mat_count} inventory materials")

                cr_count = 0
                try:
                    for c in src.execute("SELECT * FROM concrete_requests"):
                        db.execute(
                            """INSERT INTO inventory_concrete_requests (project, job_site_address,
                               area_description, pour_date, pour_time, mix_design_psi, mix_slump,
                               concrete_amount, truck_spacing, pump_size, pump_arrival_time, lab_required,
                               lab_time, drilling_required, drilling_time, requested_by, requested_signature,
                               requested_date, ordered_by, ordered_signature, ordered_date, created_at, updated_at)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (c["project"], c["job_site_address"], c["area_description"], c["pour_date"],
                             c["pour_time"], c["mix_design_psi"], c["mix_slump"], c["concrete_amount"],
                             c["truck_spacing"], c["pump_size"], c["pump_arrival_time"], c["lab_required"],
                             c["lab_time"], c["drilling_required"], c["drilling_time"], c["requested_by"],
                             c["requested_signature"], c["requested_date"], c["ordered_by"],
                             c["ordered_signature"], c["ordered_date"], c["created_at"] or now, c["updated_at"] or now))
                        cr_count += 1
                except sqlite3.OperationalError:
                    pass
                results.append(f"{cr_count} concrete requests")

                src.close()
                os.remove(tmp_path)
            except Exception as e:
                flash(f"SitePulse import failed: {e}", "error")
                return redirect(url_for("admin_import"))

        cc_file = request.files.get("tracker_backup")
        if cc_file and cc_file.filename:
            tmp_path = "/tmp/import_tracker.db"
            cc_file.save(tmp_path)
            try:
                src = sqlite3.connect(tmp_path)
                src.row_factory = sqlite3.Row
                now = datetime.utcnow().isoformat()

                project_id_map = {}
                for p in src.execute("SELECT * FROM projects"):
                    cur = db.execute(
                        """INSERT INTO tracker_projects (name, client, address, bid_due_date, estimated_value,
                           status, assigned_to, notes, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (p["name"], p["client"], p["address"], p["bid_due_date"], p["estimated_value"],
                         p["status"], p["assigned_to"], p["notes"], p["created_at"] or now, p["updated_at"] or now))
                    project_id_map[p["id"]] = cur.lastrowid
                results.append(f"{len(project_id_map)} bid tracker projects")

                quote_count = 0
                for q in src.execute("SELECT * FROM quotes"):
                    new_project_id = project_id_map.get(q["project_id"])
                    if new_project_id:
                        db.execute(
                            """INSERT INTO tracker_quotes (project_id, trade, vendor_name, vendor_contact,
                               vendor_email, vendor_phone, rfq_sent_date, status, amount, notes,
                               created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                            (new_project_id, q["trade"], q["vendor_name"], q["vendor_contact"], q["vendor_email"],
                             q["vendor_phone"], q["rfq_sent_date"], q["status"], q["amount"], q["notes"],
                             q["created_at"] or now, q["updated_at"] or now))
                        quote_count += 1
                results.append(f"{quote_count} vendor quotes")

                src.close()
                os.remove(tmp_path)
            except Exception as e:
                flash(f"Bid Tracker import failed: {e}", "error")
                return redirect(url_for("admin_import"))

        db.commit()
        flash("Imported: " + ", ".join(results) if results else "No files uploaded.")
        return redirect(url_for("admin_import"))

    return render_template("admin_import.html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 3000))
    app.run(host="0.0.0.0", port=port, debug=False)
